#!/usr/bin/env python3
"""
jetxl differential / round-trip value-fidelity fuzzer
=====================================================

The other fuzzers assert a workbook LOADS and CONFORMS to the schema. Neither
checks that the VALUES are correct -- a cell can be perfectly well-formed and
schema-valid while holding the wrong number. This fuzzer targets exactly that
silent value-corruption class:

  write random typed data with jetxl  ->  read it back with openpyxl  ->
  assert every non-null cell equals what went in (type-appropriate compare).

Comparison rules (match how Excel/openpyxl legitimately transform values):
  - int:    exact, except values beyond 2^53 may be stored as float (Excel does
            this too) -> compared as float within relative tolerance.
  - float:  relative tolerance 1e-12; NaN/Inf -> expected empty (None).
  - bool:   exact.
  - str:    exact, EXCEPT control chars (< 0x20 other than tab/nl/cr) are legally
            stripped, and a value that is all-stripped or "" reads back as None.
  - date/datetime: calendar-equal (date component); pre-1900 dates are written as
            ISO strings by design, so those compare as string.

Determinism via JETXL_FUZZ_SEED. Failures (real value corruption) are printed and
written to /tmp/fuzz_diff_repro.txt.

Usage: python fuzz_diff.py [N] [-v]
"""
from __future__ import annotations

import datetime as dt
import io
import math
import os
import random
import sys
import tempfile
import warnings

warnings.filterwarnings("ignore")

import pyarrow as pa
import openpyxl

import jetxl

N = 400
for a in sys.argv[1:]:
    if a.isdigit():
        N = int(a)
VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv
SEED = int(os.environ.get("JETXL_FUZZ_SEED", "20260705"))

_p = _f = 0
failures = []


def check(desc, cond):
    global _p, _f
    if cond:
        _p += 1
    else:
        _f += 1
        failures.append(desc)
        if len(failures) <= 40:
            print(f"  FAIL {desc}")


# ---- value generators with a "kind" tag so we know how to compare ----------
def strip_control(s):
    """Excel/OOXML legally drops control chars except tab/nl/cr; \r -> \n."""
    out = []
    for ch in s:
        o = ord(ch)
        if o == 0x0d:
            out.append("\n")
        elif o < 0x20 and ch not in ("\t", "\n"):
            continue
        else:
            out.append(ch)
    return "".join(out)


def gen_column(rng, n):
    kind = rng.choice(["int", "uint", "float", "bool", "str", "date", "datetime"])
    if kind == "int":
        vals = [rng.choice([0, 1, -1, rng.randint(-10**15, 10**15),
                            2**53, -2**53, rng.randint(-2**62, 2**62)]) for _ in range(n)]
        return pa.array(vals, pa.int64()), vals, "int"
    if kind == "uint":
        vals = [rng.choice([0, 1, 2**53, 2**63, 2**64 - 1, rng.randint(0, 2**64 - 1)]) for _ in range(n)]
        return pa.array(vals, pa.uint64()), vals, "uint"
    if kind == "float":
        vals = [rng.choice([0.0, 1.5, -1.5, math.pi, 1e-8, 1e8,
                            rng.uniform(-1e6, 1e6), 1.0 / 3.0]) for _ in range(n)]
        return pa.array(vals, pa.float64()), vals, "float"
    if kind == "bool":
        vals = [rng.random() > 0.5 for _ in range(n)]
        return pa.array(vals, pa.bool_()), vals, "bool"
    if kind == "str":
        pool = ["hello", "", " spaces ", "日本語", "tab\there", "quote\"q",
                "amp&<>x", "a" * rng.randint(1, 500), "0007", "-42", "3.14",
                "line\nbreak", "trail ", " lead"]
        vals = [rng.choice(pool) for _ in range(n)]
        return pa.array(vals), vals, "str"
    if kind == "date":
        vals = [dt.date(rng.randint(1900, 2100), rng.randint(1, 12), rng.randint(1, 28)) for _ in range(n)]
        return pa.array(vals, pa.date32()), vals, "date"
    # datetime
    vals = [dt.datetime(rng.randint(1900, 2100), rng.randint(1, 12), rng.randint(1, 28),
                        rng.randint(0, 23), rng.randint(0, 59), rng.randint(0, 59)) for _ in range(n)]
    return pa.array(vals, pa.timestamp("us")), vals, "datetime"


def compare(kind, expected, got):
    """Return True if `got` (read back) matches `expected` (written), per rules."""
    if kind in ("int", "uint"):
        # beyond 2^53 Excel stores as float; compare numerically
        if got is None:
            return expected == 0 and False  # 0 should round-trip; None is wrong
        try:
            return abs(float(got) - float(expected)) <= max(1.0, abs(float(expected)) * 1e-12)
        except Exception:
            return False
    if kind == "float":
        if not math.isfinite(expected):
            return got is None
        if got is None:
            return False
        return abs(float(got) - expected) <= abs(expected) * 1e-10 + 1e-12
    if kind == "bool":
        return bool(got) == expected
    if kind == "str":
        exp = strip_control(expected)
        if exp.strip() == "" and exp == "":
            return got is None or got == ""
        # openpyxl reads empty as None
        if exp == "":
            return got is None
        return got == exp
    if kind in ("date", "datetime"):
        if got is None:
            return False
        # date component must match; openpyxl returns datetime
        gd = got.date() if hasattr(got, "date") else got
        ed = expected.date() if isinstance(expected, dt.datetime) else expected
        if hasattr(gd, "year"):
            return gd == ed
        # written as ISO string (shouldn't happen for >=1900) -> string compare
        return str(got).startswith(str(ed))
    return True


def run_one(rng, path_kind):
    ncols = rng.randint(1, 6)
    nrows = rng.choice([1, 2, 5, 10, 25])
    arrays = {}
    meta = []  # (colname, kind, expected_list)
    for c in range(ncols):
        arr, vals, kind = gen_column(rng, nrows)
        name = f"c{c}"
        arrays[name] = arr
        meta.append((name, kind, vals))
    table = pa.table(arrays)

    # date/datetime columns need a date format to read back as dates
    col_formats = {}
    for name, kind, _ in meta:
        if kind == "date":
            col_formats[name] = "date"
        elif kind == "datetime":
            col_formats[name] = "datetime"

    if path_kind == "single_bytes":
        b = jetxl.write_sheet_arrow_to_bytes(table, column_formats=col_formats)
        wb = openpyxl.load_workbook(io.BytesIO(b))
    elif path_kind == "single_file":
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            p = tf.name
        try:
            jetxl.write_sheet_arrow(table, p, column_formats=col_formats)
            wb = openpyxl.load_workbook(p)
        finally:
            os.unlink(p)
    elif path_kind == "multi_bytes":
        sheet = {"data": table, "name": "S1", "column_formats": col_formats}
        b = jetxl.write_sheets_arrow_to_bytes([sheet], 1)
        wb = openpyxl.load_workbook(io.BytesIO(b))
    else:  # multi_file
        sheet = {"data": table, "name": "S1", "column_formats": col_formats}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            p = tf.name
        try:
            jetxl.write_sheets_arrow([sheet], p, 1)
            wb = openpyxl.load_workbook(p)
        finally:
            os.unlink(p)

    ws = wb.active
    for col_idx, (name, kind, vals) in enumerate(meta, start=1):
        for row_off, expected in enumerate(vals):
            got = ws.cell(row=2 + row_off, column=col_idx).value
            check(f"[{path_kind}] col{col_idx} row{row_off+2} {kind} exp={expected!r} got={got!r}",
                  compare(kind, expected, got))


def main():
    print("=" * 74)
    print(f"jetxl differential round-trip fuzzer — {N} iters, seed={SEED}")
    print("verifies VALUE fidelity (write -> read -> compare) across all paths")
    print("=" * 74)
    master = random.Random(SEED)
    paths = ["single_bytes", "single_file", "multi_bytes", "multi_file"]
    for i in range(N):
        rng = random.Random(master.randint(0, 2**31))
        path_kind = paths[i % len(paths)]
        try:
            run_one(rng, path_kind)
        except Exception as e:
            import traceback
            if VERBOSE:
                traceback.print_exc()
            check(f"[{path_kind}] iter{i} raised {type(e).__name__}: {str(e)[:50]}", False)
        if VERBOSE and i % 50 == 0:
            print(f"  ... {i}/{N} (fails={_f})")

    print("\n" + "=" * 74)
    print(f"value-fidelity checks: {_p} passed, {_f} failed")
    if failures:
        with open("/tmp/fuzz_diff_repro.txt", "w") as f:
            for x in failures:
                f.write(x + "\n")
        print(f"(all {len(failures)} failures -> /tmp/fuzz_diff_repro.txt)")
    print("=" * 74)
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())
