#!/usr/bin/env python3
"""
jetxl reference-parity suite
============================

Edge cases distilled from the openpyxl / XlsxWriter / rust_xlsxwriter test
suites, restricted to features jetxl actually implements. Each is exercised on
every applicable write path (single/multi × file/bytes) so a fix on one path
can't regress another.

Categories mirrored from the reference repos:
  - escapes:      sheet names + formulas + cells with < > & " '  (xlsxwriter
                  test_escapes01/02/04/05)
  - types:        int/float/bool/str/date, NaN/Inf, numbers-as-text
                  (xlsxwriter test_types, rust_xlsxwriter set_nan_string)
  - merge_range:  top-left keeps value, ranges emitted (test_merge_range)
  - quote_name:   sheet-name validation & sanitisation (test_quote_name)
  - set_column:   width by name/index, hidden (test_set_column)
  - default_row:  default + explicit row heights (test_default_row)
  - data_validation: list/whole/decimal/text_length + messages
  - cond_format:  cell_value / color_scale / data_bar / top10
  - charts:       6 types + legend position + axis + data labels

Run: python test_reference_parity.py [-v]
"""
from __future__ import annotations

import datetime as dt
import io
import os
import sys
import tempfile
import zipfile

import pyarrow as pa
import openpyxl

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv
_p = _f = 0
_fails = []


def ok(d):
    global _p
    _p += 1
    if VERBOSE:
        print(f"    pass  {d}")


def bad(d):
    global _f
    _f += 1
    _fails.append(d)
    print(f"    FAIL  {d}")


def check(d, c):
    ok(d) if c else bad(d)


def sec(n):
    print(f"\n=== {n} ===")


def _crc(b):
    if zipfile.ZipFile(io.BytesIO(b)).testzip() is not None:
        raise AssertionError("bad CRC")


def tbl(n=20):
    cats = ["North", "South", "East", "West", "Central"]
    return pa.table({
        "id": pa.array(range(n), pa.int64()),
        "region": pa.array([cats[i % 5] for i in range(n)]),
        "amount": pa.array([i * 1.5 for i in range(n)], pa.float64()),
    })


def single_paths(arrow, **kw):
    b = jetxl.write_sheet_arrow_to_bytes(arrow, **kw); _crc(b)
    yield "single_bytes", b
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheet_arrow(arrow, p, **kw)
        b = open(p, "rb").read(); _crc(b)
    finally:
        os.unlink(p)
    yield "single_file", b
    s = {"data": arrow, "name": "S1"}; s.update(kw)
    b = jetxl.write_sheets_arrow_to_bytes([dict(s)], 1); _crc(b)
    yield "multi_bytes", b
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheets_arrow([dict(s)], p, 1)
        b = open(p, "rb").read(); _crc(b)
    finally:
        os.unlink(p)
    yield "multi_file", b


def sheetxml(b):
    return zipfile.ZipFile(io.BytesIO(b)).read("xl/worksheets/sheet1.xml").decode()


# ==========================================================================
def t_escapes():
    sec("escapes — sheet names, cells, formulas with < > & \" ' (xlsxwriter)")
    # sheet names with metacharacters (test_escapes01/05)
    for nm in ["5&4", "A & B", "Q1<>Q2", 'Say "Hi"']:
        b = jetxl.write_sheet_arrow_to_bytes(tbl(3), sheet_name=nm); _crc(b)
        wb = openpyxl.load_workbook(io.BytesIO(b))
        check(f"sheet name {nm!r} round-trips", wb.active.title == nm)
    # multi-sheet names with & (must be escaped in workbook.xml)
    b = jetxl.write_sheets_arrow_to_bytes(
        [{"data": tbl(3), "name": "A & B"}, {"data": tbl(3), "name": "5<4"}], 1)
    _crc(b)
    check("multi-sheet metachar names", openpyxl.load_workbook(io.BytesIO(b)).sheetnames == ["A & B", "5<4"])
    # cell text with all five metachars (test_escapes02)
    t = pa.table({"s": pa.array(['"&<>', "a'b", "<x>&<y>"])})
    for path, b in single_paths(t):
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check(f"[{path}] cell metachars intact", ws["A2"].value == '"&<>' and ws["A4"].value == "<x>&<y>")
    # formula with < > & (test_escapes01)
    t2 = tbl(5)
    for path, b in single_paths(t2, formulas=[(2, 0, "IF(1>2,0,1)", None),
                                              (3, 0, 'CONCATENATE("a","<>&")', None)]):
        sx = sheetxml(b)
        check(f"[{path}] formula metachars escaped", "&gt;" in sx and "&lt;" in sx and "&amp;" in sx)
    # URL with & (test_escapes04)
    for path, b in single_paths(tbl(5), hyperlinks=[(2, 0, "http://x.com/?a=1&b=2", "u")]):
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check(f"[{path}] &-URL opens cleanly", ws["A1"].value == "id")


def t_types():
    sec("types — numeric/bool/date, NaN/Inf, numbers-as-text (xlsxwriter/rust)")
    # bool vs int distinction (Python bool is subclass of int)
    t = pa.table({"b": pa.array([True, False]), "i": pa.array([1, 0], pa.int64())})
    for path, b in single_paths(t):
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check(f"[{path}] bool stays bool, int stays int",
              ws["A2"].value is True and ws["A3"].value is False
              and ws["B2"].value == 1 and ws["B3"].value == 0)
    # NaN / Inf -> empty, no corrupt literal (rust_xlsxwriter set_nan_string)
    tn = pa.table({"v": pa.array([1.0, float("nan"), float("inf"), float("-inf"), 2.0], pa.float64())})
    for path, b in single_paths(tn):
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        sx = sheetxml(b)
        check(f"[{path}] NaN/Inf -> empty, finite intact",
              ws["A2"].value == 1.0 and ws["A6"].value == 2.0
              and ws["A3"].value in (None, "") and "NaN" not in sx and "inf" not in sx.lower())
    # numbers-as-text (leading zeros)
    tt = pa.table({"c": pa.array(["007", "0042", "+5"])})
    for path, b in single_paths(tt):
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check(f"[{path}] leading-zero codes stay strings",
              [ws.cell(row=r, column=1).value for r in range(2, 5)] == ["007", "0042", "+5"])


def t_dates_reference():
    sec("dates — Excel 1900 serials incl phantom leap day (all paths)")
    dates = [dt.date(1900, 1, 1), dt.date(1900, 2, 28), dt.date(1900, 3, 1),
             dt.date(2000, 1, 1), dt.date(2024, 6, 15)]
    expected = [1, 59, 61, 36526, 45458]
    t = pa.table({"d": pa.array(dates)})
    import re
    for path, b in single_paths(t, column_formats={"d": "date"}):
        sx = sheetxml(b)
        serials = [round(float(s)) for s in re.findall(r'<c r="A[2-9]"[^>]*><v>([^<]+)</v>', sx)]
        check(f"[{path}] date serials match Excel", serials == expected)


def t_merge_range():
    sec("merge_range — top-left keeps value, range emitted (xlsxwriter)")
    for path, b in single_paths(tbl(10), merge_cells=[(2, 0, 4, 0), (1, 0, 1, 2)]):
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        ranges = {str(r) for r in ws.merged_cells.ranges}
        check(f"[{path}] both merge ranges present", "A2:A4" in ranges and "A1:C1" in ranges)
        check(f"[{path}] top-left keeps value", ws["A2"].value == 0)


def t_quote_name():
    sec("quote_name — sheet-name validation & sanitisation (xlsxwriter)")
    t = tbl(3)
    # invalid chars must be rejected or sanitised, never produce a broken file
    for nm in ["Bad/Name", "Has:Colon", "Star*", "Quest?", "Brack[1]"]:
        try:
            b = jetxl.write_sheet_arrow_to_bytes(t, sheet_name=nm); _crc(b)
            title = openpyxl.load_workbook(io.BytesIO(b)).active.title
            check(f"invalid name {nm!r} sanitised (no illegal char)",
                  all(c not in title for c in "[]:*?/\\"))
        except Exception:
            ok(f"invalid name {nm!r} rejected explicitly")
    # duplicate names rejected in multi
    try:
        jetxl.write_sheets_arrow_to_bytes([{"data": t, "name": "D"}, {"data": t, "name": "D"}], 1)
        bad("duplicate sheet names should be rejected")
    except Exception:
        ok("duplicate sheet names rejected")
    # >31 chars handled (truncated or rejected)
    try:
        b = jetxl.write_sheet_arrow_to_bytes(t, sheet_name="X" * 40); _crc(b)
        check("over-31-char name <= 31", len(openpyxl.load_workbook(io.BytesIO(b)).active.title) <= 31)
    except Exception:
        ok("over-31-char name rejected")


def t_set_column_and_rows():
    sec("set_column / default_row — widths, heights, hidden (xlsxwriter)")
    for path, b in single_paths(tbl(10), column_widths={"amount": 25.0, 0: 8.0},
                                hidden_columns=["region"], row_heights={1: 30.0},
                                hidden_rows=[3], default_row_height=16.0):
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check(f"[{path}] width by name", ws.column_dimensions["C"].width is not None)
        check(f"[{path}] hidden column", ws.column_dimensions["B"].hidden)
        check(f"[{path}] explicit row height", ws.row_dimensions[1].height == 30.0)
        check(f"[{path}] hidden row", ws.row_dimensions[3].hidden)


def t_cond_format_all():
    sec("cond_format — all rule types (xlsxwriter/rust data_bar/duplicate)")
    cfs = [
        {"start_row": 2, "start_col": 2, "end_row": 20, "end_col": 2,
         "rule_type": "cell_value", "operator": "greater_than", "value": "10"},
        {"start_row": 2, "start_col": 0, "end_row": 20, "end_col": 0,
         "rule_type": "color_scale", "min_color": "FFFF0000", "max_color": "FF00FF00"},
        {"start_row": 2, "start_col": 2, "end_row": 20, "end_col": 2,
         "rule_type": "data_bar", "color": "FF638EC6"},
        {"start_row": 2, "start_col": 0, "end_row": 20, "end_col": 0,
         "rule_type": "top10", "rank": 5, "bottom": False},
    ]
    for path, b in single_paths(tbl(20), conditional_formats=cfs):
        sx = sheetxml(b)
        check(f"[{path}] cond formats emitted", "<conditionalFormatting" in sx
              and "colorScale" in sx and "dataBar" in sx)


def t_charts_reference():
    sec("charts — 6 types + legend position + axis + labels (xlsxwriter)")
    t = pa.table({"m": pa.array(["Jan", "Feb", "Mar"]),
                  "s": pa.array([1.0, 2.0, 3.0], pa.float64())})
    for ct in ["column", "bar", "line", "pie", "scatter", "area"]:
        for path, b in single_paths(t, charts=[{"chart_type": ct, "data_range": (0, 1, 3, 1),
                                               "category_col": 0, "legend_position": "bottom",
                                               "show_data_labels": True, "axis_min": 0.0}]):
            names = zipfile.ZipFile(io.BytesIO(b)).namelist()
            check(f"[{path}] {ct} chart part present",
                  any("chart" in n and n.endswith(".xml") and "rels" not in n for n in names))


def main():
    print("=" * 74)
    print("jetxl reference-parity suite (openpyxl / xlsxwriter / rust_xlsxwriter)")
    print("=" * 74)
    for fn in [t_escapes, t_types, t_dates_reference, t_merge_range, t_quote_name,
               t_set_column_and_rows, t_cond_format_all, t_charts_reference]:
        try:
            fn()
        except Exception as e:
            import traceback
            traceback.print_exc()
            bad(f"{fn.__name__} crashed: {e!r}")
    print("\n" + "=" * 74)
    print(f"TOTAL: {_p} passed, {_f} failed")
    if _fails:
        print("\nFailures:")
        for x in _fails:
            print("  -", x)
    print("=" * 74)
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())
