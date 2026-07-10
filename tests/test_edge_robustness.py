#!/usr/bin/env python3
"""
jetxl edge-case robustness suite
================================

Locks in fixes for corruption / wrong-output bugs found by systematic probing
of the input surface. Every case here previously either corrupted the file
(unopenable), silently produced wrong output, or discarded a rule. Each is now
exercised on the write paths and validated by loading with openpyxl in
warnings-as-errors mode, so any "rule discarded" regression fails the test.

Covered:
  - uint64 values above i64::MAX (were written as negative via overflow)
  - formulas passed with a leading '=' (were doubled to '==')
  - reversed ranges in conditional formats / merges / data validations
    (were emitted as invalid "B20:B2" and discarded by Excel/openpyxl)
  - invalid color strings in fonts/fills/colorScale/dataBar
    (made the whole workbook unreadable)
  - out-of-range text_rotation (made the whole workbook unreadable)
"""
from __future__ import annotations

import io
import sys
import warnings

import pyarrow as pa
import openpyxl

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv
_p = _f = 0
_fails = []


def ok(d):
    global _p
    _p += 1
    if VERBOSE:
        print(f"    pass  {d}")


def bad(d):
    global _f
    _f += 1
    _fails.append(d)
    print(f"    FAIL  {d}")


def check(d, c):
    ok(d) if c else bad(d)


def sec(n):
    print(f"\n=== {n} ===")


def tbl(n=5):
    return pa.table({"a": pa.array(list(range(1, n + 1)), pa.int64()),
                     "b": pa.array([float(i * 10) for i in range(1, n + 1)], pa.float64())})


def load_strict(b):
    """Load with openpyxl, treating its 'rule discarded' warnings as errors."""
    with warnings.catch_warnings():
        warnings.simplefilter("error", UserWarning)
        return openpyxl.load_workbook(io.BytesIO(b))


def all_arrow_paths(data, **kw):
    """Yield (label, bytes) for the four arrow paths."""
    import os
    import tempfile
    yield "single_bytes", jetxl.write_sheet_arrow_to_bytes(data, **kw)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheet_arrow(data, p, **kw)
        yield "single_file", open(p, "rb").read()
    finally:
        os.unlink(p)
    sheet = {"data": data, "name": "S1"}
    sheet.update(kw)
    yield "multi_bytes", jetxl.write_sheets_arrow_to_bytes([dict(sheet)], 1)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheets_arrow([dict(sheet)], p, 1)
        yield "multi_file", open(p, "rb").read()
    finally:
        os.unlink(p)


# ==========================================================================
def t_uint64():
    sec("uint64 above i64::MAX writes correctly (no overflow to negative)")
    cases = [
        (18446744073709551615, "uint64 max"),
        (10000000000000000000, "uint64 big"),
        (9223372036854775808, "just above i64 max"),
        (12345, "small in-range"),
    ]
    for val, name in cases:
        t = pa.table({"c": pa.array([val], pa.uint64())})
        for path, b in all_arrow_paths(t):
            ws = load_strict(b).active
            v = ws.cell(row=2, column=1).value
            # Excel stores as f64; allow float representation, must be positive & ~equal
            good = v is not None and v > 0 and abs(float(v) - val) <= max(1.0, val * 1e-15)
            check(f"[{name}/{path}] value {v} not overflowed", good)


def t_formula_equals():
    sec("Formula with leading '=' is not doubled to '=='")
    for f_in in ["=A1+B1", "A1+B1", "=SUM(A1:A5)", "IF(A1>0,1,0)"]:
        for path, b in all_arrow_paths(tbl(), formulas=[(2, 0, f_in, "0")]):
            v = load_strict(b).active.cell(row=2, column=1).value
            check(f"[{f_in!r}/{path}] single leading '='",
                  isinstance(v, str) and v.startswith("=") and not v.startswith("=="))


def t_reversed_ranges():
    sec("Reversed ranges normalized (CF / merge / data validation)")
    # conditional format reversed
    for path, b in all_arrow_paths(tbl(), conditional_formats=[
            {"start_row": 20, "start_col": 1, "end_row": 2, "end_col": 1,
             "rule_type": "data_bar", "color": "FF638EC6"}]):
        try:
            load_strict(b)  # would warn+raise if the CF rule was discarded
            ok(f"[cf/{path}] reversed CF range loads without discard")
        except UserWarning:
            bad(f"[cf/{path}] reversed CF range discarded")
    # merge reversed
    for path, b in all_arrow_paths(tbl(), merge_cells=[(20, 0, 2, 0)]):
        ws = load_strict(b).active
        check(f"[merge/{path}] reversed merge normalized",
              "A2:A20" in {str(r) for r in ws.merged_cells.ranges})
    # data validation reversed
    for path, b in all_arrow_paths(tbl(), data_validations=[
            {"start_row": 20, "start_col": 0, "end_row": 2, "end_col": 0,
             "type": "whole_number", "min": 0, "max": 100}]):
        try:
            load_strict(b)
            ok(f"[dv/{path}] reversed DV range loads")
        except UserWarning:
            bad(f"[dv/{path}] reversed DV range discarded")


def t_invalid_colors():
    sec("Invalid colors don't corrupt the workbook (font/fill/CF)")
    color_inputs = ["red", "#FF0000", "F00", "FF0000", "FFFF0000", "#FFFF0000", "zzz", "12345", ""]
    for c in color_inputs:
        for path, b in all_arrow_paths(tbl(), cell_styles=[
                {"row": 2, "col": 0, "font": {"color": c}}]):
            try:
                load_strict(b)
                ok(f"[font-color {c!r}/{path}] loads, no corruption")
            except Exception as e:
                bad(f"[font-color {c!r}/{path}] corrupted: {type(e).__name__}")
    # fills
    for c in ["#00FF00", "00FF00", "green", "badcolor"]:
        for path, b in all_arrow_paths(tbl(), cell_styles=[
                {"row": 2, "col": 0, "fill": {"pattern": "solid", "fg_color": c}}]):
            try:
                load_strict(b)
                ok(f"[fill {c!r}/{path}] loads")
            except Exception:
                bad(f"[fill {c!r}/{path}] corrupted")
    # colorScale + dataBar with bad colors
    for path, b in all_arrow_paths(tbl(), conditional_formats=[
            {"start_row": 2, "start_col": 0, "end_row": 5, "end_col": 0,
             "rule_type": "color_scale", "min_color": "bad", "max_color": "#00FF00"}]):
        try:
            load_strict(b); ok(f"[colorScale-badcolor/{path}] loads")
        except Exception:
            bad(f"[colorScale-badcolor/{path}] corrupted")
    for path, b in all_arrow_paths(tbl(), conditional_formats=[
            {"start_row": 2, "start_col": 0, "end_row": 5, "end_col": 0,
             "rule_type": "data_bar", "color": "notacolor"}]):
        try:
            load_strict(b); ok(f"[dataBar-badcolor/{path}] loads")
        except Exception:
            bad(f"[dataBar-badcolor/{path}] corrupted")


def t_text_rotation():
    sec("Out-of-range text_rotation doesn't corrupt the workbook")
    for r in [0, 45, 90, -45, -90, 180, 200, 255, -200, 1000]:
        for path, b in all_arrow_paths(tbl(), cell_styles=[
                {"row": 2, "col": 0, "alignment": {"text_rotation": r}}]):
            try:
                ws = load_strict(b).active
                tr = ws.cell(row=2, column=1).alignment.textRotation
                good = tr is None or (0 <= int(tr) <= 180) or int(tr) == 255
                check(f"[rotation {r}/{path}] valid encoding ({tr})", good)
            except Exception as e:
                bad(f"[rotation {r}/{path}] corrupted: {type(e).__name__}")


def t_table_col_out_of_range_no_panic():
    sec("Table end_col beyond schema width must NOT panic (panic=abort => crash)")
    # tbl() has 2 columns (indices 0,1). Ranges past that used to panic on a
    # slice index, which with panic="abort" aborts the whole process.
    two_col = tbl()
    cases = [
        {"name": "T", "start_row": 0, "start_col": 0, "end_row": 5, "end_col": 2},
        {"name": "T", "start_row": 0, "start_col": 0, "end_row": 5, "end_col": 99},
        {"name": "", "start_row": 0, "start_col": 0, "end_row": 5, "end_col": 2},
        {"name": "T", "start_row": 0, "start_col": 5, "end_row": 5, "end_col": 7},
    ]
    for tc in cases:
        for path, b in all_arrow_paths(two_col, tables=[dict(tc)]):
            try:
                load_strict(b)
                ok(f"[end_col={tc['end_col']} start_col={tc['start_col']}/{path}] no panic, loads")
            except Exception as e:
                bad(f"[end_col={tc['end_col']}/{path}] {type(e).__name__}: {str(e)[:40]}")


def t_duplicate_table_names():
    sec("Duplicate table names are made unique (no corruption)")
    for path, b in all_arrow_paths(tbl(), tables=[
            {"name": "D", "start_row": 0, "start_col": 0, "end_row": 5, "end_col": 0},
            {"name": "D", "start_row": 0, "start_col": 1, "end_row": 5, "end_col": 1}]):
        try:
            load_strict(b)
            ok(f"[{path}] duplicate table names load without corruption")
        except Exception as e:
            bad(f"[{path}] dup table names corrupted: {type(e).__name__}")


def t_chart_color_font_robust():
    sec("Chart colors/font sizes never corrupt the workbook")
    base = pa.table({"m": pa.array([f"M{i}" for i in range(8)]),
                     "s": pa.array([float(i) for i in range(8)], pa.float64())})
    variants = [
        {"chart_type": "column", "data_range": (0, 1, 8, 1), "category_col": 0, "title": "X", "title_color": "red"},
        {"chart_type": "column", "data_range": (0, 1, 8, 1), "category_col": 0, "title": "X", "title_color": "#FF0000"},
        {"chart_type": "column", "data_range": (0, 1, 8, 1), "category_col": 0, "title": "X", "title_color": "FF0000"},
        {"chart_type": "column", "data_range": (0, 1, 8, 1), "category_col": 0, "title": "X", "title_font_size": 0},
        {"chart_type": "column", "data_range": (0, 1, 8, 1), "category_col": 0, "title": "X", "title_font_size": 999999},
        {"chart_type": "column", "data_range": (0, 1, 8, 1), "category_col": 0, "y_axis_title": "Y", "axis_title_color": "nothex"},
    ]
    for i, ch in enumerate(variants):
        for path, b in all_arrow_paths(base, charts=[ch]):
            try:
                load_strict(b)
                ok(f"[chart-variant{i}/{path}] loads, no corruption")
            except Exception as e:
                bad(f"[chart-variant{i}/{path}] corrupted: {type(e).__name__}")


def t_hyperlink_row_zero():
    sec("Hyperlink on row 0 doesn't corrupt the workbook")
    for path, b in all_arrow_paths(tbl(), hyperlinks=[(0, 0, "http://example.com", "t")]):
        try:
            load_strict(b)
            ok(f"[{path}] hyperlink row 0 loads (clamped)")
        except Exception as e:
            bad(f"[{path}] hyperlink row 0 corrupted: {type(e).__name__}")


def t_null_type_columns():
    sec("Null-type (all-None) Arrow columns write as empty cells")
    t = pa.table({"a": pa.array([1, 2, 3], pa.int64()),
                  "x": pa.array([None, None, None], pa.null()),
                  "b": pa.array(["p", "q", "r"])})
    for path, b in all_arrow_paths(t):
        try:
            ws = load_strict(b).active
            avals = [ws.cell(row=r, column=1).value for r in range(2, 5)]
            xvals = [ws.cell(row=r, column=2).value for r in range(2, 5)]
            bvals = [ws.cell(row=r, column=3).value for r in range(2, 5)]
            check(f"[{path}] null-type col empty, neighbors intact",
                  avals == [1, 2, 3] and xvals == [None, None, None] and bvals == ["p", "q", "r"])
        except Exception as e:
            bad(f"[{path}] null-type column: {type(e).__name__}: {str(e)[:40]}")
    # pandas all-None column (arrives as Arrow null type)
    import pandas as _pd
    arrow = pa.Table.from_pandas(_pd.DataFrame({"allnone": [None, None], "v": [1, 2]}))
    for path, b in all_arrow_paths(arrow):
        try:
            load_strict(b)
            ok(f"[pandas-allnone/{path}] loads")
        except Exception as e:
            bad(f"[pandas-allnone/{path}] {type(e).__name__}")


def t_hyperlink_in_merge():
    sec("Hyperlink inside a merged range relocates to anchor (no corruption)")
    data = pa.table({"a": pa.array(list(range(1, 9)), pa.int64()),
                     "b": pa.array(list("abcdefgh")),
                     "c": pa.array(list(range(10, 90, 10)), pa.int64())})
    configs = [
        # hyperlink on a non-anchor cell of one merge
        {"merge_cells": [(2, 0, 4, 1)], "hyperlinks": [(3, 0, "https://x.com", "l")]},
        # overlapping merges + hyperlinks inside them
        {"merge_cells": [(4, 0, 8, 2), (1, 1, 5, 0), (5, 3, 14, 4)],
         "hyperlinks": [(7, 0, "#Sheet1!A1", "l"), (10, 3, "https://x.com", "l")]},
    ]
    for i, cfg in enumerate(configs):
        for path, b in all_arrow_paths(data, **cfg):
            try:
                load_strict(b)
                ok(f"[merge-hl-cfg{i}/{path}] loads (hyperlink relocated)")
            except Exception as e:
                bad(f"[merge-hl-cfg{i}/{path}] corrupted: {type(e).__name__}: {str(e)[:35]}")


def t_dict_nan_inf():
    sec("Dict path: NaN/Inf write as empty cells (no corruption)")
    import tempfile
    import os as _os
    p = tempfile.mktemp(suffix=".xlsx")
    try:
        jetxl.write_sheet({"a": [float("nan"), float("inf"), float("-inf"), 1.5],
                           "b": [1, 2, 3, 4]}, p)
        with warnings.catch_warnings():
            warnings.simplefilter("error", UserWarning)
            wb = openpyxl.load_workbook(p)
        ws = wb.active
        avals = [ws.cell(row=r, column=1).value for r in range(2, 6)]
        bvals = [ws.cell(row=r, column=2).value for r in range(2, 6)]
        check("dict NaN/Inf -> empty, finite kept, neighbor intact",
              avals[3] == 1.5 and avals[:3] == [None, None, None] and bvals == [1, 2, 3, 4])
    except Exception as e:
        bad(f"dict NaN/Inf corrupted: {type(e).__name__}: {str(e)[:40]}")
    finally:
        if _os.path.exists(p):
            _os.unlink(p)


def t_pre_1900_dates():
    sec("Pre-1900 dates write as ISO strings (Excel can't serialize them)")
    import datetime as _dt
    import tempfile
    import os as _os
    t = pa.table({"d": pa.array([_dt.date(1, 1, 1), _dt.date(1850, 6, 15),
                                 _dt.date(1900, 1, 1), _dt.date(2024, 1, 1)], pa.date32())})
    for path, b in all_arrow_paths(t, column_formats={"d": "date"}):
        try:
            ws = load_strict(b).active
            vals = [ws.cell(row=r, column=1).value for r in range(2, 6)]
            good = (isinstance(vals[0], str) and isinstance(vals[1], str)
                    and hasattr(vals[2], "year") and hasattr(vals[3], "year"))
            check(f"[{path}] pre-1900 as string, 1900+ as date", good)
        except Exception as e:
            bad(f"[{path}] pre-1900 dates: {type(e).__name__}: {str(e)[:40]}")
    p = tempfile.mktemp(suffix=".xlsx")
    try:
        jetxl.write_sheet({"d": [_dt.datetime(1, 1, 1), _dt.datetime(1850, 1, 1),
                                 _dt.datetime(2024, 1, 1)]}, p)
        with warnings.catch_warnings():
            warnings.simplefilter("error", UserWarning)
            openpyxl.load_workbook(p)
        ok("[dict] pre-1900 dates load without serial error")
    except Exception as e:
        bad(f"[dict] pre-1900 dates: {type(e).__name__}: {str(e)[:40]}")
    finally:
        if _os.path.exists(p):
            _os.unlink(p)


def main():
    print("=" * 74)
    print("jetxl edge-case robustness suite")
    print("=" * 74)
    for fn in [t_uint64, t_formula_equals, t_reversed_ranges, t_invalid_colors, t_text_rotation,
               t_table_col_out_of_range_no_panic, t_duplicate_table_names,
               t_chart_color_font_robust, t_hyperlink_row_zero,
               t_null_type_columns, t_hyperlink_in_merge,
               t_dict_nan_inf, t_pre_1900_dates]:
        try:
            fn()
        except Exception as e:
            import traceback
            traceback.print_exc()
            bad(f"{fn.__name__} crashed: {e!r}")
    print("\n" + "=" * 74)
    print(f"TOTAL: {_p} passed, {_f} failed")
    if _fails:
        print("\nFailures:")
        for x in _fails:
            print("  -", x)
    print("=" * 74)
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())
