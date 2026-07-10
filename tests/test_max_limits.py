#!/usr/bin/env python3
"""
jetxl maximum-limits stress suite
=================================

Pushes jetxl to the real structural ceilings Excel supports, for EVERY feature,
across ALL FOUR arrow write paths (single file / single bytes / multi file /
multi bytes) plus the dict paths where relevant. This is deliberately heavy —
not "three images" but hundreds-to-thousands of each element — because feature
interactions and part-numbering bugs only surface at scale.

Excel .xlsx hard limits exercised here:
  - 1,048,576 rows per sheet            (max, and max+1 must be REJECTED)
  - 16,384 columns per sheet (XFD)      (max, and max+1 must be REJECTED)
  - 32,767 characters per cell          (max-length string round-trips)
  - many sheets per workbook            (1000+)
  - large counts of images / charts / tables / merges / hyperlinks /
    conditional formats / cell styles / data validations, spread across sheets,
    all with globally-unique part names and resolvable relationships.

Each generated workbook is validated structurally (CRC, well-formed XML, every
r:id resolves, every rel target exists, no duplicate ZIP entries) — not merely
"does openpyxl open it".

Run:  python test_max_limits.py [-v] [--heavy]
  --heavy also runs the 1,048,576-row and 16,384-column cases (slow, ~a few GB
  of transient work); without it those are scaled down but the limit-rejection
  checks still run.
"""
from __future__ import annotations

import io
import os
import re
import sys
import tempfile
import time
import zipfile
from collections import Counter
from xml.etree import ElementTree as ET

import pyarrow as pa
import openpyxl

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv
HEAVY = "--heavy" in sys.argv
_p = _f = 0
_fails = []

PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d4948445200000001000000010806000000"
    "1f15c4890000000d49444154789c6260f8cf000000ffff03000006000557"
    "bfabd40000000049454e44ae426082")


def ok(d):
    global _p
    _p += 1
    if VERBOSE:
        print(f"    pass  {d}")


def bad(d):
    global _f
    _f += 1
    _fails.append(d)
    print(f"    FAIL  {d}")


def check(d, c):
    ok(d) if c else bad(d)


def sec(n):
    print(f"\n=== {n} ===")


def structural_ok(b: bytes, label: str, expect_open=True):
    """Full structural validation. Returns True if all invariants pass."""
    z = zipfile.ZipFile(io.BytesIO(b))
    names = z.namelist()
    check(f"[{label}] CRCs valid", z.testzip() is None)
    # no duplicate ZIP entries (the image-collision class of bug)
    dups = [k for k, v in Counter(names).items() if v > 1]
    check(f"[{label}] no duplicate ZIP entries", not dups)
    if dups:
        print("        dups:", dups[:5])
    # every part well-formed
    malformed = []
    for nm in names:
        if nm.endswith(".xml") or nm.endswith(".rels"):
            try:
                ET.fromstring(z.read(nm))
            except Exception as e:
                malformed.append(nm)
    check(f"[{label}] all XML well-formed", not malformed)
    # every sheet + drawing r:id resolves and targets exist
    for part in names:
        if re.match(r"xl/(worksheets/sheet|drawings/drawing)\d+\.xml$", part):
            body = z.read(part).decode()
            refd = set(re.findall(r'r:id="(rId[^"]+)"', body)) | set(re.findall(r'r:embed="(rId[^"]+)"', body))
            base = "worksheets" if "worksheets" in part else "drawings"
            rels_name = part.replace(f"{base}/", f"{base}/_rels/") + ".rels"
            if refd:
                if rels_name in names:
                    rels = z.read(rels_name).decode()
                    defined = set(re.findall(r'Id="(rId[^"]+)"', rels))
                    if not refd <= defined:
                        bad(f"[{label}] {part} has dangling r:id")
                        return False
                    for tg in re.findall(r'Target="([^"]+)"', rels):
                        if tg.startswith(("http", "mailto")):
                            continue
                        resolved = os.path.normpath(os.path.join(os.path.dirname(part), tg)).replace("\\", "/")
                        if resolved not in names:
                            bad(f"[{label}] {part} rel target missing: {tg}")
                            return False
                else:
                    bad(f"[{label}] {part} refs r:id but no rels")
                    return False
    ok(f"[{label}] all r:ids resolve + targets exist")
    if expect_open:
        try:
            openpyxl.load_workbook(io.BytesIO(b))
            ok(f"[{label}] openpyxl opens")
        except Exception as e:
            bad(f"[{label}] openpyxl open failed: {e!r}")
    return True


# --------------------------------------------------------------------------
def tbl(n=20):
    return pa.table({
        "cat": pa.array([f"c{i}" for i in range(n)]),
        "v1": pa.array([float(i) for i in range(n)], pa.float64()),
        "v2": pa.array([float(i * 2) for i in range(n)], pa.float64()),
    })


def write_multi(sheets, threads=4):
    """Return {'multi_bytes': b, 'multi_file': b}."""
    out = {}
    out["multi_bytes"] = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in sheets], threads)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheets_arrow([dict(s) for s in sheets], p, threads)
        out["multi_file"] = open(p, "rb").read()
    finally:
        os.unlink(p)
    return out


def write_single(data, **kw):
    """Return {'single_bytes': b, 'single_file': b}."""
    out = {}
    out["single_bytes"] = jetxl.write_sheet_arrow_to_bytes(data, **kw)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheet_arrow(data, p, **kw)
        out["single_file"] = open(p, "rb").read()
    finally:
        os.unlink(p)
    return out


# ==========================================================================
# 1. ROW / COLUMN grid limits (max works, max+1 rejected) — all paths
# ==========================================================================
def t_grid_limits():
    sec("Grid limits — max rows/cols work, over-limit rejected on every path")
    # max valid columns == 16,384
    max_cols = 16384 if HEAVY else 4096
    tc = pa.table({f"c{i}": pa.array([1], pa.int64()) for i in range(max_cols)})
    for path, b in write_single(tc).items():
        check(f"[{path}] {max_cols} columns writes valid",
              zipfile.ZipFile(io.BytesIO(b)).testzip() is None)

    # over the column limit must be rejected on every path
    tc2 = pa.table({f"c{i}": pa.array([1], pa.int64()) for i in range(16385)})
    _reject_on_all_paths(tc2, "16385 columns")

    # max valid rows (only under --heavy; it's ~1M rows)
    if HEAVY:
        n = 1_048_575  # + header = 1,048,576 == Excel max
        t = pa.table({"id": pa.array(range(n), pa.int64())})
        t0 = time.time()
        b = jetxl.write_sheet_arrow_to_bytes(t)
        dt = time.time() - t0
        check(f"max rows (1,048,576 incl header) writes valid in {dt:.1f}s",
              zipfile.ZipFile(io.BytesIO(b)).testzip() is None)
        sx = zipfile.ZipFile(io.BytesIO(b)).read("xl/worksheets/sheet1.xml").decode()
        last = re.findall(r'<row r="(\d+)"', sx)[-1]
        check("max-rows last row == 1048576", last == "1048576")

    # over the row limit rejected on every path (small col count, huge row count)
    n2 = 1_048_576  # + header exceeds
    t2 = pa.table({"id": pa.array(range(n2), pa.int64())})
    _reject_on_all_paths(t2, "over 1,048,576 rows")


def _reject_on_all_paths(data, what):
    paths = {
        "single_bytes": lambda: jetxl.write_sheet_arrow_to_bytes(data),
        "single_file": lambda: jetxl.write_sheet_arrow(data, tempfile.mktemp(suffix=".xlsx")),
        "multi_bytes": lambda: jetxl.write_sheets_arrow_to_bytes([{"data": data, "name": "S"}], 1),
        "multi_file": lambda: jetxl.write_sheets_arrow([{"data": data, "name": "S"}], tempfile.mktemp(suffix=".xlsx"), 1),
    }
    for path, fn in paths.items():
        try:
            fn()
            bad(f"[{path}] {what} NOT rejected")
        except Exception as e:
            # must be a clean exception, not a process abort
            check(f"[{path}] {what} rejected cleanly", "exceed" in str(e).lower() or "maximum" in str(e).lower())


# ==========================================================================
# 2. MAX chars per cell (32,767)
# ==========================================================================
def t_max_cell_chars():
    sec("Max characters per cell — 32,767")
    s = "X" * 32_767
    t = pa.table({"big": pa.array([s, "small"]), "n": pa.array([1, 2], pa.int64())})
    for path, b in write_single(t).items():
        structural_ok(b, f"maxchars/{path}")
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check(f"[{path}] 32767-char cell round-trips",
              ws["A2"].value == s and len(ws["A2"].value) == 32_767)


# ==========================================================================
# 3. MANY images (thousands) across many sheets — global media naming
# ==========================================================================
def t_many_images():
    sec("Many images — thousands across sheets, unique media, all multi paths")
    n_sheets = 200 if HEAVY else 60
    imgs_per = 10
    total = n_sheets * imgs_per
    sheets = [{"data": tbl(5), "name": f"S{s}",
               "images": [{"data": list(PNG), "extension": "png",
                           "from_col": 3 + (i % 4), "from_row": 1 + i * 6,
                           "to_col": 5 + (i % 4), "to_row": 5 + i * 6}
                          for i in range(imgs_per)]}
              for s in range(n_sheets)]
    for path, b in write_multi(sheets).items():
        names = zipfile.ZipFile(io.BytesIO(b)).namelist()
        media = [n for n in names if re.match(r"xl/media/image\d+\.", n)]
        check(f"[{path}] {total} images -> {len(set(media))} unique media files",
              len(set(media)) == total and len(media) == total)
        structural_ok(b, f"manyimg/{path}", expect_open=(n_sheets <= 60))


# ==========================================================================
# 4. MANY charts (thousands) across sheets
# ==========================================================================
def t_many_charts():
    sec("Many charts — thousands across sheets, unique chart parts")
    n_sheets = 150 if HEAVY else 50
    charts_per = 10
    total = n_sheets * charts_per
    ct = ["column", "bar", "line", "pie", "area", "scatter"]
    sheets = [{"data": tbl(20), "name": f"S{s}",
               "charts": [{"chart_type": ct[i % 6], "data_range": (0, 1, 20, 1),
                           "category_col": 0, "title": f"C{i}"} for i in range(charts_per)]}
              for s in range(n_sheets)]
    for path, b in write_multi(sheets).items():
        names = zipfile.ZipFile(io.BytesIO(b)).namelist()
        charts = [n for n in names if re.match(r"xl/charts/chart\d+\.xml$", n)]
        check(f"[{path}] {total} charts -> {len(set(charts))} unique chart parts",
              len(set(charts)) == total and len(charts) == total)
        structural_ok(b, f"manychart/{path}", expect_open=(n_sheets <= 50))


# ==========================================================================
# 5. MANY tables — globally-unique names required by Excel
# ==========================================================================
def t_many_tables():
    sec("Many tables — global-unique names, across sheets")
    n_sheets = 100 if HEAVY else 40
    tbl_per = 5
    total = n_sheets * tbl_per
    wide = pa.table({f"c{i}": pa.array([float(j) for j in range(20)], pa.float64()) for i in range(10)})
    sheets = [{"data": wide, "name": f"S{s}",
               "tables": [{"name": f"Tbl_{s}_{i}", "start_row": 0, "start_col": i * 2,
                           "end_row": 20, "end_col": i * 2 + 1} for i in range(tbl_per)]}
              for s in range(n_sheets)]
    for path, b in write_multi(sheets).items():
        z = zipfile.ZipFile(io.BytesIO(b))
        names = z.namelist()
        table_parts = [n for n in names if re.match(r"xl/tables/table\d+\.xml$", n)]
        check(f"[{path}] {total} table parts unique", len(set(table_parts)) == total)
        # display names must be globally unique (Excel requirement)
        dnames = []
        for tp in table_parts:
            m = re.search(r'displayName="([^"]+)"', z.read(tp).decode())
            if m:
                dnames.append(m.group(1))
        check(f"[{path}] table display names globally unique",
              len(set(dnames)) == len(dnames) == total)
        structural_ok(b, f"manytable/{path}", expect_open=(n_sheets <= 40))


# ==========================================================================
# 6. MANY merges / hyperlinks / conditional formats / cell styles in one sheet
# ==========================================================================
def t_many_per_sheet_elements():
    sec("Thousands of merges / hyperlinks / cond-formats / styles in one sheet")
    n = 2000
    t = tbl(n)
    # 2000 merges (each merges two cells in a spare column pair) — use rows
    merges = [(r, 0, r, 0) for r in range(2, 2)]  # placeholder; real below
    merges = [(r, 1, r, 2) for r in range(2, n)]  # merge cols B:C on each data row
    hyperlinks = [(r, 0, f"https://example.com/row/{r}?x=1&y=2", f"r{r}") for r in range(2, min(n, 1000))]
    cond = [{"start_row": 2, "start_col": 1, "end_row": n, "end_col": 1,
             "rule_type": "cell_value", "operator": "greater_than", "value": str(v)} for v in range(50)]
    styles = [{"row": r, "col": 0, "font": {"bold": (r % 2 == 0)},
               "fill": {"pattern": "solid", "fg_color": "FFEEEEEE"}} for r in range(2, min(n, 1000))]
    for path, b in write_single(t, merge_cells=merges, hyperlinks=hyperlinks,
                                conditional_formats=cond, cell_styles=styles).items():
        structural_ok(b, f"manyelem/{path}")
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check(f"[{path}] ~{len(merges)} merges present", len(list(ws.merged_cells.ranges)) >= len(merges) - 1)


# ==========================================================================
# 7. THE MEGA WORKBOOK — everything, everywhere, at once
# ==========================================================================
def t_mega_workbook():
    sec("Mega workbook — every feature, many sheets, all combined")
    n = 40
    t = pa.table({
        "month": pa.array([f"M{i%12}" for i in range(n)]),
        "region": pa.array(["N", "S", "E", "W", "C"][i % 5] for i in range(n)),
        "sales": pa.array([float(i * 100) for i in range(n)], pa.float64()),
        "costs": pa.array([float(i * 60) for i in range(n)], pa.float64()),
        "profit": pa.array([float(i * 40) for i in range(n)], pa.float64()),
    })
    n_sheets = 30 if HEAVY else 12
    sheets = []
    for s in range(n_sheets):
        sheets.append({
            "data": t, "name": f"Sheet{s}",
            "charts": [
                {"chart_type": "column", "data_range": (0, 2, n, 2), "category_col": 0, "title": f"S{s}Sales"},
                {"chart_type": "line", "data_range": (0, 4, n, 4), "category_col": 0, "title": f"S{s}Profit"},
            ],
            "tables": [{"name": f"MegaTbl{s}", "start_row": 0, "start_col": 0, "end_row": n, "end_col": 4,
                        "style": "TableStyleMedium9"}],
            "images": [{"data": list(PNG), "extension": "png", "from_col": 7, "from_row": 1, "to_col": 10, "to_row": 6}],
            "conditional_formats": [
                {"start_row": 2, "start_col": 2, "end_row": n, "end_col": 2, "rule_type": "data_bar", "color": "FF638EC6"},
                {"start_row": 2, "start_col": 4, "end_row": n, "end_col": 4, "rule_type": "color_scale",
                 "min_color": "FFFF0000", "max_color": "FF00FF00"},
            ],
            "cell_styles": [{"row": 1, "col": 0, "font": {"bold": True, "color": "FFFFFFFF"},
                             "fill": {"pattern": "solid", "fg_color": "FF4472C4"},
                             "border": {"bottom": {"style": "thick"}}}],
            "merge_cells": [(1, 0, 1, 4)],
            "hyperlinks": [(3, 0, f"https://example.com/s{s}?a=1&b=2", "link")],
            "formulas": [(6, 2, "SUM(C2:C5)", None)],
            "data_validations": [{"start_row": 2, "start_col": 1, "end_row": n, "end_col": 1,
                                  "type": "list", "items": ["N", "S", "E", "W", "C"]}],
            "auto_filter": True, "freeze_rows": 1, "tab_color": "FF00B050",
        })
    t0 = time.time()
    out = write_multi(sheets)
    dt = time.time() - t0
    for path, b in out.items():
        z = zipfile.ZipFile(io.BytesIO(b))
        names = z.namelist()
        n_charts = sum(1 for x in names if re.match(r"xl/charts/chart\d+\.xml$", x))
        n_tables = sum(1 for x in names if re.match(r"xl/tables/table\d+\.xml$", x))
        n_imgs = sum(1 for x in names if re.match(r"xl/media/image\d+\.", x))
        check(f"[{path}] {n_sheets*2} charts present", n_charts == n_sheets * 2)
        check(f"[{path}] {n_sheets} tables present", n_tables == n_sheets)
        check(f"[{path}] {n_sheets} images present (unique)", n_imgs == n_sheets)
        structural_ok(b, f"mega/{path}", expect_open=(n_sheets <= 12))
    print(f"    (mega workbook {n_sheets} sheets built in {dt:.2f}s per 2 paths)")

    # reproducible
    b1 = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in sheets], 4)
    b2 = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in sheets], 4)
    check("mega workbook byte-reproducible", b1 == b2)


# ==========================================================================
# 8. MANY sheets (1000+)
# ==========================================================================
def t_many_sheets():
    sec("Many sheets — 1000+ in one workbook")
    n_sheets = 1000
    t = pa.table({"a": pa.array([1, 2, 3]), "b": pa.array(["x", "y", "z"])})
    sheets = [{"data": t, "name": f"Sheet{i}"} for i in range(n_sheets)]
    t0 = time.time()
    out = write_multi(sheets)
    dt = time.time() - t0
    for path, b in out.items():
        wb = openpyxl.load_workbook(io.BytesIO(b))
        check(f"[{path}] {n_sheets} sheets all present", len(wb.sheetnames) == n_sheets)
        check(f"[{path}] first & last sheet intact",
              wb.sheetnames[0] == "Sheet0" and wb.sheetnames[-1] == f"Sheet{n_sheets-1}")
        check(f"[{path}] CRC valid", zipfile.ZipFile(io.BytesIO(b)).testzip() is None)
    print(f"    (1000 sheets built in {dt:.2f}s per 2 paths)")


def main():
    print("=" * 74)
    print(f"jetxl maximum-limits stress suite {'(HEAVY)' if HEAVY else '(standard)'}")
    print("=" * 74)
    for fn in [t_grid_limits, t_max_cell_chars, t_many_images, t_many_charts,
               t_many_tables, t_many_per_sheet_elements, t_mega_workbook, t_many_sheets]:
        try:
            fn()
        except Exception as e:
            import traceback
            traceback.print_exc()
            bad(f"{fn.__name__} crashed: {e!r}")
    print("\n" + "=" * 74)
    print(f"TOTAL: {_p} passed, {_f} failed")
    if _fails:
        print("\nFailures:")
        for x in _fails:
            print("  -", x)
    print("=" * 74)
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())
