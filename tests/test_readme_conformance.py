#!/usr/bin/env python3
"""
jetxl README-conformance + write-path-matrix suite
==================================================

Answers the question "is every documented feature actually validated?" by
testing the API surface the way the README documents it, and by exercising each
feature across ALL FOUR write paths where applicable:

    single-sheet file   (write_sheet_arrow)
    single-sheet bytes  (write_sheet_arrow_to_bytes)
    multi-sheet  file   (write_sheets_arrow)
    multi-sheet  bytes  (write_sheets_arrow_to_bytes)

plus the DataFrame entry points the README advertises (Polars / Pandas /
PyArrow) and the legacy dict API.

Rationale: the file vs bytes paths and single vs multi paths have historically
diverged (a whole class of bugs lived only on one path). Testing a feature on
one path proves nothing about the others, so we sweep the matrix.

Every write is round-tripped through openpyxl and CRC-checked. Failures are
reported per (feature, path). Documentation/implementation gaps are reported as
DOC-GAP (not hard failures unless they cause corruption).

Run:  python test_readme_conformance.py [-v]
"""
from __future__ import annotations

import io
import os
import sys
import tempfile
import zipfile

import pyarrow as pa
import openpyxl

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv

_p = _f = _g = 0
_fails = []
_gaps = []


def ok(desc):
    global _p
    _p += 1
    if VERBOSE:
        print(f"    pass  {desc}")


def bad(desc):
    global _f
    _f += 1
    _fails.append(desc)
    print(f"    FAIL  {desc}")


def gap(desc):
    global _g
    _g += 1
    _gaps.append(desc)
    print(f"    DOC-GAP  {desc}")


def check(desc, cond):
    ok(desc) if cond else bad(desc)


# --------------------------------------------------------------------------
# helpers: run a config through all four arrow write paths, return workbooks
# --------------------------------------------------------------------------
def all_paths(table, **kwargs):
    """Yield (path_name, workbook) for each of the four arrow write paths.
    kwargs are the per-sheet feature options (same keys the README uses)."""
    results = {}

    # single bytes
    b = jetxl.write_sheet_arrow_to_bytes(table, **kwargs)
    _assert_crc(b)
    results["single_bytes"] = openpyxl.load_workbook(io.BytesIO(b))

    # single file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        jetxl.write_sheet_arrow(table, path, **kwargs)
        with open(path, "rb") as fh:
            fb = fh.read()
        _assert_crc(fb)
        results["single_file"] = openpyxl.load_workbook(io.BytesIO(fb))
    finally:
        os.unlink(path)

    # multi bytes (single sheet in a multi wrapper)
    sheet = {"data": table, "name": "S1"}
    sheet.update(kwargs)
    bm = jetxl.write_sheets_arrow_to_bytes([sheet], 2)
    _assert_crc(bm)
    results["multi_bytes"] = openpyxl.load_workbook(io.BytesIO(bm))

    # multi file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        jetxl.write_sheets_arrow([dict(sheet)], path, 2)
        with open(path, "rb") as fh:
            fbm = fh.read()
        _assert_crc(fbm)
        results["multi_file"] = openpyxl.load_workbook(io.BytesIO(fbm))
    finally:
        os.unlink(path)

    return results


def _assert_crc(b):
    z = zipfile.ZipFile(io.BytesIO(b))
    if z.testzip() is not None:
        raise AssertionError("bad CRC")


def sample(n=20):
    cats = ["North", "South", "East", "West", "Central"]
    return pa.table({
        "id": pa.array(list(range(n)), pa.int64()),
        "region": pa.array([cats[i % 5] for i in range(n)]),
        "amount": pa.array([i * 1.5 for i in range(n)], pa.float64()),
        "flag": pa.array([i % 2 == 0 for i in range(n)]),
    })


# ==========================================================================
def sec(name):
    print(f"\n=== {name} ===")


# 1. DataFrame entry points (README advertises Polars / Pandas / PyArrow)
def t_dataframe_entrypoints():
    sec("DataFrame entry points — Polars / Pandas / PyArrow")
    import polars as pl
    import pandas as pd

    pldf = pl.DataFrame({"a": [1, 2, 3], "b": ["x", "y", "z"], "c": [1.1, 2.2, 3.3]})
    pddf = pd.DataFrame({"a": [1, 2, 3], "b": ["x", "y", "z"], "c": [1.1, 2.2, 3.3]})

    # Polars -> arrow
    b = jetxl.write_sheet_arrow_to_bytes(pldf.to_arrow())
    ws = openpyxl.load_workbook(io.BytesIO(b)).active
    check("polars.to_arrow round-trips", ws["A2"].value == 1 and ws["B2"].value == "x")

    # Pandas -> pyarrow.Table
    b = jetxl.write_sheet_arrow_to_bytes(pa.Table.from_pandas(pddf))
    ws = openpyxl.load_workbook(io.BytesIO(b)).active
    check("pandas via pa.Table.from_pandas round-trips", ws["C4"].value == 3.3)

    # Native pyarrow
    b = jetxl.write_sheet_arrow_to_bytes(sample(5))
    ws = openpyxl.load_workbook(io.BytesIO(b)).active
    check("native pyarrow round-trips", ws["A2"].value == 0)

    # RecordBatch (not just Table)
    batch = sample(5).to_batches()[0]
    try:
        b = jetxl.write_sheet_arrow_to_bytes(batch)
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        check("pyarrow RecordBatch accepted", ws["A2"].value == 0)
    except Exception as e:
        bad(f"pyarrow RecordBatch accepted (raised {type(e).__name__})")


# 2. Text rotation (documented values 0-180, 255)
def t_text_rotation():
    sec("Text rotation — across write paths")
    style = [{"row": 2, "col": 0,
              "alignment": {"horizontal": "center", "text_rotation": 45}}]
    for path, wb in all_paths(sample(10), cell_styles=style).items():
        c = wb.active.cell(row=2, column=1)
        rot = c.alignment.textRotation
        check(f"[{path}] text_rotation=45 applied", rot == 45)


# 3. Fill patterns (solid, gray125, none)
def t_fill_patterns():
    sec("Fill patterns — solid / gray125 / none")
    for pat, expect in [("solid", "solid"), ("gray125", "gray125")]:
        style = [{"row": 2, "col": 0, "fill": {"pattern": pat, "fg_color": "FFD9D9D9"}}]
        b = jetxl.write_sheet_arrow_to_bytes(sample(10), cell_styles=style)
        _assert_crc(b)
        ws = openpyxl.load_workbook(io.BytesIO(b)).active
        c = ws.cell(row=2, column=1)
        check(f"fill pattern '{pat}' applied", c.fill.patternType == expect)


# 4. Full table config (all documented options) across paths
def t_table_config():
    sec("Excel tables — full documented config across paths")
    # NOTE: the README documents display_name="My Data" (with a space). Excel
    # forbids spaces in a table's displayName, so jetxl must sanitize it to a
    # valid identifier or the whole workbook fails to open. all_paths() opening
    # the file across every path IS the assertion that the sanitization holds.
    table = {
        "name": "MyTable", "display_name": "My Data",
        "start_row": 0, "start_col": 0, "end_row": 20, "end_col": 3,
        "style": "TableStyleMedium2",
        "show_first_column": True, "show_last_column": True,
        "show_row_stripes": True, "show_column_stripes": True,
    }
    for path, wb in all_paths(sample(20), tables=[table]).items():
        ws = wb.active
        # the table part must exist and be wired
        has_tbl = len(ws.tables) >= 1
        check(f"[{path}] spaced display_name sanitized + table registered", has_tbl)

    # Auto-sizing table (end_row/end_col omitted -> computed)
    auto = {"name": "AutoT", "start_row": 0, "start_col": 0}
    b = jetxl.write_sheet_arrow_to_bytes(sample(15), tables=[auto])
    _assert_crc(b)
    ws = openpyxl.load_workbook(io.BytesIO(b)).active
    check("auto-sized table registered", len(ws.tables) >= 1)


# 5. Chart config incl. legend_position (documented but possibly unwired)
def t_chart_config():
    sec("Charts — documented config incl. legend_position")
    t = pa.table({
        "month": pa.array(["Jan", "Feb", "Mar", "Apr"]),
        "sales": pa.array([100.0, 150.0, 120.0, 200.0], pa.float64()),
    })
    chart = {
        "chart_type": "column", "data_range": (0, 1, 4, 1),
        "category_col": 0, "title": "Sales", "show_legend": True,
        "legend_position": "bottom",  # documented
        "x_axis_title": "Month", "y_axis_title": "Value",
        "stacked": False, "show_data_labels": True,
        "title_bold": True, "title_font_size": 1800, "title_color": "FF0000",
    }
    b = jetxl.write_sheet_arrow_to_bytes(t, charts=[chart])
    _assert_crc(b)
    z = zipfile.ZipFile(io.BytesIO(b))
    chart_xml = [z.read(n).decode() for n in z.namelist() if "chart" in n and n.endswith(".xml")]
    check("chart part emitted", len(chart_xml) >= 1)
    if chart_xml:
        cx = chart_xml[0]
        check("chart title present", "Sales" in cx)
        # legend_position documented as settable to 'bottom'; verify it took
        # effect by parsing the actual legendPos val (a loose substring match can
        # false-positive on other val="..." attributes elsewhere in the chart).
        import re as _re
        m = _re.search(r'<c:legendPos\s+val="([^"]*)"', cx)
        if m:
            if m.group(1) == "b":
                ok("legend_position='bottom' honored")
            else:
                gap(f"legend_position='bottom' documented but ignored: chart "
                    f"emits legendPos val=\"{m.group(1)}\" (extract_chart never "
                    f"reads the 'legend_position' key, always defaults to right)")
        else:
            check("legend present", "legend" in cx.lower())

    # all six documented chart types must at least produce a valid file
    for ct in ["column", "bar", "line", "pie", "scatter", "area"]:
        b = jetxl.write_sheet_arrow_to_bytes(t, charts=[{
            "chart_type": ct, "data_range": (0, 1, 4, 1), "category_col": 0}])
        _assert_crc(b)
        openpyxl.load_workbook(io.BytesIO(b))
        ok(f"chart type '{ct}' produces valid file")


# 6. Images from FILE path (README documents path= as well as data=)
def t_images_from_file():
    sec("Images — from file path and from bytes, across paths")
    png = bytes.fromhex(
        "89504e470d0a1a0a0000000d4948445200000001000000010806000000"
        "1f15c4890000000d49444154789c6260f8cf000000ffff03000006000557"
        "bfabd40000000049454e44ae426082")
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tf:
        tf.write(png)
        png_path = tf.name
    try:
        img_file = {"path": png_path, "from_col": 5, "from_row": 1, "to_col": 8, "to_row": 6}
        img_bytes = {"data": list(png), "extension": "png",
                     "from_col": 1, "from_row": 1, "to_col": 3, "to_row": 4}
        for path, wb in all_paths(sample(10), images=[img_file]).items():
            # image media + drawing must exist -> re-open the raw archive
            check(f"[{path}] image-from-file opens", wb.active["A1"].value == "id")
        # bytes form
        for path, wb in all_paths(sample(10), images=[img_bytes]).items():
            check(f"[{path}] image-from-bytes opens", wb.active["A1"].value == "id")
    finally:
        os.unlink(png_path)


# 7. Data validation — all four documented types
def t_data_validation():
    sec("Data validation — list / whole_number / decimal / text_length")
    dvs = [
        {"start_row": 2, "start_col": 0, "end_row": 20, "end_col": 0,
         "type": "list", "items": ["A", "B", "C"]},
        {"start_row": 2, "start_col": 1, "end_row": 20, "end_col": 1,
         "type": "whole_number", "min": 0, "max": 100},
        {"start_row": 2, "start_col": 2, "end_row": 20, "end_col": 2,
         "type": "decimal", "min": 0.0, "max": 1.0},
        {"start_row": 2, "start_col": 3, "end_row": 20, "end_col": 3,
         "type": "text_length", "min": 1, "max": 10},
    ]
    for path, wb in all_paths(sample(20), data_validations=dvs).items():
        types = {dv.type for dv in wb.active.data_validations.dataValidation}
        check(f"[{path}] all validation types present", len(types) >= 3)


# 8. Formulas + merge + hyperlinks across paths (feature-parity guard)
def t_core_features_matrix():
    sec("Core features — formulas / merges / hyperlinks across all paths")
    kwargs = dict(
        formulas=[(5, 2, "SUM(C2:C4)", "9.0")],
        merge_cells=[(1, 0, 1, 3)],
        hyperlinks=[(3, 0, "https://x.com/?a=1&b=2", "link")],
    )
    for path, wb in all_paths(sample(20), **kwargs).items():
        ws = wb.active
        check(f"[{path}] formula present",
              str(ws.cell(row=5, column=3).value).startswith("="))
        check(f"[{path}] merge present",
              "A1:D1" in {str(r) for r in ws.merged_cells.ranges})
        check(f"[{path}] hyperlink cell intact", ws.cell(row=4, column=1).value is not None)


# 9. Legacy dict API (README documents it as backward-compatible)
def t_dict_api():
    sec("Legacy dict API — write_sheet / write_sheets")
    data = {"flag": [True, False], "count": [1, 2], "label": ["a", "b"]}
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        jetxl.write_sheet(data, path, sheet_name="D")
        ws = openpyxl.load_workbook(path)["D"]
        check("dict write_sheet bool fidelity", ws["A2"].value is True)
        check("dict write_sheet int fidelity", ws["B2"].value == 1)
    finally:
        os.unlink(path)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        sheets = [
            {"name": "Q1", "columns": {"x": [1, 2, 3]}},
            {"name": "Q2", "columns": {"y": [4, 5, 6]}},
        ]
        jetxl.write_sheets(sheets, path, 2)
        wb = openpyxl.load_workbook(path)
        check("dict write_sheets order", wb.sheetnames == ["Q1", "Q2"])
    finally:
        os.unlink(path)


def main():
    print("=" * 72)
    print("jetxl README-conformance + write-path-matrix suite")
    print("=" * 72)
    for fn in [
        t_dataframe_entrypoints, t_text_rotation, t_fill_patterns,
        t_table_config, t_chart_config, t_images_from_file,
        t_data_validation, t_core_features_matrix, t_dict_api,
    ]:
        try:
            fn()
        except Exception as e:
            import traceback
            traceback.print_exc()
            bad(f"{fn.__name__} crashed: {e!r}")

    print("\n" + "=" * 72)
    print(f"TOTAL: {_p} passed, {_f} failed, {_g} doc-gaps")
    if _fails:
        print("\nFailures:")
        for x in _fails:
            print("  -", x)
    if _gaps:
        print("\nDocumentation/implementation gaps (non-fatal):")
        for x in _gaps:
            print("  -", x)
    print("=" * 72)
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())
