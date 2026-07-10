#!/usr/bin/env python3
"""
jetxl comprehensive property-based fuzzer (expanded)
====================================================

Extends fuzz_jetxl.py to the surfaces it did not cover:

  A. COMPOSITE / EXOTIC Arrow types (struct, list, large_list, fixed_size_list,
     map, binary, fixed_size_binary, decimal128, duration, time32, interval,
     nested dictionaries). Property: jetxl must NOT panic on any of them — it
     either writes a loadable/conformant workbook, or raises a clean, catchable
     Python exception. (panic="abort" would turn a panic into a process crash.)

  B. MULTI-SHEET configs — many sheets, each with independent random schemas and
     independent random feature configs (styles/merges/CF/DV/tables/charts/
     images/hyperlinks/formulas), across write_sheets_arrow and
     write_sheets_arrow_to_bytes. Probes cross-sheet interactions (shared media,
     per-sheet drawings, table-name uniqueness across sheets).

  C. DICT-BASED paths — write_sheet / write_sheets (the legacy dict API), with
     randomized python value types per column incl. None, mixed types, NaN/Inf,
     huge ints, unicode, datetimes.

Properties asserted (per output that IS produced):
  P1  no crash / no process abort
  P2  loads in openpyxl with warnings-as-errors
  P3  conforms to the ECMA-376 schema (Part-3 ext wildcard + W3C xml: attrs honored)

Deterministic (JETXL_FUZZ_SEED). On failure the config is printed + written to
/tmp/fuzz_expanded_repro.txt for a reproducible repro.

Usage: python fuzz_jetxl_expanded.py [N] [-v]
"""
from __future__ import annotations

import datetime as dt
import io
import os
import random
import sys
import tempfile
import warnings
import zipfile

warnings.filterwarnings("ignore")

import pyarrow as pa
import openpyxl

import jetxl

N = 300
for a in sys.argv[1:]:
    if a.isdigit():
        N = int(a)
VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv
SEED = int(os.environ.get("JETXL_FUZZ_SEED", "20260705"))

SCHEMA_DIR = os.environ.get("JETXL_ECMA_SCHEMAS", os.path.join(os.path.dirname(__file__), "ecma", "schemas_transitional"))
_schemas = None


def load_schemas():
    global _schemas
    if _schemas is not None:
        return _schemas
    try:
        import xmlschema
        if not os.path.isfile(os.path.join(SCHEMA_DIR, "sml.xsd")):
            _schemas = {}
            return _schemas
        _schemas = {
            "sml": xmlschema.XMLSchema(os.path.join(SCHEMA_DIR, "sml.xsd")),
            "chart": xmlschema.XMLSchema(os.path.join(SCHEMA_DIR, "dml-chart.xsd")),
            "drawing": xmlschema.XMLSchema(os.path.join(SCHEMA_DIR, "dml-spreadsheetDrawing.xsd")),
        }
    except Exception:
        _schemas = {}
    return _schemas


def schema_for(part):
    import re
    s = load_schemas()
    if not s:
        return None
    if re.match(r"xl/(worksheets/sheet|workbook|styles|sharedStrings|tables/table)", part):
        return s["sml"]
    if "charts/chart" in part and part.endswith(".xml"):
        return s["chart"]
    if "drawings/drawing" in part and part.endswith(".xml") and "rels" not in part:
        return s["drawing"]
    return None


def schema_violations(b):
    s = load_schemas()
    if not s:
        return []
    out = []
    z = zipfile.ZipFile(io.BytesIO(b))
    for part in z.namelist():
        sch = schema_for(part)
        if sch is None:
            continue
        try:
            for e in sch.iter_errors(z.read(part).decode("utf-8")):
                path = (e.path or "").lower()
                reason = str(e.reason)
                if "ext" in path:
                    continue
                if ("1998/namespace" in reason.lower() or "xml:space" in reason.lower()
                        or "xml:lang" in reason.lower()):
                    continue
                out.append((part, reason[:80]))
                break
        except Exception as ex:
            out.append((part, f"validator error: {ex!r}"[:80]))
    return out


PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d4948445200000001000000010806000000"
    "1f15c4890000000d49444154789c6260f8cf000000ffff03000006000557"
    "bfabd40000000049454e44ae426082")

UNICODE = ["café", "日本語", "Ω≈ç√", "emoji😀", "a\tb", " lead", "trail ", "",
           "   ", "quote\"x", "amp&<>", "\x01c", "n\x00b", "A" * 200, "0007", "-42"]
COLORS = ["FF0000", "FFFF0000", "#00FF00", "00ff00", "red", "F00", "zzz", "", "638EC6"]


# ==========================================================================
# A. COMPOSITE / EXOTIC Arrow types — must not panic
# ==========================================================================
def composite_arrays(rng):
    """Return a list of (label, arrow_array) of exotic types."""
    out = []
    try: out.append(("struct", pa.array([{"a": 1, "b": "x"}, {"a": 2, "b": "y"}])))
    except Exception: pass
    try: out.append(("list_int", pa.array([[1, 2, 3], [4, 5]])))
    except Exception: pass
    try: out.append(("large_list", pa.array([[1, 2], [3]], pa.large_list(pa.int64()))))
    except Exception: pass
    try: out.append(("fixed_size_list", pa.array([[1, 2], [3, 4]], pa.list_(pa.int64(), 2))))
    except Exception: pass
    try: out.append(("list_str", pa.array([["a", "b"], ["c"]])))
    except Exception: pass
    try: out.append(("map", pa.array([[("k", 1)], [("m", 2)]], pa.map_(pa.string(), pa.int64()))))
    except Exception: pass
    try: out.append(("binary", pa.array([b"hello", b"world"], pa.binary())))
    except Exception: pass
    try: out.append(("fixed_size_binary", pa.array([b"ab", b"cd"], pa.binary(2))))
    except Exception: pass
    try: out.append(("decimal128", pa.array([pa.scalar(3, pa.decimal128(10, 2)), pa.scalar(5, pa.decimal128(10, 2))])))
    except Exception: pass
    try: out.append(("decimal256", pa.array([pa.scalar(3, pa.decimal256(20, 4))])))
    except Exception: pass
    try: out.append(("duration_s", pa.array([1000, 2000], pa.duration("s"))))
    except Exception: pass
    try: out.append(("time32_s", pa.array([3600, 7200], pa.time32("s"))))
    except Exception: pass
    try: out.append(("time64_us", pa.array([3600_000000, 7200_000000], pa.time64("us"))))
    except Exception: pass
    try: out.append(("interval", pa.array([pa.scalar(1, pa.month_day_nano_interval())])))
    except Exception: pass
    try: out.append(("null_type", pa.array([None, None], pa.null())))
    except Exception: pass
    try: out.append(("nested_list_of_struct", pa.array([[{"a": 1}], [{"a": 2}]])))
    except Exception: pass
    try: out.append(("dict_of_int", pa.array([1, 2, 1]).dictionary_encode()))
    except Exception: pass
    return out


def fuzz_composite(rng, stats, failures):
    arrays = composite_arrays(rng)
    for label, arr in arrays:
        n = len(arr)
        # place the exotic column alongside a normal one
        tbl = pa.table({"ok": pa.array(list(range(n)), pa.int64()), "exotic": arr})
        stats["composite_tried"] += 1
        # P1: must not panic. Clean exception is acceptable for unsupported types.
        try:
            b = jetxl.write_sheet_arrow_to_bytes(tbl)
        except (OSError, ValueError, TypeError) as e:
            # clean rejection — acceptable, not a failure
            stats["composite_clean_reject"] += 1
            continue
        except BaseException as e:
            failures.append(("A_PANIC", label, repr(e)[:100], {"type": label}))
            continue
        # if it DID produce output, it must load + conform
        stats["composite_wrote"] += 1
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("error", UserWarning)
                openpyxl.load_workbook(io.BytesIO(b))
        except Exception as e:
            failures.append(("A_LOAD", label, repr(e)[:100], {"type": label}))
            continue
        v = schema_violations(b)
        if v:
            failures.append(("A_SCHEMA", label, str(v[0])[:100], {"type": label}))


# ==========================================================================
# Shared random generators (reuse the base fuzzer's logic, inline for indep.)
# ==========================================================================
def rng_column(rng, n):
    kind = rng.choice(["int64", "uint64", "float64", "bool", "string", "large_string",
                       "date32", "timestamp_us", "categorical", "all_null", "int32", "float32"])
    def mn(vals, ty):
        return pa.array([v if rng.random() > 0.1 else None for v in vals], ty)
    if kind == "int64":
        return mn([rng.choice([0, -1, 2**62, -2**62, rng.randint(-10**9, 10**9)]) for _ in range(n)], pa.int64())
    if kind == "uint64":
        return mn([rng.choice([0, 2**63, 2**64 - 1, rng.randint(0, 2**64 - 1)]) for _ in range(n)], pa.uint64())
    if kind == "int32":
        return mn([rng.randint(-2**31, 2**31 - 1) for _ in range(n)], pa.int32())
    if kind == "float64":
        return mn([rng.choice([0.0, -0.0, 1e308, float("nan"), float("inf"), rng.uniform(-1e6, 1e6)]) for _ in range(n)], pa.float64())
    if kind == "float32":
        return mn([rng.uniform(-1e5, 1e5) for _ in range(n)], pa.float32())
    if kind == "bool":
        return mn([rng.random() > 0.5 for _ in range(n)], pa.bool_())
    if kind == "string":
        return pa.array([rng.choice(UNICODE) if rng.random() > 0.15 else None for _ in range(n)])
    if kind == "large_string":
        return pa.array([rng.choice(UNICODE) for _ in range(n)], pa.large_string())
    if kind == "date32":
        return mn([dt.date(rng.randint(1900, 2100), rng.randint(1, 12), rng.randint(1, 28)) for _ in range(n)], pa.date32())
    if kind == "timestamp_us":
        return mn([dt.datetime(rng.randint(1900, 2100), rng.randint(1, 12), rng.randint(1, 28), rng.randint(0, 23)) for _ in range(n)], pa.timestamp("us"))
    if kind == "categorical":
        return pa.array([rng.choice(UNICODE[:6]) for _ in range(n)]).dictionary_encode()
    if kind == "all_null":
        return pa.array([None] * n, rng.choice([pa.int64(), pa.float64(), pa.string(), pa.null()]))
    return pa.array(list(range(n)), pa.int64())


def rng_table(rng):
    ncols = rng.randint(1, 6)
    nrows = rng.choice([1, 1, 2, 5, 10, 25])
    cols = {}
    for c in range(ncols):
        cols[rng.choice(["col", "值", "a b", "x_1", "F"]) + str(c)] = rng_column(rng, nrows)
    return pa.table(cols), ncols, nrows


def rng_config(rng, ncols, nrows):
    cfg = {}
    last = nrows + 1
    rr = lambda: rng.choice([rng.randint(1, last), rng.randint(1, last + 4)])
    rc = lambda: rng.choice([rng.randint(0, ncols - 1), rng.randint(0, ncols + 3)])
    if rng.random() < 0.5:
        styles = []
        for _ in range(rng.randint(1, 3)):
            st = {"row": rr(), "col": rc()}
            if rng.random() < 0.7:
                st["font"] = {"bold": rng.random() < 0.5}
                if rng.random() < 0.6: st["font"]["color"] = rng.choice(COLORS)
                if rng.random() < 0.4: st["font"]["size"] = rng.choice([0, 11, 409, 1000])
            if rng.random() < 0.4: st["fill"] = {"pattern": "solid", "fg_color": rng.choice(COLORS)}
            if rng.random() < 0.3: st["alignment"] = {"text_rotation": rng.choice([0, 45, -90, 200, 255])}
            styles.append(st)
        cfg["cell_styles"] = styles
    if rng.random() < 0.3:
        cfg["merge_cells"] = [(rr(), rc(), rr(), rc()) for _ in range(rng.randint(1, 3))]
    if rng.random() < 0.3:
        cfs = []
        for _ in range(rng.randint(1, 2)):
            rt = rng.choice(["cell_value", "color_scale", "data_bar", "top10"])
            cf = {"start_row": rr(), "start_col": rc(), "end_row": rr(), "end_col": rc(), "rule_type": rt}
            if rt == "cell_value":
                cf["operator"] = rng.choice(["greater_than", "less_than", "between"]); cf["value"] = str(rng.randint(0, 99))
                if cf["operator"] == "between": cf["value2"] = str(rng.randint(100, 200))
            elif rt == "color_scale":
                cf["min_color"] = rng.choice(COLORS); cf["max_color"] = rng.choice(COLORS)
                if rng.random() < 0.5: cf["mid_color"] = rng.choice(COLORS)
            elif rt == "data_bar":
                cf["color"] = rng.choice(COLORS)
            else:
                cf["rank"] = rng.randint(1, 20); cf["bottom"] = rng.random() < 0.5
            cfs.append(cf)
        cfg["conditional_formats"] = cfs
    if rng.random() < 0.3:
        dvs = []
        for _ in range(rng.randint(1, 2)):
            vt = rng.choice(["list", "whole_number", "decimal"])
            dv = {"start_row": rr(), "start_col": rc(), "end_row": rr(), "end_col": rc(), "type": vt}
            if vt == "list": dv["items"] = [rng.choice(UNICODE[:5]) for _ in range(rng.randint(1, 4))]
            else: dv["min"] = rng.randint(0, 50); dv["max"] = rng.randint(0, 50)
            dvs.append(dv)
        cfg["data_validations"] = dvs
    if rng.random() < 0.3:
        cfg["tables"] = [{"name": rng.choice(["T", "", "My Table", "值", "D"]),
                          "start_row": 0, "start_col": rng.randint(0, max(0, ncols - 1)),
                          "end_row": rng.choice([nrows, nrows + 2]),
                          "end_col": rng.choice([ncols - 1, ncols + 2]),
                          "style": rng.choice(["TableStyleMedium9", "NotReal"])} for _ in range(rng.randint(1, 2))]
    if rng.random() < 0.25:
        charts = []
        for _ in range(rng.randint(1, 2)):
            ct = rng.choice(["column", "bar", "line", "pie", "scatter", "area"])
            ch = {"chart_type": ct, "data_range": (0, min(1, ncols - 1), nrows, min(1, ncols - 1)), "category_col": 0}
            if rng.random() < 0.6: ch["title"] = rng.choice(UNICODE[:5])
            if rng.random() < 0.4: ch["title_color"] = rng.choice(COLORS)
            if rng.random() < 0.4: ch["title_font_size"] = rng.choice([0, 1400, 999999])
            if rng.random() < 0.3: ch["y_axis_title"] = "Y"; ch["axis_title_color"] = rng.choice(COLORS)
            charts.append(ch)
        cfg["charts"] = charts
    if rng.random() < 0.2:
        cfg["images"] = [{"data": list(PNG), "extension": "png", "from_col": rng.randint(0, ncols),
                          "from_row": rng.randint(0, last), "to_col": rng.randint(0, ncols + 3), "to_row": rng.randint(0, last + 3)}]
    if rng.random() < 0.3:
        cfg["hyperlinks"] = [(rr(), rc(), rng.choice(["https://x.com", "mailto:a@b.com", "#Sheet1!A1", ""]), "l") for _ in range(rng.randint(1, 2))]
    if rng.random() < 0.2:
        cfg["formulas"] = [(rr(), rc(), rng.choice(["=A1+B1", "A1+B1", "=SUM(A1:A5)"]), "0") for _ in range(rng.randint(1, 2))]
    if rng.random() < 0.3: cfg["auto_filter"] = True
    if rng.random() < 0.3: cfg["freeze_rows"] = rng.choice([0, 1, 100])
    return cfg


# ==========================================================================
# B. MULTI-SHEET fuzzing
# ==========================================================================
def fuzz_multisheet(rng, stats, failures):
    nsheets = rng.randint(2, 6)
    sheets = []
    used_names = set()
    for i in range(nsheets):
        data, nc, nr = rng_table(rng)
        cfg = rng_config(rng, nc, nr)
        name = rng.choice(["Sheet", "数据", "S", "Tab", "值"]) + str(i)
        # occasionally force a duplicate-ish name to probe sheet-name handling
        sheet = {"data": data, "name": name}
        sheet.update(cfg)
        sheets.append(sheet)

    # bytes path
    stats["multi_tried"] += 1
    try:
        b = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in sheets], rng.randint(1, 8))
    except (OSError, ValueError, TypeError) as e:
        stats["multi_clean_reject"] += 1
        b = None
    except BaseException as e:
        failures.append(("B_PANIC_bytes", f"{nsheets}sheets", repr(e)[:100], {"sheets": [s["name"] for s in sheets]}))
        b = None
    if b is not None:
        stats["multi_wrote"] += 1
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("error", UserWarning)
                openpyxl.load_workbook(io.BytesIO(b))
        except Exception as e:
            failures.append(("B_LOAD_bytes", f"{nsheets}sheets", repr(e)[:100], {"sheets": [s["name"] for s in sheets]}))
        else:
            v = schema_violations(b)
            if v:
                failures.append(("B_SCHEMA_bytes", f"{nsheets}sheets", str(v[0])[:100], {"sheets": [s["name"] for s in sheets]}))

    # file path (sometimes)
    if rng.random() < 0.5:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            p = tf.name
        try:
            jetxl.write_sheets_arrow([dict(s) for s in sheets], p, rng.randint(1, 8))
            fb = open(p, "rb").read()
            with warnings.catch_warnings():
                warnings.simplefilter("error", UserWarning)
                openpyxl.load_workbook(io.BytesIO(fb))
        except (OSError, ValueError, TypeError):
            pass
        except BaseException as e:
            failures.append(("B_FILE", f"{nsheets}sheets", repr(e)[:100], {"sheets": [s["name"] for s in sheets]}))
        finally:
            if os.path.exists(p):
                os.unlink(p)


# ==========================================================================
# C. DICT-based path fuzzing
# ==========================================================================
def rng_dict_value(rng):
    return rng.choice([
        rng.randint(-10**6, 10**6), rng.uniform(-1e6, 1e6), rng.choice(UNICODE),
        rng.random() > 0.5, None, float("nan"), float("inf"), 10**20,
        dt.datetime(2024, 1, 1), dt.date(2023, 6, 15), -0.0,
    ])


def fuzz_dict(rng, stats, failures):
    ncols = rng.randint(1, 5)
    nrows = rng.choice([1, 2, 5, 10])
    cols = {}
    for c in range(ncols):
        name = rng.choice(["a", "col", "值", "x y"]) + str(c)
        # each column: sometimes homogeneous, sometimes mixed types
        if rng.random() < 0.5:
            base = rng.choice(["int", "float", "str", "bool"])
            vals = []
            for _ in range(nrows):
                if base == "int": vals.append(rng.randint(-1000, 1000) if rng.random() > 0.1 else None)
                elif base == "float": vals.append(rng.choice([rng.uniform(-1e3, 1e3), float("nan"), float("inf")]))
                elif base == "str": vals.append(rng.choice(UNICODE))
                else: vals.append(rng.random() > 0.5)
            cols[name] = vals
        else:
            cols[name] = [rng_dict_value(rng) for _ in range(nrows)]

    stats["dict_tried"] += 1
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheet(cols, p, sheet_name=rng.choice(["Sheet1", "数据", "T"]))
        stats["dict_wrote"] += 1
        fb = open(p, "rb").read()
        with warnings.catch_warnings():
            warnings.simplefilter("error", UserWarning)
            openpyxl.load_workbook(io.BytesIO(fb))
        v = schema_violations(fb)
        if v:
            failures.append(("C_SCHEMA", "dict", str(v[0])[:100], cols))
    except (OSError, ValueError, TypeError):
        stats["dict_clean_reject"] += 1
    except BaseException as e:
        failures.append(("C_PANIC_or_LOAD", "dict", repr(e)[:100], cols))
    finally:
        if os.path.exists(p):
            os.unlink(p)


# ==========================================================================
def main():
    print("=" * 74)
    print(f"jetxl EXPANDED property fuzzing — {N} iterations, seed={SEED}")
    print(f"schema validation (P3): {'ENABLED' if load_schemas() else 'unavailable'}")
    print("coverage: A) composite Arrow types  B) multi-sheet  C) dict paths")
    print("=" * 74)

    master = random.Random(SEED)
    stats = {k: 0 for k in ["composite_tried", "composite_wrote", "composite_clean_reject",
                            "multi_tried", "multi_wrote", "multi_clean_reject",
                            "dict_tried", "dict_wrote", "dict_clean_reject"]}
    failures = []

    for i in range(N):
        rng = random.Random(master.randint(0, 2**31))
        try:
            fuzz_composite(rng, stats, failures)
            fuzz_multisheet(rng, stats, failures)
            fuzz_dict(rng, stats, failures)
        except Exception as e:
            import traceback
            traceback.print_exc()
            failures.append(("HARNESS", str(i), repr(e)[:100], {}))
        if VERBOSE and i % 40 == 0:
            print(f"  ... {i}/{N}  failures so far: {len(failures)}")

    print("\n" + "=" * 74)
    print("coverage stats:")
    print(f"  A composite: {stats['composite_tried']} tried, "
          f"{stats['composite_wrote']} wrote, {stats['composite_clean_reject']} clean-rejected")
    print(f"  B multi:     {stats['multi_tried']} tried, "
          f"{stats['multi_wrote']} wrote, {stats['multi_clean_reject']} clean-rejected")
    print(f"  C dict:      {stats['dict_tried']} tried, "
          f"{stats['dict_wrote']} wrote, {stats['dict_clean_reject']} clean-rejected")
    print(f"\ntotal failures: {len(failures)}")
    if failures:
        from collections import Counter
        print("by class:", dict(Counter(f[0] for f in failures)))
        print("\n--- first 12 ---")
        for kind, label, msg, cfg in failures[:12]:
            print(f"\n[{kind}] {label}")
            print(f"  {msg}")
        with open("/tmp/fuzz_expanded_repro.txt", "w") as f:
            for kind, label, msg, cfg in failures:
                f.write(f"{kind}\t{label}\t{msg}\n\t{cfg}\n\n")
        print("\n(repros -> /tmp/fuzz_expanded_repro.txt)")
    print("=" * 74)
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
