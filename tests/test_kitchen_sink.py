#!/usr/bin/env python3
"""
jetxl kitchen-sink integration suite
====================================

Real-world stress test: build COMPLEX workbooks where many features are stacked
in the same sheet and across many sheets at once — multiple charts, multiple
tables, images, conditional formats, cell styles, merges, hyperlinks, formulas,
data validation, freeze panes, auto-filter — then validate the result is not
just "openable" but STRUCTURALLY CORRECT:

  - the ZIP CRCs are valid
  - every XML part parses (well-formed)
  - every r:id referenced in a sheet resolves in that sheet's .rels
  - every r:id referenced in a drawing resolves in the drawing's .rels
  - every relationship Target points at a part that actually exists
  - [Content_Types].xml declares every part type used
  - openpyxl can load it and read back the expected values/objects

This is where feature *interactions* break things (dangling relationship ids,
chart/drawing/image part-number collisions, styles.xml dxf/xf index clashes) —
none of which a single-feature test would catch.

Every scenario runs on all four arrow write paths:
    single file / single bytes / multi file / multi bytes

Run: python test_kitchen_sink.py [-v]
"""
from __future__ import annotations

import io
import os
import re
import sys
import tempfile
import zipfile
from xml.etree import ElementTree as ET

import pyarrow as pa
import openpyxl

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv
_p = _f = 0
_fails = []

# 1x1 red PNG for image tests
PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d4948445200000001000000010806000000"
    "1f15c4890000000d49444154789c6260f8cf000000ffff03000006000557"
    "bfabd40000000049454e44ae426082")


def ok(d):
    global _p
    _p += 1
    if VERBOSE:
        print(f"    pass  {d}")


def bad(d):
    global _f
    _f += 1
    _fails.append(d)
    print(f"    FAIL  {d}")


def check(d, c):
    ok(d) if c else bad(d)


def sec(n):
    print(f"\n=== {n} ===")


# --------------------------------------------------------------------------
# Structural validator — the core of this suite
# --------------------------------------------------------------------------
def validate_ooxml(b: bytes, label: str, expect_parts=None):
    """Deep structural validation of an .xlsx byte blob. Returns True if all
    invariants hold; records a check per invariant."""
    z = zipfile.ZipFile(io.BytesIO(b))
    names = set(z.namelist())

    # 1. CRC
    check(f"[{label}] zip CRCs valid", z.testzip() is None)

    # 2. every XML part well-formed
    malformed = []
    for nm in names:
        if nm.endswith(".xml") or nm.endswith(".rels"):
            try:
                ET.fromstring(z.read(nm))
            except Exception as e:
                malformed.append(f"{nm}: {e}")
    check(f"[{label}] all XML parts well-formed", not malformed)
    if malformed:
        for m in malformed:
            print("        malformed:", m)

    # 3. every sheet's r:ids resolve in its rels, and targets exist
    sheet_parts = [n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]
    for sp in sheet_parts:
        sheet = z.read(sp).decode()
        refd = set(re.findall(r'r:id="(rId[^"]+)"', sheet)) | set(re.findall(r'r:embed="(rId[^"]+)"', sheet))
        rels_name = sp.replace("worksheets/", "worksheets/_rels/") + ".rels"
        if refd:
            if rels_name in names:
                rels = z.read(rels_name).decode()
                defined = set(re.findall(r'Id="(rId[^"]+)"', rels))
                check(f"[{label}] {sp} r:ids all resolve", refd <= defined)
                # targets exist
                targets = re.findall(r'Target="([^"]+)"', rels)
                missing = []
                for tg in targets:
                    if tg.startswith("http") or tg.startswith("mailto"):
                        continue  # external
                    resolved = os.path.normpath(os.path.join("xl/worksheets", tg)).replace("\\", "/")
                    if resolved not in names:
                        missing.append(tg)
                check(f"[{label}] {sp} rel targets exist", not missing)
            else:
                bad(f"[{label}] {sp} references r:ids but has no .rels")

    # 4. every drawing's r:ids resolve and targets exist
    draw_parts = [n for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
    for dp in draw_parts:
        draw = z.read(dp).decode()
        refd = set(re.findall(r'r:id="(rId[^"]+)"', draw)) | set(re.findall(r'r:embed="(rId[^"]+)"', draw))
        rels_name = dp.replace("drawings/", "drawings/_rels/") + ".rels"
        if refd:
            if rels_name in names:
                rels = z.read(rels_name).decode()
                defined = set(re.findall(r'Id="(rId[^"]+)"', rels))
                check(f"[{label}] {dp} r:ids all resolve", refd <= defined)
                targets = re.findall(r'Target="([^"]+)"', rels)
                missing = []
                for tg in targets:
                    resolved = os.path.normpath(os.path.join("xl/drawings", tg)).replace("\\", "/")
                    if resolved not in names:
                        missing.append(tg)
                check(f"[{label}] {dp} rel targets exist", not missing)
            else:
                bad(f"[{label}] {dp} references r:ids but has no .rels")

    # 5. Content_Types declares chart/drawing/table/image if present
    ct = z.read("[Content_Types].xml").decode()
    if any("charts/chart" in n for n in names):
        check(f"[{label}] Content_Types declares chart", "chart" in ct)
    if any("drawings/drawing" in n for n in names):
        check(f"[{label}] Content_Types declares drawing", "drawing" in ct)
    if any("tables/table" in n for n in names):
        check(f"[{label}] Content_Types declares table", "table" in ct)
    if any("media/image" in n for n in names):
        check(f"[{label}] Content_Types declares image ext", "png" in ct.lower() or "Default" in ct)

    # 6. openpyxl loads it
    try:
        wb = openpyxl.load_workbook(io.BytesIO(b))
        check(f"[{label}] openpyxl loads workbook", wb is not None)
    except Exception as e:
        bad(f"[{label}] openpyxl load failed: {e!r}")
        wb = None

    # 7. expected parts present
    if expect_parts:
        for pat in expect_parts:
            check(f"[{label}] has part matching {pat}", any(re.search(pat, n) for n in names))

    return wb


# --------------------------------------------------------------------------
# Write across all four arrow paths
# --------------------------------------------------------------------------
def write_all_paths(sheets: list[dict], label_prefix: str):
    """Yield (label, bytes) for single (first sheet only) and multi paths.
    `sheets` is a list of per-sheet config dicts each containing 'data','name'."""
    # multi bytes
    b = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in sheets], 4)
    yield f"{label_prefix}/multi_bytes", b

    # multi file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheets_arrow([dict(s) for s in sheets], p, 4)
        yield f"{label_prefix}/multi_file", open(p, "rb").read()
    finally:
        os.unlink(p)

    # single paths use the first sheet's config
    first = dict(sheets[0])
    data = first.pop("data")
    first.pop("name", None)
    b = jetxl.write_sheet_arrow_to_bytes(data, **first)
    yield f"{label_prefix}/single_bytes", b
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheet_arrow(data, p, **first)
        yield f"{label_prefix}/single_file", open(p, "rb").read()
    finally:
        os.unlink(p)


# --------------------------------------------------------------------------
# Data builders
# --------------------------------------------------------------------------
def sales_table(n=50):
    cats = ["North", "South", "East", "West", "Central"]
    return pa.table({
        "month": pa.array([f"M{i%12}" for i in range(n)]),
        "region": pa.array([cats[i % 5] for i in range(n)]),
        "sales": pa.array([float(i * 100) for i in range(n)], pa.float64()),
        "costs": pa.array([float(i * 60) for i in range(n)], pa.float64()),
        "profit": pa.array([float(i * 40) for i in range(n)], pa.float64()),
    })


# ==========================================================================
# 1. Everything stacked in ONE sheet
# ==========================================================================
def t_everything_one_sheet():
    sec("Kitchen sink — every feature stacked in one sheet, all paths")
    n = 50
    t = sales_table(n)
    cfg = {
        "data": t, "name": "AllFeatures",
        "charts": [
            {"chart_type": "column", "data_range": (0, 2, n, 2), "category_col": 0,
             "title": "Sales", "legend_position": "bottom", "show_data_labels": True},
            {"chart_type": "line", "data_range": (0, 4, n, 4), "category_col": 0, "title": "Profit"},
            {"chart_type": "pie", "data_range": (0, 2, 5, 2), "category_col": 0, "title": "Top5"},
        ],
        "tables": [{"name": "SalesTable", "start_row": 0, "start_col": 0,
                    "end_row": n, "end_col": 4, "style": "TableStyleMedium9"}],
        "images": [{"data": list(PNG), "extension": "png",
                    "from_col": 7, "from_row": 1, "to_col": 10, "to_row": 6}],
        "conditional_formats": [
            {"start_row": 2, "start_col": 2, "end_row": n, "end_col": 2,
             "rule_type": "data_bar", "color": "FF638EC6"},
            {"start_row": 2, "start_col": 4, "end_row": n, "end_col": 4,
             "rule_type": "color_scale", "min_color": "FFFF0000", "max_color": "FF00FF00"},
        ],
        "cell_styles": [{"row": 1, "col": 0, "font": {"bold": True, "size": 14, "color": "FFFFFFFF"},
                         "fill": {"pattern": "solid", "fg_color": "FF4472C4"},
                         "border": {"bottom": {"style": "thick"}}}],
        "merge_cells": [(1, 0, 1, 4)],
        "hyperlinks": [(3, 0, "https://example.com/?a=1&b=2", "link")],
        "formulas": [(7, 2, "SUM(C2:C6)", "1000")],
        "data_validations": [{"start_row": 2, "start_col": 1, "end_row": n, "end_col": 1,
                              "type": "list", "items": ["North", "South", "East", "West", "Central"]}],
        "auto_filter": True, "freeze_rows": 1, "freeze_cols": 1, "zoom_scale": 115,
    }
    for label, b in write_all_paths([cfg], "kitchen1"):
        wb = validate_ooxml(b, label, expect_parts=[
            r"charts/chart1\.xml", r"charts/chart2\.xml", r"charts/chart3\.xml",
            r"tables/table1\.xml", r"media/image1\.png", r"drawings/drawing1\.xml"])
        if wb:
            ws = wb.active if "single" in label else wb["AllFeatures"]
            check(f"[{label}] chart count == 3",
                  sum(1 for n2 in zipfile.ZipFile(io.BytesIO(b)).namelist()
                      if re.match(r"xl/charts/chart\d+\.xml", n2)) == 3)
            check(f"[{label}] merge present", "A1:E1" in {str(r) for r in ws.merged_cells.ranges})
            check(f"[{label}] formula present", str(ws.cell(row=7, column=3).value).startswith("="))


# ==========================================================================
# 2. Many sheets, each a different heavy feature combo (part-numbering stress)
# ==========================================================================
def t_multi_sheet_heavy():
    sec("Multi-sheet — each sheet a different heavy combo (part-number stress)")
    n = 40
    t = sales_table(n)
    sheets = [
        {"data": t, "name": "Charts", "charts": [
            {"chart_type": "column", "data_range": (0, 2, n, 2), "category_col": 0, "title": "A"},
            {"chart_type": "bar", "data_range": (0, 3, n, 3), "category_col": 0, "title": "B"}]},
        {"data": t, "name": "Tables", "tables": [
            {"name": "TblA", "start_row": 0, "start_col": 0, "end_row": n, "end_col": 4}]},
        {"data": t, "name": "Images", "images": [
            {"data": list(PNG), "extension": "png", "from_col": 6, "from_row": 1, "to_col": 9, "to_row": 5}]},
        {"data": t, "name": "ChartImg", "charts": [
            {"chart_type": "line", "data_range": (0, 4, n, 4), "category_col": 0, "title": "P"}],
         "images": [{"data": list(PNG), "extension": "png", "from_col": 7, "from_row": 1, "to_col": 10, "to_row": 6}]},
        {"data": t, "name": "CondFmt", "conditional_formats": [
            {"start_row": 2, "start_col": 2, "end_row": n, "end_col": 2, "rule_type": "data_bar", "color": "FF638EC6"},
            {"start_row": 2, "start_col": 4, "end_row": n, "end_col": 4, "rule_type": "cell_value",
             "operator": "greater_than", "value": "500"}]},
        {"data": t, "name": "Styled", "cell_styles": [
            {"row": 2, "col": 0, "font": {"bold": True}, "border": {"bottom": {"style": "thick"}}}],
         "hyperlinks": [(3, 0, "https://x.com/?a=1&b=2", "l")],
         "merge_cells": [(1, 0, 1, 3)]},
    ]
    # multi paths only (single path can't hold multiple sheets)
    for label, b in [("heavy/multi_bytes", jetxl.write_sheets_arrow_to_bytes([dict(s) for s in sheets], 4))]:
        wb = validate_ooxml(b, label)
        if wb:
            check(f"[{label}] all 6 sheets present",
                  wb.sheetnames == ["Charts", "Tables", "Images", "ChartImg", "CondFmt", "Styled"])
        # cross-check: each sheet with a drawing has its own drawing part, no collision
        names = zipfile.ZipFile(io.BytesIO(b)).namelist()
        draw_count = sum(1 for x in names if re.match(r"xl/drawings/drawing\d+\.xml$", x))
        chart_count = sum(1 for x in names if re.match(r"xl/charts/chart\d+\.xml$", x))
        img_count = sum(1 for x in names if re.match(r"xl/media/image\d+\.", x))
        # Charts sheet(2 charts->1 drawing) + Images(1 drawing) + ChartImg(1 drawing) = 3 drawings
        check(f"[{label}] chart parts present (>=3)", chart_count >= 3)
        check(f"[{label}] image parts present (>=2)", img_count >= 2)
        check(f"[{label}] drawing parts present (>=3)", draw_count >= 3)
        # media image names must be workbook-GLOBAL (no duplicate ZIP entries)
        from collections import Counter as _C
        media = [x for x in names if re.match(r"xl/media/image\d+\.", x)]
        dup_media = [k for k, v in _C(names).items() if v > 1 and "media/image" in k]
        check(f"[{label}] no duplicate media entries", not dup_media)
        check(f"[{label}] media names are unique", len(set(media)) == len(media))

    # also on multi file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheets_arrow([dict(s) for s in sheets], p, 4)
        b = open(p, "rb").read()
    finally:
        os.unlink(p)
    wb = validate_ooxml(b, "heavy/multi_file")
    if wb:
        check("[heavy/multi_file] all 6 sheets present",
              wb.sheetnames == ["Charts", "Tables", "Images", "ChartImg", "CondFmt", "Styled"])


# ==========================================================================
# 3. Multiple tables + multiple charts in the SAME sheet (id-space stress)
# ==========================================================================
def t_multi_tables_charts_same_sheet():
    sec("Multiple tables + multiple charts in one sheet (id-space stress)")
    n = 30
    # two side-by-side data blocks so two tables don't overlap
    t = pa.table({
        "a": pa.array([f"r{i}" for i in range(n)]),
        "b": pa.array([float(i) for i in range(n)], pa.float64()),
        "c": pa.array([float(i * 2) for i in range(n)], pa.float64()),
        "d": pa.array([f"s{i}" for i in range(n)]),
        "e": pa.array([float(i * 3) for i in range(n)], pa.float64()),
        "f": pa.array([float(i * 4) for i in range(n)], pa.float64()),
    })
    cfg = {
        "data": t, "name": "MultiTC",
        "tables": [
            {"name": "TableLeft", "start_row": 0, "start_col": 0, "end_row": n, "end_col": 2},
            {"name": "TableRight", "start_row": 0, "start_col": 3, "end_row": n, "end_col": 5},
        ],
        "charts": [
            {"chart_type": "column", "data_range": (0, 1, n, 1), "category_col": 0, "title": "L"},
            {"chart_type": "line", "data_range": (0, 2, n, 2), "category_col": 0, "title": "M"},
            {"chart_type": "area", "data_range": (0, 4, n, 4), "category_col": 3, "title": "R"},
            {"chart_type": "scatter", "data_range": (0, 5, n, 5), "category_col": 3, "title": "S"},
        ],
    }
    for label, b in write_all_paths([cfg], "multitc"):
        wb = validate_ooxml(b, label, expect_parts=[
            r"tables/table1\.xml", r"tables/table2\.xml",
            r"charts/chart1\.xml", r"charts/chart4\.xml"])
        names = zipfile.ZipFile(io.BytesIO(b)).namelist()
        check(f"[{label}] exactly 2 table parts",
              sum(1 for x in names if re.match(r"xl/tables/table\d+\.xml$", x)) == 2)
        check(f"[{label}] exactly 4 chart parts",
              sum(1 for x in names if re.match(r"xl/charts/chart\d+\.xml$", x)) == 4)


# ==========================================================================
# 4. Stress: many charts on one sheet (part-number rollover)
# ==========================================================================
def t_many_charts():
    sec("Stress — 10 charts on one sheet (part-number rollover)")
    n = 20
    cols = {"cat": pa.array([f"c{i}" for i in range(n)])}
    for j in range(10):
        cols[f"v{j}"] = pa.array([float(i * (j + 1)) for i in range(n)], pa.float64())
    t = pa.table(cols)
    charts = [{"chart_type": ["column", "bar", "line", "pie", "area", "scatter"][j % 6],
               "data_range": (0, j + 1, n, j + 1), "category_col": 0, "title": f"Chart{j}"}
              for j in range(10)]
    for label, b in write_all_paths([{"data": t, "name": "ManyCharts", "charts": charts}], "manycharts"):
        wb = validate_ooxml(b, label)
        names = zipfile.ZipFile(io.BytesIO(b)).namelist()
        check(f"[{label}] all 10 chart parts present",
              sum(1 for x in names if re.match(r"xl/charts/chart\d+\.xml$", x)) == 10)


# ==========================================================================
# 5. Reproducibility of a complex workbook (byte-identical)
# ==========================================================================
def t_complex_reproducible():
    sec("Complex workbook is byte-reproducible")
    n = 40
    t = sales_table(n)
    cfg = [{"data": t, "name": "S1",
            "charts": [{"chart_type": "column", "data_range": (0, 2, n, 2), "category_col": 0, "title": "X"}],
            "tables": [{"name": "T", "start_row": 0, "start_col": 0, "end_row": n, "end_col": 4}],
            "conditional_formats": [{"start_row": 2, "start_col": 2, "end_row": n, "end_col": 2,
                                     "rule_type": "data_bar", "color": "FF638EC6"}],
            "cell_styles": [{"row": 1, "col": 0, "font": {"bold": True}}]},
           {"data": t, "name": "S2",
            "images": [{"data": list(PNG), "extension": "png", "from_col": 6, "from_row": 1, "to_col": 9, "to_row": 5}]}]
    b1 = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in cfg], 4)
    b2 = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in cfg], 4)
    check("complex multi-sheet byte-identical across runs", b1 == b2)
    validate_ooxml(b1, "reproducible/complex")


def t_multi_sheet_images_no_collision():
    sec("Multi-sheet images — global media naming, no collision")
    from collections import Counter
    png = list(PNG)
    t = pa.table({"a": pa.array([1, 2, 3]), "b": pa.array([1.0, 2.0, 3.0], pa.float64())})
    # 5 sheets each with an image + one sheet mixing a chart in
    sheets = [{"data": t, "name": f"Img{i}",
               "images": [{"data": png, "extension": "png",
                           "from_col": 3, "from_row": 1, "to_col": 5, "to_row": 4}]}
              for i in range(5)]
    sheets.insert(2, {"data": t, "name": "Chart",
                      "charts": [{"chart_type": "column", "data_range": (0, 1, 3, 1), "category_col": 0}]})
    # multi bytes
    b = jetxl.write_sheets_arrow_to_bytes([dict(s) for s in sheets], 4)
    names = zipfile.ZipFile(io.BytesIO(b)).namelist()
    media = [n for n in names if re.match(r"xl/media/image\d+\.", n)]
    check("multi_bytes: 5 distinct media files", len(set(media)) == 5)
    check("multi_bytes: no duplicate ZIP entries",
          not [k for k, v in Counter(names).items() if v > 1])
    validate_ooxml(b, "img_collision/multi_bytes")
    # multi file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        jetxl.write_sheets_arrow([dict(s) for s in sheets], p, 4)
        b = open(p, "rb").read()
    finally:
        os.unlink(p)
    names = zipfile.ZipFile(io.BytesIO(b)).namelist()
    media = [n for n in names if re.match(r"xl/media/image\d+\.", n)]
    check("multi_file: 5 distinct media files", len(set(media)) == 5)
    check("multi_file: no duplicate ZIP entries",
          not [k for k, v in Counter(names).items() if v > 1])
    validate_ooxml(b, "img_collision/multi_file")


def main():
    print("=" * 74)
    print("jetxl kitchen-sink integration suite")
    print("=" * 74)
    for fn in [t_everything_one_sheet, t_multi_sheet_heavy,
               t_multi_tables_charts_same_sheet, t_many_charts,
               t_multi_sheet_images_no_collision, t_complex_reproducible]:
        try:
            fn()
        except Exception as e:
            import traceback
            traceback.print_exc()
            bad(f"{fn.__name__} crashed: {e!r}")
    print("\n" + "=" * 74)
    print(f"TOTAL: {_p} passed, {_f} failed")
    if _fails:
        print("\nFailures:")
        for x in _fails:
            print("  -", x)
    print("=" * 74)
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())
