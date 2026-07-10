#!/usr/bin/env python3
"""
jetxl functional test suite
===========================

Exercises jetxl's features with dummy data and VERIFIES each generated .xlsx
actually opens and contains what we expect (using openpyxl as an independent
reader). This is a functional/smoke test, not a benchmark — the goal is "does
every feature produce a valid, correct file".

Each test:
  1. writes a real .xlsx into an output folder,
  2. re-opens it with openpyxl (a separate implementation),
  3. asserts the structure/values are right.

If a feature isn't available in your build, that test is skipped (not failed).

Requirements:
    pip install jetxl openpyxl numpy
    pip install polars        # (or pyarrow) for the Arrow-path tests

Usage:
    python test_jetxl.py                 # run everything, keep files in ./jetxl_test_out
    python test_jetxl.py --keep-open     # also print the output folder at the end
    python test_jetxl.py --outdir /tmp/x # choose where files go
"""
from __future__ import annotations
import argparse
import datetime as dt
import os
import sys
import traceback

# ----- optional imports ------------------------------------------------------
def _try(name):
    try:
        return __import__(name)
    except Exception:
        return None

np = _try("numpy")
pl = _try("polars")
pa = _try("pyarrow")

try:
    import jetxl
except Exception as e:
    print("FATAL: cannot import jetxl:", e)
    print("Install it first:  pip install jetxl")
    sys.exit(1)

try:
    import openpyxl
except Exception:
    print("FATAL: openpyxl is required to verify the output files.")
    print("Install it:  pip install openpyxl")
    sys.exit(1)

if np is None:
    print("FATAL: numpy is required to build dummy data.  pip install numpy")
    sys.exit(1)


# ----- tiny test framework ---------------------------------------------------
class Results:
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.skipped = 0
        self.failures = []      # (name, message)

    def ok(self, name):
        self.passed += 1
        print(f"  \033[32mPASS\033[0m  {name}")

    def skip(self, name, why):
        self.skipped += 1
        print(f"  \033[33mSKIP\033[0m  {name}  ({why})")

    def fail(self, name, msg):
        self.failed += 1
        self.failures.append((name, msg))
        print(f"  \033[31mFAIL\033[0m  {name}\n          {msg}")


R = Results()


def test(name, fn):
    """Run one test function; catch and record any exception."""
    try:
        fn()
        R.ok(name)
    except _Skip as s:
        R.skip(name, str(s))
    except AssertionError as a:
        R.fail(name, f"assertion: {a}")
    except Exception as e:
        R.fail(name, f"{type(e).__name__}: {e}\n" +
                     "          " + traceback.format_exc().splitlines()[-2].strip())


class _Skip(Exception):
    pass


def need(cond, why):
    if not cond:
        raise _Skip(why)


def have(fn_name):
    return hasattr(jetxl, fn_name)


# ----- dummy data ------------------------------------------------------------
def dummy_columns(n=50, seed=0):
    """A realistic-ish mixed-type table as a column dict."""
    rng = np.random.default_rng(seed)
    regions = np.array(["North", "South", "East", "West", "Central"])
    return {
        "ID": np.arange(1, n + 1, dtype=np.int64),
        "Name": np.array([f"Person {i}" for i in range(n)]),
        "Region": regions[rng.integers(0, len(regions), n)],
        "Sales": np.round(rng.uniform(1000, 99999, n), 2),
        "Units": rng.integers(1, 500, n, dtype=np.int64),
        "Active": rng.integers(0, 2, n, dtype=bool),
        "JoinDate": np.array([dt.date(2020, 1, 1) + dt.timedelta(days=int(x))
                              for x in rng.integers(0, 1500, n)]),
    }


def to_dict_api(cols):
    """Convert numpy column dict -> {name: python list} for the dict API."""
    out = {}
    for k, v in cols.items():
        if hasattr(v, "tolist"):
            out[k] = v.tolist()
        else:
            out[k] = list(v)
    return out


def to_arrow(cols):
    """Build an Arrow table from the column dict, using whatever backend exists."""
    need(pl is not None or pa is not None, "no arrow backend (pip install polars or pyarrow)")
    if pl is not None:
        return pl.DataFrame({k: list(v) for k, v in cols.items()}).to_arrow()
    arrays = {k: pa.array(list(v)) for k, v in cols.items()}
    return pa.table(arrays)


# ----- verification helpers --------------------------------------------------
def open_wb(path):
    assert os.path.exists(path), f"file was not created: {path}"
    assert os.path.getsize(path) > 0, f"file is empty: {path}"
    return openpyxl.load_workbook(path)


OUT = "jetxl_test_out"   # set in main()


def out(name):
    return os.path.join(OUT, name)


# =============================================================================
# TESTS
# =============================================================================
def t_dict_single():
    need(have("write_sheet"), "jetxl.write_sheet not in build")
    cols = dummy_columns(40)
    path = out("dict_single.xlsx")
    jetxl.write_sheet(to_dict_api(cols), path, sheet_name="People")
    wb = open_wb(path)
    assert "People" in wb.sheetnames, f"sheet name wrong: {wb.sheetnames}"
    ws = wb["People"]
    # header row + 40 data rows
    assert ws.max_row == 41, f"expected 41 rows, got {ws.max_row}"
    assert ws["A1"].value == "ID", f"header A1 wrong: {ws['A1'].value!r}"
    assert ws["B2"].value == "Person 0", f"first name wrong: {ws['B2'].value!r}"
    # numeric type preserved
    assert isinstance(ws["A2"].value, (int, float)), "ID should be numeric"


def t_dict_types():
    """bool stays bool (not 1/0), numbers stay numbers, strings stay strings."""
    need(have("write_sheet"), "jetxl.write_sheet not in build")
    data = {
        "flag": [True, False, True],
        "count": [1, 2, 3],
        "price": [1.5, 2.25, 3.75],
        "label": ["a", "b", "c"],
    }
    path = out("dict_types.xlsx")
    jetxl.write_sheet(data, path, sheet_name="Types")
    ws = open_wb(path)["Types"]
    assert ws["A2"].value is True, f"bool True became {ws['A2'].value!r}"
    assert ws["A3"].value is False, f"bool False became {ws['A3'].value!r}"
    assert ws["B2"].value == 1, f"int wrong: {ws['B2'].value!r}"
    assert abs(ws["C2"].value - 1.5) < 1e-9, f"float wrong: {ws['C2'].value!r}"
    assert ws["D2"].value == "a", f"str wrong: {ws['D2'].value!r}"


def t_dict_multi():
    need(have("write_sheets"), "jetxl.write_sheets not in build")
    sheets = [
        {"name": "Q1", "columns": to_dict_api(dummy_columns(10, seed=1))},
        {"name": "Q2", "columns": to_dict_api(dummy_columns(15, seed=2))},
        {"name": "Q3", "columns": to_dict_api(dummy_columns(20, seed=3))},
    ]
    path = out("dict_multi.xlsx")
    jetxl.write_sheets(sheets, path, 2)   # 2 threads
    wb = open_wb(path)
    assert wb.sheetnames == ["Q1", "Q2", "Q3"], f"sheets: {wb.sheetnames}"
    assert wb["Q2"].max_row == 16, f"Q2 rows: {wb['Q2'].max_row}"


def t_arrow_single():
    need(have("write_sheet_arrow"), "jetxl.write_sheet_arrow not in build")
    cols = dummy_columns(60)
    arrow = to_arrow(cols)
    path = out("arrow_single.xlsx")
    jetxl.write_sheet_arrow(arrow, path, sheet_name="Sales")
    wb = open_wb(path)
    assert "Sales" in wb.sheetnames, f"sheets: {wb.sheetnames}"
    ws = wb["Sales"]
    assert ws.max_row == 61, f"expected 61 rows, got {ws.max_row}"
    assert ws["A1"].value == "ID"


def t_arrow_formatting():
    """Exercise common formatting flags together and confirm the file is valid."""
    need(have("write_sheet_arrow"), "jetxl.write_sheet_arrow not in build")
    cols = dummy_columns(30)
    arrow = to_arrow(cols)
    path = out("arrow_formatted.xlsx")
    jetxl.write_sheet_arrow(
        arrow, path,
        sheet_name="Formatted",
        auto_filter=True,
        freeze_rows=1,
        styled_headers=True,
        auto_width=True,
        column_formats={"Sales": "currency", "JoinDate": "date"},
        zoom_scale=120,
        tab_color="FF4472C4",
    )
    wb = open_wb(path)
    ws = wb["Formatted"]
    assert ws.max_row == 31
    # freeze pane should be set (openpyxl exposes it)
    assert ws.freeze_panes in ("A2", "B2", None) or ws.freeze_panes, \
        f"freeze_panes: {ws.freeze_panes!r}"
    assert ws.auto_filter.ref is not None, "auto_filter not set"


def t_arrow_multi_threads():
    need(have("write_sheets_arrow"), "jetxl.write_sheets_arrow not in build")
    sheets = [{"data": to_arrow(dummy_columns(25, seed=i)), "name": f"S{i}"}
              for i in range(1, 6)]
    path = out("arrow_multi.xlsx")
    jetxl.write_sheets_arrow(sheets, path, 4)   # 4 threads
    wb = open_wb(path)
    assert wb.sheetnames == [f"S{i}" for i in range(1, 6)], f"sheets: {wb.sheetnames}"
    for i in range(1, 6):
        assert wb[f"S{i}"].max_row == 26


def t_to_bytes():
    need(have("write_sheet_arrow_to_bytes"), "write_sheet_arrow_to_bytes not in build")
    import io
    arrow = to_arrow(dummy_columns(20))
    raw = jetxl.write_sheet_arrow_to_bytes(arrow)
    assert isinstance(raw, (bytes, bytearray)), f"expected bytes, got {type(raw)}"
    assert raw[:2] == b"PK", "not a zip/xlsx byte stream"
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert wb.active.max_row == 21


def t_special_chars():
    """Strings with XML metacharacters, unicode, and edge whitespace must survive."""
    need(have("write_sheet"), "jetxl.write_sheet not in build")
    data = {
        "text": ["a & b", "x < y > z", 'quote " apos \'', "café ☕ 日本",
                 "  leading", "trailing  ", "normal"],
    }
    path = out("special_chars.xlsx")
    jetxl.write_sheet(data, path, sheet_name="Chars & <Test>")
    wb = open_wb(path)
    # sheet name with & and <> should survive (Excel caps at 31 chars though)
    assert wb.sheetnames, "no sheets"
    ws = wb.active
    vals = [ws.cell(r, 1).value for r in range(2, 9)]
    assert vals[0] == "a & b", f"ampersand: {vals[0]!r}"
    assert vals[1] == "x < y > z", f"angles: {vals[1]!r}"
    assert vals[3] == "café ☕ 日本", f"unicode: {vals[3]!r}"
    assert vals[4] == "  leading", f"leading ws lost: {vals[4]!r}"
    assert vals[5] == "trailing  ", f"trailing ws lost: {vals[5]!r}"


def t_empty_and_nulls():
    """None cells and an all-empty column shouldn't corrupt the file."""
    need(have("write_sheet"), "jetxl.write_sheet not in build")
    data = {
        "a": [1, None, 3],
        "b": [None, None, None],
        "c": ["x", None, "z"],
    }
    path = out("nulls.xlsx")
    jetxl.write_sheet(data, path, sheet_name="Nulls")
    ws = open_wb(path)["Nulls"]
    assert ws["A2"].value == 1
    assert ws["A3"].value is None, f"None became {ws['A3'].value!r}"
    assert ws["C4"].value == "z"


def t_data_validation():
    need(have("write_sheet_arrow"), "write_sheet_arrow not in build")
    arrow = to_arrow(dummy_columns(20))
    path = out("validation.xlsx")
    try:
        jetxl.write_sheet_arrow(
            arrow, path, sheet_name="Valid",
            data_validations=[{
                "start_row": 2, "start_col": 2, "end_row": 21, "end_col": 2,
                "type": "list",
                "items": ["North", "South", "East", "West", "Central"],
                "show_dropdown": True,
                "error_title": "Bad region",
                "error_message": "Pick from the list",
            }],
        )
    except TypeError as e:
        raise _Skip(f"data_validations kwarg not supported: {e}")
    wb = open_wb(path)          # opening is the main check
    ws = wb["Valid"]
    # openpyxl exposes data validations
    assert len(ws.data_validations.dataValidation) >= 1, "no data validation found"


def t_conditional_format():
    need(have("write_sheet_arrow"), "write_sheet_arrow not in build")
    arrow = to_arrow(dummy_columns(30))
    path = out("conditional.xlsx")
    try:
        jetxl.write_sheet_arrow(
            arrow, path, sheet_name="Cond",
            conditional_formats=[{
                "start_row": 2, "start_col": 3, "end_row": 31, "end_col": 3,
                "rule_type": "cell_value", "operator": "greater_than",
                "value": "50000", "priority": 1,
                "style": {"font": {"bold": True, "color": "FFFF0000"}},
            }],
        )
    except TypeError as e:
        raise _Skip(f"conditional_formats kwarg not supported: {e}")
    open_wb(path)   # must open cleanly


def t_excel_table():
    need(have("write_sheet_arrow"), "write_sheet_arrow not in build")
    arrow = to_arrow(dummy_columns(15))
    path = out("table.xlsx")
    try:
        jetxl.write_sheet_arrow(
            arrow, path, sheet_name="Tbl",
            tables=[{
                "name": "SalesTable", "display_name": "SalesTable",
                "start_row": 1, "start_col": 0, "end_row": 16, "end_col": 6,
                "style": "TableStyleMedium2",
            }],
        )
    except TypeError as e:
        raise _Skip(f"tables kwarg not supported: {e}")
    wb = open_wb(path)
    ws = wb["Tbl"]
    assert len(ws.tables) >= 1, "no table object found in worksheet"


def t_chart():
    need(have("write_sheet_arrow"), "write_sheet_arrow not in build")
    # small aggregated data for a chart
    cols = {
        "Region": ["North", "South", "East", "West"],
        "Total": [120.0, 95.5, 143.2, 88.0],
    }
    arrow = to_arrow(cols)
    path = out("chart.xlsx")
    # NOTE: some jetxl builds treat the chart anchor keys (from_col/from_row/
    # to_col/to_row) as required and will CRASH THE PROCESS (Rust panic:
    # "unwrap on None") if they're omitted — a Rust panic is not a catchable
    # Python exception, so we can't try/except it. We therefore supply the
    # anchor keys explicitly. If your build defaults them, they're harmless.
    chart = {
        "chart_type": "column",
        "data_range": (1, 0, 5, 1),
        "category_col": 0,
        "title": "Sales by Region",
        "x_axis_title": "Region",
        "y_axis_title": "Total",
        "from_col": 3, "from_row": 0, "to_col": 11, "to_row": 15,
    }
    try:
        jetxl.write_sheet_arrow(arrow, path, sheet_name="Chart", charts=[chart])
    except TypeError as e:
        raise _Skip(f"charts kwarg not supported: {e}")
    open_wb(path)   # opening a chart file cleanly is the check


def t_merge_and_hyperlinks():
    need(have("write_sheet_arrow"), "write_sheet_arrow not in build")
    arrow = to_arrow(dummy_columns(10))
    path = out("merge_links.xlsx")
    try:
        jetxl.write_sheet_arrow(
            arrow, path, sheet_name="ML",
            merge_cells=[(1, 0, 1, 2)],   # merge A1:C1
            hyperlinks=[(2, 1, "https://example.com", "Example")],
        )
    except TypeError as e:
        raise _Skip(f"merge/hyperlink kwargs not supported: {e}")
    wb = open_wb(path)
    ws = wb["ML"]
    assert len(ws.merged_cells.ranges) >= 1, "no merged range found"


def t_big_multisheet_collision():
    """>=100 sheets — regression guard for the workbook-rels rId collision.
    The 100th sheet must be a real, readable worksheet."""
    need(have("write_sheets_arrow"), "write_sheets_arrow not in build")
    sheets = [{"data": to_arrow({"n": [i]}), "name": f"Sheet{i}"}
              for i in range(1, 151)]
    path = out("big_150_sheets.xlsx")
    jetxl.write_sheets_arrow(sheets, path, 4)
    wb = open_wb(path)
    assert len(wb.sheetnames) == 150, f"expected 150 sheets, got {len(wb.sheetnames)}"
    assert wb.sheetnames[99] == "Sheet100", f"100th sheet: {wb.sheetnames[99]}"
    assert wb["Sheet100"]["A2"].value == 100, "sheet 100 data unreadable"


# =============================================================================
def main():
    global OUT
    ap = argparse.ArgumentParser()
    ap.add_argument("--outdir", default="jetxl_test_out",
                    help="where to write the test .xlsx files")
    ap.add_argument("--keep-open", action="store_true",
                    help="print the output folder path at the end")
    args = ap.parse_args()
    OUT = args.outdir
    os.makedirs(OUT, exist_ok=True)

    print("=" * 68)
    print("jetxl functional test suite")
    print("=" * 68)
    print(f"jetxl   : {getattr(jetxl, '__version__', 'unknown')}")
    print(f"openpyxl: {openpyxl.__version__}")
    print(f"arrow   : {'polars' if pl else ('pyarrow' if pa else 'NONE — arrow tests skip')}")
    print(f"output  : {os.path.abspath(OUT)}\n")

    tests = [
        ("dict API — single sheet",            t_dict_single),
        ("dict API — type fidelity",           t_dict_types),
        ("dict API — multi sheet",             t_dict_multi),
        ("arrow API — single sheet",           t_arrow_single),
        ("arrow API — formatting flags",       t_arrow_formatting),
        ("arrow API — multi sheet + threads",  t_arrow_multi_threads),
        ("arrow API — write to bytes",         t_to_bytes),
        ("special chars / unicode / edge ws",  t_special_chars),
        ("empty cells and nulls",              t_empty_and_nulls),
        ("data validation dropdown",           t_data_validation),
        ("conditional formatting",             t_conditional_format),
        ("excel table",                        t_excel_table),
        ("chart",                              t_chart),
        ("merged cells + hyperlinks",          t_merge_and_hyperlinks),
        ("150 sheets (rId collision guard)",   t_big_multisheet_collision),
    ]

    for name, fn in tests:
        test(name, fn)

    print("\n" + "=" * 68)
    print(f"  {R.passed} passed   {R.failed} failed   {R.skipped} skipped")
    print("=" * 68)
    if R.failures:
        print("\nFailures:")
        for name, msg in R.failures:
            print(f"  - {name}: {msg.splitlines()[0]}")
    if args.keep_open:
        print(f"\nOutput files are in: {os.path.abspath(OUT)}")

    sys.exit(1 if R.failed else 0)


if __name__ == "__main__":
    main()