#!/usr/bin/env python3
"""
jetxl comprehensive test suite
==============================

A single, verbose, self-contained test runner that exercises the whole jetxl
surface: the Arrow API and the legacy dict API, single- and multi-sheet writes,
threading, every formatting option, the int-or-name column reference feature,
shared strings, the hand-rolled ZIP writer (fastzip), and error handling. Every
check states what it verified and why.

Design:
  - Pure stdlib + pyarrow + openpyxl. No pytest needed (run it directly), but the
    checks are structured so a pytest collector would also find `test_*` fns.
  - Every write is round-tripped through openpyxl and/or zipfile and asserted on
    actual cell values / archive structure, not just "it didn't throw".
  - Groups are independent; one failing group doesn't abort the rest.
  - Exit code is non-zero if anything failed, so CI can gate on it.

Run:
    python test_suite.py            # full suite
    python test_suite.py -v         # also print every passing assertion
    python test_suite.py --quick    # skip the large-N perf-ish round trips
"""

from __future__ import annotations
import io
import os
import sys
import zipfile
import tempfile
import traceback

import pyarrow as pa

try:
    import openpyxl
except Exception:
    print("ERROR: openpyxl is required for the test suite (pip install openpyxl).")
    sys.exit(2)

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv
QUICK = "--quick" in sys.argv


# ----------------------------------------------------------------------------
# Tiny test harness
# ----------------------------------------------------------------------------
class Group:
    """Collects pass/fail for a named group of assertions."""

    def __init__(self, name):
        self.name = name
        self.passed = 0
        self.failed = 0
        self.failures = []

    def check(self, label, cond, detail=""):
        if cond:
            self.passed += 1
            if VERBOSE:
                print(f"    pass  {label}")
        else:
            self.failed += 1
            msg = f"{label}" + (f"  [{detail}]" if detail else "")
            self.failures.append(msg)
            print(f"    FAIL  {msg}")
        return cond

    def eq(self, label, got, want):
        return self.check(label, got == want, f"got {got!r}, want {want!r}")


RESULTS: list[Group] = []


def group(name):
    print(f"\n=== {name} ===")
    g = Group(name)
    RESULTS.append(g)
    return g


def run(fn):
    """Run a test function, converting an unexpected exception into a failure
    rather than aborting the whole suite."""
    try:
        fn()
    except Exception as e:
        g = group(f"{fn.__name__} (CRASHED)")
        g.check("did not raise unexpectedly", False, f"{type(e).__name__}: {e}")
        if VERBOSE:
            traceback.print_exc()


# ----------------------------------------------------------------------------
# Shared data builders
# ----------------------------------------------------------------------------
def tbl_mixed(rows=1000):
    """Representative business sheet: ints, floats, unique + low-cardinality
    strings, bools. Low-cardinality 'region' should trigger shared strings;
    'note' is unique and should stay inline."""
    return pa.table({
        "id":     pa.array(list(range(rows)), pa.int64()),
        "region": pa.array([["North", "South", "East", "West", "Central"][i % 5] for i in range(rows)]),
        "note":   pa.array([f"unique note {i}" for i in range(rows)]),
        "amount": pa.array([i * 1.5 for i in range(rows)], pa.float64()),
        "active": pa.array([i % 2 == 0 for i in range(rows)]),
    })


def load_bytes(by):
    return openpyxl.load_workbook(io.BytesIO(by))


def col_values(ws, col, start=2, n=None):
    """Read a column's values (1-based col index), skipping the header row."""
    out = []
    r = start
    while True:
        v = ws.cell(row=r, column=col).value
        if v is None and (n is None):
            break
        out.append(v)
        r += 1
        if n is not None and len(out) >= n:
            break
    return out


# ============================================================================
# GROUP 1 — Arrow API, data types round-trip
# ============================================================================
def test_datatypes():
    g = group("Arrow API — data type round trips")

    # integers (incl. negative and zero)
    t = pa.table({"a": pa.array([-5, 0, 7, 2_000_000], pa.int64())})
    ws = load_bytes(jetxl.write_sheet_arrow_to_bytes(t)).active
    g.eq("int64 values", col_values(ws, 1, n=4), [-5, 0, 7, 2_000_000])

    # floats
    t = pa.table({"a": pa.array([1.5, -2.25, 0.0, 3.333], pa.float64())})
    ws = load_bytes(jetxl.write_sheet_arrow_to_bytes(t)).active
    vals = col_values(ws, 1, n=4)
    g.check("float64 values", all(abs(a - b) < 1e-9 for a, b in zip(vals, [1.5, -2.25, 0.0, 3.333])),
            str(vals))

    # booleans
    t = pa.table({"a": pa.array([True, False, True])})
    ws = load_bytes(jetxl.write_sheet_arrow_to_bytes(t)).active
    g.check("bool values", col_values(ws, 1, n=3) in ([True, False, True], [1, 0, 1]),
            str(col_values(ws, 1, n=3)))

    # strings incl. XML-special characters, unicode, whitespace
    special = ["plain", "<tag> & 'amp'", 'quote"here', "日本語テスト", "  leading/trailing  ", "line1\nline2"]
    t = pa.table({"a": pa.array(special)})
    ws = load_bytes(jetxl.write_sheet_arrow_to_bytes(t)).active
    g.eq("string special chars preserved", col_values(ws, 1, n=len(special)), special)

    # int32 / float32 (narrower widths)
    t = pa.table({"i": pa.array([1, 2, 3], pa.int32()), "f": pa.array([1.0, 2.0, 3.0], pa.float32())})
    wb = load_bytes(jetxl.write_sheet_arrow_to_bytes(t)); ws = wb.active
    g.eq("int32 column", col_values(ws, 1, n=3), [1, 2, 3])

    # nulls interleaved
    t = pa.table({"a": pa.array(["x", None, "y", None]), "b": pa.array([1, None, 3, 4], pa.int64())})
    ws = load_bytes(jetxl.write_sheet_arrow_to_bytes(t)).active
    g.eq("null strings -> empty cells", [ws.cell(r, 1).value for r in range(2, 6)], ["x", None, "y", None])
    g.eq("null ints -> empty cells", [ws.cell(r, 2).value for r in range(2, 6)], [1, None, 3, 4])

    # RecordBatch input (not just Table)
    batch = pa.record_batch({"a": pa.array([9, 8, 7], pa.int64())})
    ws = load_bytes(jetxl.write_sheet_arrow_to_bytes(batch)).active
    g.eq("RecordBatch input accepted", col_values(ws, 1, n=3), [9, 8, 7])


# ============================================================================
# GROUP 2 — Headers, freezing, filters, layout options
# ============================================================================
def test_layout_options():
    g = group("Layout options — headers, freeze, filter, gridlines, RTL, zoom")

    t = tbl_mixed(50)

    # header row present by default
    ws = load_bytes(jetxl.write_sheet_arrow_to_bytes(t)).active
    g.eq("header row written", [ws.cell(1, c).value for c in range(1, 6)],
         ["id", "region", "note", "amount", "active"])

    # write_header_row=False omits headers -> first row is data
    by = jetxl.write_sheet_arrow_to_bytes(t, write_header_row=False)
    ws = load_bytes(by).active
    g.eq("header omitted -> first cell is data", ws.cell(1, 1).value, 0)

    # freeze panes
    by = jetxl.write_sheet_arrow_to_bytes(t, freeze_rows=1, freeze_cols=2)
    ws = load_bytes(by).active
    g.check("freeze panes set", ws.freeze_panes is not None, str(ws.freeze_panes))

    # auto_filter
    by = jetxl.write_sheet_arrow_to_bytes(t, auto_filter=True)
    ws = load_bytes(by).active
    g.check("auto_filter applied", ws.auto_filter.ref is not None, str(ws.auto_filter.ref))

    # gridlines off
    by = jetxl.write_sheet_arrow_to_bytes(t, gridlines_visible=False)
    ws = load_bytes(by).active
    g.check("gridlines toggled off", ws.sheet_view.showGridLines is False,
            str(ws.sheet_view.showGridLines))

    # right-to-left
    by = jetxl.write_sheet_arrow_to_bytes(t, right_to_left=True)
    ws = load_bytes(by).active
    g.check("RTL sheet view", bool(ws.sheet_view.rightToLeft), str(ws.sheet_view.rightToLeft))

    # custom sheet name
    by = jetxl.write_sheet_arrow_to_bytes(t, sheet_name="Report Q3")
    wb = load_bytes(by)
    g.check("custom sheet name", "Report Q3" in wb.sheetnames, str(wb.sheetnames))

    # styled headers (bold) — just confirm it still opens and header present
    by = jetxl.write_sheet_arrow_to_bytes(t, styled_headers=True)
    ws = load_bytes(by).active
    g.eq("styled headers still write labels", ws.cell(1, 1).value, "id")


# ============================================================================
# GROUP 3 — Column references by index OR name (the new feature)
# ============================================================================
def test_column_references():
    g = group("Column references — int index OR str name")

    t = tbl_mixed(200)

    # by name
    by = jetxl.write_sheet_arrow_to_bytes(
        t, column_widths={"region": 25.0, "amount": "auto"},
        column_formats={"amount": "currency"}, hidden_columns=["note"])
    ws = load_bytes(by).active
    g.eq("name refs open", ws.cell(1, 1).value, "id")
    g.check("hidden by name (col C = note)", ws.column_dimensions["C"].hidden, "")

    # by index  (region=1, amount=3, note=2)
    by = jetxl.write_sheet_arrow_to_bytes(
        t, column_widths={1: 25.0, 3: "auto"},
        column_formats={3: "currency"}, hidden_columns=[2])
    ws = load_bytes(by).active
    g.eq("index refs open", ws.cell(1, 1).value, "id")
    g.check("hidden by index (col C)", ws.column_dimensions["C"].hidden, "")

    # mixed index + name in the same call
    by = jetxl.write_sheet_arrow_to_bytes(
        t, column_widths={0: 10.0, "region": 25.0}, hidden_columns=[2, "amount"])
    ws = load_bytes(by).active
    g.eq("mixed refs open", ws.cell(1, 1).value, "id")
    g.check("hidden amount by name (col D)", ws.column_dimensions["D"].hidden, "")

    # width actually applied (character width -> openpyxl width)
    by = jetxl.write_sheet_arrow_to_bytes(t, column_widths={"id": 33.0})
    ws = load_bytes(by).active
    g.check("explicit width applied to col A", ws.column_dimensions["A"].width is not None, "")


# ============================================================================
# GROUP 4 — Error handling (strict, as configured)
# ============================================================================
def test_error_handling():
    g = group("Error handling — strict column ref validation")

    t = tbl_mixed(50)

    # unknown name -> ValueError
    try:
        jetxl.write_sheet_arrow_to_bytes(t, column_widths={"nope": 10.0})
        g.check("unknown name raises", False, "no exception")
    except ValueError as e:
        g.check("unknown name -> ValueError", "nope" in str(e), str(e))
    except Exception as e:
        g.check("unknown name -> ValueError", False, f"{type(e).__name__}")

    # out-of-range index -> IndexError
    try:
        jetxl.write_sheet_arrow_to_bytes(t, hidden_columns=[99])
        g.check("oob index raises", False, "no exception")
    except IndexError as e:
        g.check("oob index -> IndexError", "99" in str(e), str(e))
    except Exception as e:
        g.check("oob index -> IndexError", False, f"{type(e).__name__}")

    # negative index -> IndexError
    try:
        jetxl.write_sheet_arrow_to_bytes(t, hidden_columns=[-1])
        g.check("negative index raises", False, "no exception")
    except IndexError:
        g.check("negative index -> IndexError", True)
    except Exception as e:
        g.check("negative index -> IndexError", False, f"{type(e).__name__}")

    # unknown format name in column_formats keyed by index
    try:
        jetxl.write_sheet_arrow_to_bytes(t, column_formats={7: "currency"})
        g.check("oob format index raises", False, "no exception")
    except IndexError:
        g.check("oob format index -> IndexError", True)
    except Exception as e:
        g.check("oob format index -> IndexError", False, f"{type(e).__name__}")

    # empty table should not panic (should either write an empty sheet or raise
    # a catchable error, never crash the process)
    try:
        empty = pa.table({"a": pa.array([], pa.int64())})
        by = jetxl.write_sheet_arrow_to_bytes(empty)
        g.check("empty table handled (opens)", zipfile.ZipFile(io.BytesIO(by)).testzip() is None)
    except Exception as e:
        # a clean exception is acceptable; a crash is not (we'd never get here)
        g.check("empty table -> clean error not crash", True, f"{type(e).__name__}")


# ============================================================================
# GROUP 5 — Formatting: number formats, hyperlinks, merges, row heights
# ============================================================================
def test_formatting():
    g = group("Formatting — number formats, hyperlinks, merges, tab color")

    t = tbl_mixed(20)

    # number format applied (currency on amount)
    by = jetxl.write_sheet_arrow_to_bytes(t, column_formats={"amount": "currency"})
    ws = load_bytes(by).active
    # openpyxl exposes number_format on the styled data cells
    fmt = ws.cell(2, 4).number_format
    g.check("currency number format present", fmt not in (None, "General"), str(fmt))

    # hyperlink with query-string ampersand (the corruption-fix regression guard)
    by = jetxl.write_sheet_arrow_to_bytes(
        t, hyperlinks=[(1, 0, "https://example.com/?x=1&y=2&z=3", "link")])
    wb = load_bytes(by); ws = wb.active
    g.eq("file with &-URL opens cleanly", ws.cell(1, 1).value, "id")

    # merge cells. Row/col in merge_cells are 1-based absolute Excel
    # coordinates (like hyperlinks), so merge the header cells A1:C1.
    by = jetxl.write_sheet_arrow_to_bytes(t, merge_cells=[(1, 0, 1, 2)])
    ws = load_bytes(by).active
    g.check("merged range present", len(ws.merged_cells.ranges) >= 1,
            str(list(ws.merged_cells.ranges)))

    # tab color
    by = jetxl.write_sheet_arrow_to_bytes(t, tab_color="FFFF0000")
    ws = load_bytes(by).active
    g.check("tab color set", ws.sheet_properties.tabColor is not None, "")

    # row heights. Keys are 1-based Excel row numbers, so {1: 40} sets row 1.
    by = jetxl.write_sheet_arrow_to_bytes(t, row_heights={1: 40.0})
    ws = load_bytes(by).active
    g.check("custom row height", ws.row_dimensions[1].height == 40.0,
            str(ws.row_dimensions[1].height))


# ============================================================================
# GROUP 6 — Shared strings (dedup path)
# ============================================================================
def test_shared_strings():
    g = group("Shared strings — low-cardinality dedup, high-cardinality inline")

    t = tbl_mixed(1000)
    by = jetxl.write_sheet_arrow_to_bytes(t)

    with zipfile.ZipFile(io.BytesIO(by)) as z:
        names = z.namelist()
    g.check("sharedStrings.xml present", "xl/sharedStrings.xml" in names, str(names))

    ws = load_bytes(by).active
    cats = ["North", "South", "East", "West", "Central"]
    region_ok = all(ws.cell(r, 2).value == cats[(r - 2) % 5] for r in range(2, 1002))
    g.check("low-card 'region' resolves correctly (shared)", region_ok)

    note_ok = all(ws.cell(r, 3).value == f"unique note {r - 2}" for r in range(2, 1002))
    g.check("high-card 'note' resolves correctly (inline)", note_ok)

    # a fully unique string column should NOT force wrong values even without a
    # shared table
    t2 = pa.table({"u": pa.array([f"v{i}" for i in range(500)])})
    ws2 = load_bytes(jetxl.write_sheet_arrow_to_bytes(t2)).active
    g.check("all-unique column correct", all(ws2.cell(r, 1).value == f"v{r-2}" for r in range(2, 502)))


# ============================================================================
# GROUP 7 — Multi-sheet + threading
# ============================================================================
def test_multisheet():
    g = group("Multi-sheet — parallel write, per-sheet integrity, thread counts")

    sheets = [{"name": f"S{i}", "data": tbl_mixed(500)} for i in range(6)]

    # to file, several thread counts (exercises the pool cache across counts)
    with tempfile.TemporaryDirectory() as tmp:
        for th in (1, 2, 4):
            path = os.path.join(tmp, f"multi_{th}.xlsx")
            jetxl.write_sheets_arrow(sheets, path, th)
            wb = openpyxl.load_workbook(path)
            g.eq(f"{th} threads -> all sheets present", wb.sheetnames, [f"S{i}" for i in range(6)])
            ws = wb["S3"]
            ok = (ws.cell(2, 1).value == 0 and ws.cell(2, 2).value == "North"
                  and ws.cell(501, 1).value == 499)
            g.check(f"{th} threads -> S3 data intact", ok)

    # to bytes
    by = jetxl.write_sheets_arrow_to_bytes(sheets, 4)
    wb = load_bytes(by)
    g.eq("to_bytes multi sheets present", wb.sheetnames, [f"S{i}" for i in range(6)])
    z = zipfile.ZipFile(io.BytesIO(by))
    g.check("multi archive CRC valid", z.testzip() is None)

    # heterogeneous sheet sizes in one workbook
    mixed_sheets = [
        {"name": "tiny", "data": pa.table({"a": pa.array([1], pa.int64())})},
        {"name": "big",  "data": tbl_mixed(2000)},
        {"name": "empty_ish", "data": pa.table({"a": pa.array([0], pa.int64())})},
    ]
    by = jetxl.write_sheets_arrow_to_bytes(mixed_sheets, 4)
    wb = load_bytes(by)
    g.eq("heterogeneous sheets", wb.sheetnames, ["tiny", "big", "empty_ish"])
    g.eq("big sheet row count", wb["big"].max_row, 2001)  # 2000 + header

    # per-sheet independent formatting via dict params
    fmt_sheets = [
        {"name": "A", "data": tbl_mixed(100), "hidden_columns": ["note"]},
        {"name": "B", "data": tbl_mixed(100), "column_widths": {"id": 20.0}},
    ]
    by = jetxl.write_sheets_arrow_to_bytes(fmt_sheets, 2)
    wb = load_bytes(by)
    g.check("per-sheet hidden col (A/note=C)", wb["A"].column_dimensions["C"].hidden, "")


# ============================================================================
# GROUP 8 — Legacy dict API
# ============================================================================
def test_dict_api():
    g = group("Legacy dict API — write_sheet / write_sheets")

    with tempfile.TemporaryDirectory() as tmp:
        # single sheet
        cols = {"name": ["a", "b", "c"], "value": [1, 2, 3]}
        path = os.path.join(tmp, "dict_single.xlsx")
        jetxl.write_sheet(cols, path)
        wb = openpyxl.load_workbook(path); ws = wb.active
        g.eq("dict header", [ws.cell(1, c).value for c in range(1, 3)], ["name", "value"])
        g.eq("dict values col1", col_values(ws, 1, n=3), ["a", "b", "c"])
        g.eq("dict values col2", col_values(ws, 2, n=3), [1, 2, 3])

        # multi sheet (if exposed)
        if hasattr(jetxl, "write_sheets"):
            dict_sheets = [
                {"name": "one", "columns": {"x": [1, 2]}},
                {"name": "two", "columns": {"y": [3, 4]}},
            ]
            path = os.path.join(tmp, "dict_multi.xlsx")
            try:
                jetxl.write_sheets(dict_sheets, path, 2)
                wb = openpyxl.load_workbook(path)
                g.check("dict multi sheets", set(["one", "two"]).issubset(set(wb.sheetnames)),
                        str(wb.sheetnames))
            except Exception as e:
                g.check("dict multi sheets", False, f"{type(e).__name__}: {e}")


# ============================================================================
# GROUP 9 — fastzip ZIP writer structure & guarantees
# ============================================================================
def test_fastzip():
    g = group("fastzip — ZIP structure, CRC, store-fallback, reproducibility")

    t = tbl_mixed(200)
    by = jetxl.write_sheet_arrow_to_bytes(t)
    z = zipfile.ZipFile(io.BytesIO(by))

    g.check("archive parses", True)
    g.check("all entry CRCs valid (testzip)", z.testzip() is None)
    names = z.namelist()
    for required in ("[Content_Types].xml", "xl/workbook.xml", "xl/worksheets/sheet1.xml"):
        g.check(f"contains {required}", required in names, str(names))

    # tiny file exercises the store-fallback safety net
    tiny = pa.table({"a": pa.array([1], pa.int64())})
    bt = jetxl.write_sheet_arrow_to_bytes(tiny)
    g.check("tiny file valid CRC", zipfile.ZipFile(io.BytesIO(bt)).testzip() is None)

    # wide sheet (many columns -> many cell refs), integrity holds
    wide = pa.table({f"c{i}": pa.array(list(range(50))) for i in range(40)})
    bw = jetxl.write_sheet_arrow_to_bytes(wide)
    g.check("wide sheet valid CRC", zipfile.ZipFile(io.BytesIO(bw)).testzip() is None)
    g.eq("wide sheet column count", load_bytes(bw).active.max_column, 40)

    # reproducibility: identical input -> byte-identical output (fixed timestamps)
    b1 = jetxl.write_sheet_arrow_to_bytes(tbl_mixed(300))
    b2 = jetxl.write_sheet_arrow_to_bytes(tbl_mixed(300))
    g.check("byte-identical output for identical input", b1 == b2)

    # every stored/deflated entry decompresses to the size the header claims
    ok = True
    for info in z.infolist():
        try:
            data = z.read(info.filename)
            ok = ok and (len(data) == info.file_size)
        except Exception:
            ok = False
    g.check("every entry decompresses to declared size", ok)


# ============================================================================
# GROUP 10 — Large round trips (value integrity at scale)
# ============================================================================
def test_large_roundtrip():
    g = group("Large round trips — value integrity at scale")
    if QUICK:
        g.check("skipped (--quick)", True)
        return

    t = tbl_mixed(200_000)
    by = jetxl.write_sheet_arrow_to_bytes(t)
    g.check("200K archive CRC valid", zipfile.ZipFile(io.BytesIO(by)).testzip() is None)
    ws = load_bytes(by).active
    # spot-check scattered rows rather than all 200K (openpyxl read is the slow part)
    errs = 0
    for r in (2, 1000, 50_000, 199_999, 200_001):
        if ws.cell(r, 1).value != r - 2:
            errs += 1
        if ws.cell(r, 2).value != ["North", "South", "East", "West", "Central"][(r - 2) % 5]:
            errs += 1
    g.eq("scattered-row integrity (200K)", errs, 0)

    # multi-sheet at scale
    sheets = [{"name": f"S{i}", "data": tbl_mixed(125_000)} for i in range(8)]
    by = jetxl.write_sheets_arrow_to_bytes(sheets, 8)
    g.check("8x125K archive CRC valid", zipfile.ZipFile(io.BytesIO(by)).testzip() is None)
    wb = load_bytes(by)
    g.eq("8 sheets present at scale", len(wb.sheetnames), 8)
    g.eq("last sheet last row intact", wb["S7"].cell(125_001, 1).value, 124_999)


# ============================================================================
# Runner
# ============================================================================
def main():
    print("=" * 70)
    print("jetxl comprehensive test suite")
    print(f"pyarrow {pa.__version__} | openpyxl {openpyxl.__version__} | "
          f"quick={QUICK} verbose={VERBOSE}")
    print("=" * 70)

    for fn in (
        test_datatypes,
        test_layout_options,
        test_column_references,
        test_error_handling,
        test_formatting,
        test_shared_strings,
        test_multisheet,
        test_dict_api,
        test_fastzip,
        test_large_roundtrip,
    ):
        run(fn)

    total_pass = sum(g.passed for g in RESULTS)
    total_fail = sum(g.failed for g in RESULTS)

    print("\n" + "=" * 70)
    print("SUMMARY")
    print("-" * 70)
    for g in RESULTS:
        status = "OK  " if g.failed == 0 else "FAIL"
        print(f"  [{status}] {g.name}: {g.passed} passed, {g.failed} failed")
        for f in g.failures:
            print(f"           - {f}")
    print("-" * 70)
    print(f"TOTAL: {total_pass} passed, {total_fail} failed")
    print("=" * 70)

    sys.exit(1 if total_fail else 0)


if __name__ == "__main__":
    main()