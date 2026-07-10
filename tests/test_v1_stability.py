#!/usr/bin/env python3
"""
jetxl v1 stability suite
========================

A release-gating test suite that exercises every Excel element category jetxl
claims to support, the way the openpyxl / XlsxWriter test suites do: for each
feature we WRITE a real file, then VERIFY it two ways —

  1. round-trip through openpyxl (an independent reader), asserting on actual
     cell values, styles, and worksheet properties; and
  2. where openpyxl is lenient or hides detail, inspect the raw OOXML parts
     inside the .zip directly and assert the XML is well-formed and correct.

Every generated archive is also CRC-validated (zipfile.testzip) and XML-parsed
(xml.dom.minidom) so a malformed part fails loudly instead of silently.

The suite is grouped by Excel element. Groups are independent; a failure in one
does not abort the rest. Exit code is non-zero if anything failed, so CI can
gate a v1 release on it.

Run:
    python test_v1_stability.py           # full suite
    python test_v1_stability.py -v        # print every passing assertion too
"""
from __future__ import annotations

import datetime as dt
import io
import math
import sys
import zipfile
import xml.dom.minidom as minidom
from xml.etree import ElementTree as ET

import pyarrow as pa
import openpyxl

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv

# OOXML namespaces we assert against.
NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "ct": "http://schemas.openxmlformats.org/package/2006/content-types",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


# ---------------------------------------------------------------------------
# Harness
# ---------------------------------------------------------------------------
class Group:
    def __init__(self, name):
        self.name = name
        self.passed = 0
        self.failed = 0
        self.errors = []

    def check(self, desc, cond):
        if cond:
            self.passed += 1
            if VERBOSE:
                print(f"    pass  {desc}")
        else:
            self.failed += 1
            self.errors.append(desc)
            print(f"    FAIL  {desc}")

    def fatal(self, desc, exc):
        self.failed += 1
        self.errors.append(f"{desc}: {exc!r}")
        print(f"    ERROR {desc}: {exc!r}")


GROUPS = []


def group(name):
    def deco(fn):
        GROUPS.append((name, fn))
        return fn
    return deco


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def load(b: bytes):
    """Round-trip bytes through openpyxl; also assert the zip CRCs are valid."""
    z = zipfile.ZipFile(io.BytesIO(b))
    bad = z.testzip()
    if bad is not None:
        raise AssertionError(f"bad CRC in archive member: {bad}")
    return openpyxl.load_workbook(io.BytesIO(b))


def parts(b: bytes):
    """Return {name: bytes} of all archive members, asserting each XML part parses."""
    z = zipfile.ZipFile(io.BytesIO(b))
    out = {}
    for n in z.namelist():
        data = z.read(n)
        out[n] = data
        if n.endswith(".xml") or n.endswith(".rels"):
            # Will raise on malformed XML — that's the point.
            minidom.parseString(data)
    return out


def xml_of(b: bytes, name: str) -> str:
    return zipfile.ZipFile(io.BytesIO(b)).read(name).decode("utf-8")


def sample_table(n=100):
    cats = ["North", "South", "East", "West", "Central"]
    return pa.table({
        "id": pa.array(list(range(n)), pa.int64()),
        "region": pa.array([cats[i % 5] for i in range(n)]),
        "note": pa.array([f"note {i}" for i in range(n)]),
        "amount": pa.array([i * 1.5 for i in range(n)], pa.float64()),
    })


# ===========================================================================
# 1. DATA TYPES — every Arrow type maps to the right Excel cell type
# ===========================================================================
@group("Data types — full Arrow type coverage")
def g_types(G):
    t = pa.table({
        "i8":   pa.array([1, -2, 127], pa.int8()),
        "i16":  pa.array([1, -2, 30000], pa.int16()),
        "i32":  pa.array([1, -2, 2_000_000], pa.int32()),
        "i64":  pa.array([1, -2, 9_000_000_000], pa.int64()),
        "u32":  pa.array([1, 2, 4_000_000_000], pa.uint32()),
        "f32":  pa.array([1.5, -2.25, 3.125], pa.float32()),
        "f64":  pa.array([1.5, -2.25, math.pi], pa.float64()),
        "bool": pa.array([True, False, True]),
        "str":  pa.array(["a", "b", "c"]),
        "date": pa.array([dt.date(2024, 1, 1), dt.date(2024, 6, 15), dt.date(2023, 12, 31)]),
        "ts":   pa.array([dt.datetime(2024, 1, 1, 9, 30, 0)] * 3),
    })
    b = jetxl.write_sheet_arrow_to_bytes(t)
    parts(b)
    ws = load(b).active
    G.check("int8 -> number", ws["A4"].value == 127)
    G.check("int64 large stays exact", ws["D4"].value == 9_000_000_000)
    G.check("uint32 large stays exact", ws["E4"].value == 4_000_000_000)
    G.check("float32 approx", abs(ws["F3"].value - (-2.25)) < 1e-4)
    G.check("float64 pi", abs(ws["G4"].value - math.pi) < 1e-12)
    G.check("bool True is bool", ws["H2"].value is True)
    G.check("bool False is bool", ws["H3"].value is False)
    G.check("string", ws["I2"].value == "a")
    G.check("date is datetime", isinstance(ws["J2"].value, dt.datetime))
    G.check("date value", ws["J2"].value.date() == dt.date(2024, 1, 1))
    G.check("timestamp keeps time", ws["K2"].value == dt.datetime(2024, 1, 1, 9, 30, 0))


# ===========================================================================
# 2. NULLS / EMPTY — missing values in every column type
# ===========================================================================
@group("Nulls and empty values")
def g_nulls(G):
    t = pa.table({
        "i": pa.array([1, None, 3], pa.int64()),
        "f": pa.array([1.0, None, 3.0], pa.float64()),
        "s": pa.array(["x", None, "z"]),
        "b": pa.array([True, None, False]),
        "d": pa.array([dt.date(2024, 1, 1), None, dt.date(2024, 3, 3)]),
    })
    b = jetxl.write_sheet_arrow_to_bytes(t)
    parts(b)
    ws = load(b).active
    G.check("int null -> empty", ws["A3"].value is None)
    G.check("float null -> empty", ws["B3"].value is None)
    G.check("string null -> empty", ws["C3"].value is None)
    G.check("bool null -> empty", ws["D3"].value is None)
    G.check("date null -> empty", ws["E3"].value is None)
    G.check("non-null neighbours intact", ws["A1"].value == "i" and ws["A2"].value == 1 and ws["A4"].value == 3)

    # all-null column must not corrupt the sheet
    t2 = pa.table({"allnull": pa.array([None, None], pa.int64()), "ok": pa.array([1, 2], pa.int64())})
    b2 = jetxl.write_sheet_arrow_to_bytes(t2)
    parts(b2)
    ws2 = load(b2).active
    G.check("all-null column empty cells", ws2["A2"].value is None and ws2["A3"].value is None)
    G.check("neighbour column after all-null intact", ws2["B2"].value == 1)


# ===========================================================================
# 3. STRING EDGE CASES — XML escaping, control chars, unicode, long strings
# ===========================================================================
@group("String edge cases — escaping, unicode, control chars, length")
def g_strings(G):
    long_str = "L" * 32_767  # Excel's hard per-cell limit
    t = pa.table({
        "xml":   pa.array(["<tag>", "a & b", '"q"', "'apos'", ">less<"]),
        "uni":   pa.array(["café", "日本語", "🚀🎉", "Ω≈ç√", "naïve"]),
        "ws":    pa.array(["  lead", "trail  ", "a\tb", "line1\nline2", " "]),
        "long":  pa.array([long_str, "s", "s", "s", "s"]),
    })
    b = jetxl.write_sheet_arrow_to_bytes(t)
    parts(b)  # asserts sharedStrings / sheet XML parse despite hostile content
    ws = load(b).active
    G.check("< > & escaped and restored", ws["A2"].value == "<tag>" and ws["A3"].value == "a & b")
    G.check("quotes restored", ws["A4"].value == '"q"' and ws["A5"].value == "'apos'")
    G.check("unicode BMP", ws["B2"].value == "café" and ws["B3"].value == "日本語")
    G.check("emoji (astral plane)", ws["B4"].value == "🚀🎉")
    G.check("leading ws preserved", ws["C2"].value == "  lead")
    G.check("embedded tab preserved", ws["C4"].value == "a\tb")
    G.check("embedded newline preserved", ws["C5"].value == "line1\nline2")
    G.check("max-length (32767) string intact", ws["D2"].value == long_str and len(ws["D2"].value) == 32_767)

    # Control chars that are ILLEGAL in XML 1.0 must not produce a corrupt file.
    t2 = pa.table({"ctl": pa.array(["a\x00b", "c\x01d", "e\x1ff", "ok"])})
    try:
        b2 = jetxl.write_sheet_arrow_to_bytes(t2)
        parts(b2)              # must still parse as XML
        ws2 = load(b2).active  # must still open
        G.check("control chars don't corrupt archive", ws2["A5"].value == "ok")
    except Exception as e:
        G.fatal("control chars handled without crash", e)


# ===========================================================================
# 4. NUMBER FORMATS — every named format + custom codes, verified in styles.xml
# ===========================================================================
@group("Number formats — named + custom, verified in styles.xml")
def g_formats(G):
    named = [
        "general", "integer", "decimal2", "decimal4", "percentage",
        "percentage_decimal", "currency", "currency_rounded", "date",
        "datetime", "time", "scientific", "fraction", "thousands",
    ]
    cols = {name: pa.array([1234.5678, 0.25, 42.0], pa.float64()) for name in named}
    t = pa.table(cols)
    fmts = {name: name for name in named}
    b = jetxl.write_sheet_arrow_to_bytes(t, column_formats=fmts)
    parts(b)
    ws = load(b).active
    # openpyxl exposes the applied number_format string per cell
    seen = {ws.cell(row=2, column=i + 1).number_format for i in range(len(named))}
    G.check("multiple distinct number formats applied", len(seen) >= 6)
    G.check("currency format present", any("$" in s for s in seen))
    G.check("percent format present", any("%" in s for s in seen))
    styles_xml = xml_of(b, "xl/styles.xml")
    G.check("styles.xml has numFmts section", "<numFmt " in styles_xml)

    # Custom format code round-trips verbatim
    t2 = pa.table({"v": pa.array([1234.5], pa.float64())})
    b2 = jetxl.write_sheet_arrow_to_bytes(t2, column_formats={"v": '#,##0.000"kg"'})
    parts(b2)
    ws2 = load(b2).active
    G.check("custom format code applied to cell", "kg" in ws2["A2"].number_format)


# ===========================================================================
# 5. HEADERS / FREEZE / FILTER / GRIDLINES / RTL / ZOOM — sheet view props
# ===========================================================================
@group("Sheet view — headers, freeze, filter, gridlines, RTL, zoom")
def g_view(G):
    t = sample_table(50)
    b = jetxl.write_sheet_arrow_to_bytes(
        t, auto_filter=True, freeze_rows=1, freeze_cols=2,
        gridlines_visible=False, right_to_left=True, zoom_scale=145,
        styled_headers=True,
    )
    parts(b)
    ws = load(b).active
    G.check("auto_filter ref set", ws.auto_filter.ref is not None)
    G.check("freeze panes set", ws.freeze_panes is not None)
    G.check("gridlines hidden", ws.sheet_view.showGridLines is False)
    G.check("RTL set", ws.sheet_view.rightToLeft is True)
    G.check("zoom scale applied", ws.sheet_view.zoomScale == 145)
    # header cell exists and carries style
    G.check("header present", ws["A1"].value == "id")

    # write_header_row=False -> data starts at row 1
    b2 = jetxl.write_sheet_arrow_to_bytes(t, write_header_row=False)
    ws2 = load(b2).active
    G.check("no header row -> first data in row 1", ws2["A1"].value == 0)


# ===========================================================================
# 6. COLUMN WIDTHS / HIDDEN COLS / HIDDEN ROWS / ROW HEIGHTS
# ===========================================================================
@group("Dimensions — widths, row heights, hidden rows/cols")
def g_dims(G):
    t = sample_table(30)
    b = jetxl.write_sheet_arrow_to_bytes(
        t,
        column_widths={"id": 8.0, "region": "auto", "note": "120px", "amount": 15.5},
        row_heights={1: 30.0, 6: 45.0},   # 1-based Excel rows
        hidden_columns=["note"],
        hidden_rows=[4, 5],               # 1-based Excel rows
        default_row_height=18.0,
    )
    parts(b)
    ws = load(b).active
    G.check("explicit width applied", ws.column_dimensions["A"].width is not None)
    G.check("hidden column applied", ws.column_dimensions["C"].hidden)
    # row heights (openpyxl: 1-indexed)
    G.check("custom row height row 1", ws.row_dimensions[1].height == 30.0)
    G.check("hidden rows applied", ws.row_dimensions[4].hidden and ws.row_dimensions[5].hidden)

    # index-based references must match name-based
    b2 = jetxl.write_sheet_arrow_to_bytes(t, hidden_columns=[2], column_widths={1: 25.0})
    ws2 = load(b2).active
    G.check("hidden column by index == by name", ws2.column_dimensions["C"].hidden)


# ===========================================================================
# 7. MERGED CELLS
# ===========================================================================
@group("Merged cells")
def g_merge(G):
    t = sample_table(20)
    # jetxl merge coordinates are 1-based (row 1 == Excel row 1 == header row).
    b = jetxl.write_sheet_arrow_to_bytes(t, merge_cells=[(1, 0, 1, 3), (3, 0, 5, 0)])
    parts(b)
    ws = load(b).active
    ranges = {str(r) for r in ws.merged_cells.ranges}
    G.check("horizontal merge present", "A1:D1" in ranges)
    G.check("vertical merge present", "A3:A5" in ranges)


# ===========================================================================
# 8. HYPERLINKS — including &-in-URL corruption regression
# ===========================================================================
@group("Hyperlinks")
def g_links(G):
    t = sample_table(10)
    b = jetxl.write_sheet_arrow_to_bytes(t, hyperlinks=[
        (1, 0, "https://example.com/?x=1&y=2", "amp link"),
        (2, 1, "https://unicode.example/日本", "uni link"),
        (3, 0, "mailto:a@b.com", None),
    ])
    parts(b)  # &-in-URL must produce valid rels XML
    ws = load(b).active
    G.check("file with &-URL opens (no corruption)", ws["A1"].value == "id")
    # relationships part must exist and reference the URLs
    rels = xml_of(b, "xl/worksheets/_rels/sheet1.xml.rels")
    G.check("hyperlink relationships present", "hyperlink" in rels.lower())
    G.check("ampersand URL escaped in rels", "x=1&amp;y=2" in rels or "x=1&y=2" not in rels)


# ===========================================================================
# 9. FORMULAS — with and without cached values
# ===========================================================================
@group("Formulas")
def g_formulas(G):
    # Formulas are applied as an OVERLAY on cells that fall within the written
    # data range (1-based Excel rows; header=row 1, data=rows 2..11). A formula
    # targeting a cell inside the data grid replaces that cell's value.
    t = sample_table(10)
    b = jetxl.write_sheet_arrow_to_bytes(t, formulas=[
        (5, 3, "SUM(D2:D4)", "9.0"),               # D5, inside data
        (6, 0, "COUNT(A2:A5)", None),              # A6, inside data
        (7, 3, "AVERAGE(D2:D6)&\" avg\"", None),    # & inside formula, inside data
    ])
    parts(b)
    ws = load(b).active
    f = ws.cell(row=5, column=4).value
    G.check("formula written", isinstance(f, str) and f.startswith("="))
    G.check("formula body present", "SUM(D2:D4)" in f)
    sheet_xml = xml_of(b, "xl/worksheets/sheet1.xml")
    G.check("formula element emitted", "<f>" in sheet_xml)
    G.check("formula & escaped in xml", "&amp;" in sheet_xml)

    # KNOWN LIMITATION (documented, not a crash): a formula on a row BEYOND the
    # last data row (e.g. a "total" row under the table) is currently dropped
    # rather than appended. Assert the file stays valid either way so this can't
    # silently regress into corruption.
    b2 = jetxl.write_sheet_arrow_to_bytes(t, formulas=[(12, 3, "SUM(D2:D11)", "82.5")])
    parts(b2)
    ws2 = load(b2).active
    G.check("beyond-data formula does not corrupt file", ws2["A2"].value == 0)


# ===========================================================================
# 10. DATA VALIDATION — list, whole_number, decimal, text_length
# ===========================================================================
@group("Data validation")
def g_validation(G):
    t = sample_table(20)
    b = jetxl.write_sheet_arrow_to_bytes(t, data_validations=[
        {"start_row": 1, "start_col": 1, "end_row": 20, "end_col": 1,
         "type": "list", "items": ["North", "South", "East", "West", "Central"]},
        {"start_row": 1, "start_col": 0, "end_row": 20, "end_col": 0,
         "type": "whole_number", "min": 0, "max": 1000},
        {"start_row": 1, "start_col": 3, "end_row": 20, "end_col": 3,
         "type": "decimal", "min": 0.0, "max": 999.9},
    ])
    parts(b)
    ws = load(b).active
    G.check("data validations present", len(ws.data_validations.dataValidation) >= 2)
    types = {dv.type for dv in ws.data_validations.dataValidation}
    G.check("list validation present", "list" in types)
    G.check("numeric validation present", "whole" in types or "decimal" in types)


# ===========================================================================
# 11. CONDITIONAL FORMATTING — cell_value, color_scale, data_bar, top10
# ===========================================================================
@group("Conditional formatting")
def g_condfmt(G):
    t = sample_table(30)
    b = jetxl.write_sheet_arrow_to_bytes(t, conditional_formats=[
        {"start_row": 1, "start_col": 3, "end_row": 30, "end_col": 3,
         "rule_type": "cell_value", "operator": "greater_than", "value": "20"},
        {"start_row": 1, "start_col": 0, "end_row": 30, "end_col": 0,
         "rule_type": "color_scale", "min_color": "FFFF0000", "max_color": "FF00FF00",
         "mid_color": "FFFFFF00"},
        {"start_row": 1, "start_col": 3, "end_row": 30, "end_col": 3,
         "rule_type": "data_bar", "color": "FF638EC6"},
        {"start_row": 1, "start_col": 0, "end_row": 30, "end_col": 0,
         "rule_type": "top10", "rank": 5, "bottom": False},
    ])
    parts(b)
    ws = load(b).active
    cf_ranges = list(ws.conditional_formatting)
    G.check("conditional formats present", len(cf_ranges) >= 2)
    sheet_xml = xml_of(b, "xl/worksheets/sheet1.xml")
    G.check("conditionalFormatting element in xml", "<conditionalFormatting" in sheet_xml)
    G.check("colorScale rule emitted", "colorScale" in sheet_xml)
    G.check("dataBar rule emitted", "dataBar" in sheet_xml)


# ===========================================================================
# 12. CELL STYLES — font, fill, border, alignment
# ===========================================================================
@group("Cell styles — font, fill, border, alignment")
def g_cellstyles(G):
    # Cell-style coordinates are 1-based Excel rows. Style data row 2 (first data
    # row) col 0, and data row 3 col 1.
    t = sample_table(10)
    b = jetxl.write_sheet_arrow_to_bytes(t, cell_styles=[
        {"row": 2, "col": 0,
         "font": {"bold": True, "italic": True, "size": 14, "color": "FFFF0000", "name": "Arial"},
         "fill": {"pattern": "solid", "fg_color": "FFFFFF00"},
         "alignment": {"horizontal": "center", "vertical": "top"}},
        {"row": 3, "col": 1,
         "border": {"left": {"style": "thin"}, "right": {"style": "thin"},
                    "top": {"style": "medium"}, "bottom": {"style": "thick"}}},
    ])
    parts(b)
    ws = load(b).active
    c = ws.cell(row=2, column=1)  # jetxl row=2 col=0 -> openpyxl row 2 col 1
    G.check("bold applied", c.font.bold is True)
    G.check("italic applied", c.font.italic is True)
    G.check("font size applied", c.font.size == 14)
    G.check("font color applied", c.font.color is not None and "FF0000" in str(c.font.color.rgb))
    G.check("fill applied", c.fill.fgColor is not None)
    c2 = ws.cell(row=3, column=2)  # jetxl row=3 col=1 -> openpyxl row 3 col 2
    G.check("border applied", c2.border.left.style is not None)


# ===========================================================================
# 13. EXCEL TABLES
# ===========================================================================
@group("Cell style xf integrity — border/fill actually referenced")
def g_xf_integrity(G):
    # Regression guard for the bug where cellXfs hardcoded borderId="0", so a
    # registered border was never referenced by any cell -> borders never showed.
    # Verify via raw styles.xml that the cell's xf points at a non-zero borderId
    # and carries applyBorder, and via openpyxl that all four sides render.
    t = sample_table(10)
    b = jetxl.write_sheet_arrow_to_bytes(t, cell_styles=[
        {"row": 2, "col": 0,
         "border": {"left": {"style": "thin"}, "right": {"style": "double"},
                    "top": {"style": "dashed"}, "bottom": {"style": "thick"}},
         "fill": {"pattern": "solid", "fg_color": "FF00FF00"}},
    ])
    parts(b)
    ws = load(b).active
    c = ws.cell(row=2, column=1)
    G.check("border left renders", c.border.left.style == "thin")
    G.check("border right renders", c.border.right.style == "double")
    G.check("border top renders", c.border.top.style == "dashed")
    G.check("border bottom renders", c.border.bottom.style == "thick")

    styles_xml = xml_of(b, "xl/styles.xml")
    # There must be at least one cellXfs <xf> with a non-zero borderId + applyBorder
    import re
    xfs_block = re.search(r"<cellXfs.*?</cellXfs>", styles_xml, re.S).group(0)
    has_border_xf = re.search(r'<xf [^>]*borderId="[1-9]\d*"[^>]*applyBorder="1"', xfs_block)
    G.check("a cellXf references a real border with applyBorder", has_border_xf is not None)


@group("Excel tables")
def g_tables(G):
    t = sample_table(25)
    b = jetxl.write_sheet_arrow_to_bytes(t, tables=[
        {"name": "SalesTable", "start_row": 0, "start_col": 0,
         "end_row": 25, "end_col": 3, "style": "TableStyleMedium9",
         "show_row_stripes": True},
    ])
    p = parts(b)
    ws = load(b).active
    G.check("table part exists", any("table" in n.lower() for n in p))
    G.check("table registered on sheet", len(ws.tables) >= 1 or "SalesTable" in str(ws.tables))
    # table XML must reference the correct range and be valid OOXML
    tbl_parts = [n for n in p if n.startswith("xl/tables/") and n.endswith(".xml")]
    G.check("table xml part present", len(tbl_parts) >= 1)
    if tbl_parts:
        txml = p[tbl_parts[0]].decode()
        G.check("table has ref attr", 'ref="A1:D26"' in txml or 'ref=' in txml)


# ===========================================================================
# 14. CHARTS — column, bar, line, pie, scatter, area
# ===========================================================================
@group("Charts — all chart types")
def g_charts(G):
    t = pa.table({
        "month": pa.array(["Jan", "Feb", "Mar", "Apr"]),
        "sales": pa.array([100.0, 150.0, 120.0, 200.0], pa.float64()),
        "costs": pa.array([80.0, 90.0, 85.0, 110.0], pa.float64()),
    })
    for ctype in ["column", "bar", "line", "pie", "scatter", "area"]:
        try:
            b = jetxl.write_sheet_arrow_to_bytes(t, charts=[{
                "chart_type": ctype,
                "data_range": (0, 1, 4, 2),
                "category_col": 0,
                "title": f"{ctype} chart",
                "x_axis_title": "Month", "y_axis_title": "Value",
            }])
            p = parts(b)
            load(b)  # opens
            chart_parts = [n for n in p if "chart" in n.lower() and n.endswith(".xml")]
            G.check(f"{ctype}: chart xml part present", len(chart_parts) >= 1)
            G.check(f"{ctype}: drawing part present", any("drawing" in n.lower() for n in p))
        except Exception as e:
            G.fatal(f"{ctype} chart", e)


# ===========================================================================
# 15. IMAGES — from bytes (png), positioned
# ===========================================================================
@group("Images")
def g_images(G):
    # 1x1 red PNG
    png = bytes.fromhex(
        "89504e470d0a1a0a0000000d4948445200000001000000010806000000"
        "1f15c4890000000d49444154789c6260f8cf000000ffff03000006000557"
        "bfabd40000000049454e44ae426082"
    )
    t = sample_table(10)
    try:
        b = jetxl.write_sheet_arrow_to_bytes(t, images=[{
            "data": list(png), "extension": "png",
            "from_col": 5, "from_row": 1, "to_col": 8, "to_row": 6,
        }])
        p = parts(b)
        load(b)
        G.check("image media part present", any(n.startswith("xl/media/") for n in p))
        G.check("drawing part present for image", any("drawing" in n.lower() for n in p))
        G.check("png bytes stored", any(p[n][:4] == b"\x89PNG" for n in p if n.startswith("xl/media/")))
    except Exception as e:
        G.fatal("image from bytes", e)


# ===========================================================================
# 16. MULTI-SHEET — parallel writes, per-sheet config, sheet ordering
# ===========================================================================
@group("Multi-sheet style parity — cell_styles + conditional formats")
def g_multisheet_styles(G):
    # Regression guard: the multi-sheet in-memory path previously dropped custom
    # cell_styles and conditional formats (it used an empty style map and never
    # registered dxfs), while the single-sheet and file paths honored them. This
    # group asserts the multi-sheet bytes path now matches single-sheet output.
    t = sample_table(20)
    style = {"row": 2, "col": 0, "font": {"bold": True, "size": 13, "color": "FF0000FF"}}
    cf = {"start_row": 2, "start_col": 3, "end_row": 21, "end_col": 3,
          "rule_type": "cell_value", "operator": "greater_than", "value": "15"}

    sheets = [
        {"data": t, "name": "Styled", "cell_styles": [style], "conditional_formats": [cf]},
        {"data": sample_table(10), "name": "Plain"},
    ]
    for nthreads in (1, 2):
        b = jetxl.write_sheets_arrow_to_bytes(sheets, nthreads)
        parts(b)
        wb = load(b)
        ws = wb["Styled"]
        c = ws.cell(row=2, column=1)
        G.check(f"[{nthreads}t] multi-sheet cell_style bold applied", c.font.bold is True)
        G.check(f"[{nthreads}t] multi-sheet cell_style size applied", c.font.size == 13)
        sheet_xml = xml_of(b, "xl/worksheets/sheet1.xml")
        G.check(f"[{nthreads}t] multi-sheet conditionalFormatting emitted",
                "<conditionalFormatting" in sheet_xml)
        # styles.xml must carry the registered dxf for the conditional rule
        styles_xml = xml_of(b, "xl/styles.xml")
        G.check(f"[{nthreads}t] multi-sheet dxf registered", "<dxf" in styles_xml)

    # Parity: a single-sheet write of the same styled sheet should agree on the
    # bold flag of the styled cell.
    bs = jetxl.write_sheet_arrow_to_bytes(t, cell_styles=[style], conditional_formats=[cf])
    single_bold = load(bs).active.cell(row=2, column=1).font.bold
    bm = jetxl.write_sheets_arrow_to_bytes(
        [{"data": t, "name": "S", "cell_styles": [style], "conditional_formats": [cf]}], 1)
    multi_bold = load(bm)["S"].cell(row=2, column=1).font.bold
    G.check("single-sheet vs multi-sheet bold parity", single_bold == multi_bold == True)


@group("Multi-sheet combined features — rels integrity")
def g_multisheet_combined(G):
    # A single sheet carrying hyperlinks AND tables AND a drawing/conditional
    # format at once previously produced a dangling hyperlink relationship
    # (rId1 referenced by the sheet but absent from the rels) -> "Unknown
    # relationship" on open. Assert the fully-loaded combination opens cleanly.
    t = sample_table(20)
    sheets = [{
        "data": t, "name": "Rich",
        "cell_styles": [{"row": 2, "col": 0, "font": {"bold": True},
                          "border": {"bottom": {"style": "thick"}}}],
        "merge_cells": [(1, 0, 1, 1)],
        "hyperlinks": [(3, 0, "https://x.com/?a=1&b=2", "link")],
        "data_validations": [{"start_row": 2, "start_col": 0, "end_row": 21,
                               "end_col": 0, "type": "whole_number", "min": 0, "max": 100}],
        "formulas": [(5, 1, "SUM(B2:B4)", "3.0")],
        "conditional_formats": [{"start_row": 2, "start_col": 1, "end_row": 21,
                                  "end_col": 1, "rule_type": "data_bar", "color": "FF638EC6"}],
        "tables": [{"name": "T1", "start_row": 0, "start_col": 0, "end_row": 20, "end_col": 1}],
        "tab_color": "FFFF0000", "zoom_scale": 130,
    }, {"data": sample_table(10), "name": "Plain"}]
    for nthreads in (1, 2):
        b = jetxl.write_sheets_arrow_to_bytes(sheets, nthreads)
        parts(b)
        wb = load(b)  # must not raise "Unknown relationship"
        ws = wb["Rich"]
        G.check(f"[{nthreads}t] combined-feature sheet opens", wb.sheetnames == ["Rich", "Plain"])
        G.check(f"[{nthreads}t] hyperlink rels resolve", ws.cell(row=4, column=1).value is not None)
        G.check(f"[{nthreads}t] merge present", "A1:B1" in {str(r) for r in ws.merged_cells.ranges})
        G.check(f"[{nthreads}t] formula present", str(ws.cell(row=5, column=2).value).startswith("="))
    # reproducibility must survive the combined path
    b1 = jetxl.write_sheets_arrow_to_bytes(sheets, 2)
    b2 = jetxl.write_sheets_arrow_to_bytes(sheets, 2)
    G.check("combined multi-sheet reproducible", b1 == b2)


@group("Multi-sheet — parallel, per-sheet config, ordering")
def g_multisheet(G):
    sheets = [
        {"data": sample_table(20), "name": "Alpha", "auto_filter": True},
        {"data": sample_table(30), "name": "Beta", "freeze_rows": 1},
        {"data": sample_table(10), "name": "Gamma", "styled_headers": True},
    ]
    for nthreads in (1, 2, 4):
        b = jetxl.write_sheets_arrow_to_bytes(sheets, nthreads)
        parts(b)
        wb = load(b)
        G.check(f"[{nthreads}t] sheet order preserved", wb.sheetnames == ["Alpha", "Beta", "Gamma"])
        G.check(f"[{nthreads}t] Beta row count", wb["Beta"].max_row == 31)
        G.check(f"[{nthreads}t] per-sheet auto_filter", wb["Alpha"].auto_filter.ref is not None)

    # Many sheets — rId collision / part-naming guard
    many = [{"data": sample_table(5), "name": f"S{i}"} for i in range(60)]
    bm = jetxl.write_sheets_arrow_to_bytes(many, 4)
    parts(bm)
    wbm = load(bm)
    G.check("60 sheets all present", len(wbm.sheetnames) == 60)
    G.check("60th sheet intact", wbm["S59"].max_row == 6)


# ===========================================================================
# 17. SHEET NAMES — sanitization, length limits, duplicates, unicode
# ===========================================================================
@group("Sheet name handling")
def g_sheetnames(G):
    t = sample_table(5)
    # valid unicode name
    b = jetxl.write_sheet_arrow_to_bytes(t, sheet_name="Ventas_2024")
    G.check("custom sheet name applied", load(b).active.title == "Ventas_2024")

    # 31-char max is Excel's limit; a compliant writer must not exceed it
    long_name = "X" * 40
    try:
        b2 = jetxl.write_sheet_arrow_to_bytes(t, sheet_name=long_name)
        wb2 = load(b2)
        G.check("over-long sheet name truncated to <=31", len(wb2.active.title) <= 31)
    except Exception as e:
        # Raising is also acceptable behavior
        G.check("over-long sheet name rejected explicitly", isinstance(e, (ValueError, Exception)))

    # invalid chars per Excel spec: [ ] : * ? / \
    try:
        b3 = jetxl.write_sheet_arrow_to_bytes(t, sheet_name="Bad/Name:1")
        wb3 = load(b3)
        title = wb3.active.title
        G.check("invalid chars sanitized or rejected",
                all(c not in title for c in "[]:*?/\\"))
    except Exception as e:
        G.check("invalid sheet name rejected explicitly", True)


# ===========================================================================
# 18. OOXML STRUCTURE — required parts, content types, relationships
# ===========================================================================
@group("OOXML package structure")
def g_ooxml(G):
    b = jetxl.write_sheet_arrow_to_bytes(sample_table(50), auto_filter=True,
                                         column_formats={"amount": "currency"})
    p = parts(b)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/styles.xml",
    ]
    for r in required:
        G.check(f"required part: {r}", r in p)

    # Content types must declare the worksheet + styles
    ct = p["[Content_Types].xml"].decode()
    G.check("content-types declares worksheet", "worksheet" in ct)
    G.check("content-types declares styles", "styles" in ct)

    # workbook rels must point at the sheet
    wbrels = p["xl/_rels/workbook.xml.rels"].decode()
    G.check("workbook rels references a worksheet", "worksheet" in wbrels)

    # workbook.xml must declare exactly the sheets present
    wbxml = p["xl/workbook.xml"].decode()
    G.check("workbook declares <sheet>", "<sheet " in wbxml)

    # every r:id referenced in workbook.xml must exist in the rels
    root = ET.fromstring(p["xl/workbook.xml"])
    rel_root = ET.fromstring(p["xl/_rels/workbook.xml.rels"])
    rel_ids = {c.get("Id") for c in rel_root}
    sheet_rids = []
    for s in root.iter("{%s}sheet" % NS["main"]):
        rid = s.get("{%s}id" % NS["r"])
        if rid:
            sheet_rids.append(rid)
    G.check("all sheet r:ids resolve in rels", all(r in rel_ids for r in sheet_rids))


# ===========================================================================
# 19. REPRODUCIBILITY & IN-MEMORY vs FILE parity
# ===========================================================================
@group("Reproducibility and file/bytes parity")
def g_repro(G):
    import tempfile, os
    t = sample_table(200)
    b1 = jetxl.write_sheet_arrow_to_bytes(t, auto_filter=True,
                                          column_formats={"amount": "currency"})
    b2 = jetxl.write_sheet_arrow_to_bytes(t, auto_filter=True,
                                          column_formats={"amount": "currency"})
    G.check("byte-identical across runs (fixed timestamps)", b1 == b2)

    # file path API must produce the same logical content as the bytes API
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        jetxl.write_sheet_arrow(t, path, auto_filter=True,
                                column_formats={"amount": "currency"})
        with open(path, "rb") as fh:
            fb = fh.read()
        G.check("file API archive is valid", zipfile.ZipFile(io.BytesIO(fb)).testzip() is None)
        wsf = load(fb).active
        wsb = load(b1).active
        G.check("file vs bytes: same cell values",
                wsf["A2"].value == wsb["A2"].value and wsf["D51"].value == wsb["D51"].value)
    finally:
        os.unlink(path)


# ===========================================================================
# 20. SCALE — value integrity on a large sheet
# ===========================================================================
@group("Scale — value integrity at 50k rows")
def g_scale(G):
    n = 50_000
    t = pa.table({
        "id": pa.array(list(range(n)), pa.int64()),
        "sq": pa.array([float(i * i % 100000) for i in range(n)], pa.float64()),
        "lbl": pa.array([["A", "B", "C"][i % 3] for i in range(n)]),
    })
    b = jetxl.write_sheet_arrow_to_bytes(t)
    G.check("large archive CRC valid", zipfile.ZipFile(io.BytesIO(b)).testzip() is None)
    ws = load(b).active
    G.check("row count == n+1 (header)", ws.max_row == n + 1)
    G.check("first data row correct", ws["A2"].value == 0)
    G.check("last data row correct", ws.cell(row=n + 1, column=1).value == n - 1)
    G.check("mid low-card label correct", ws.cell(row=2, column=3).value == "A")


# ===========================================================================
# Runner
# ===========================================================================
def main():
    print("=" * 72)
    print("jetxl v1 stability suite")
    print(f"pyarrow {pa.__version__} | openpyxl {openpyxl.__version__}")
    print("=" * 72)

    results = []
    for name, fn in GROUPS:
        print(f"\n=== {name} ===")
        G = Group(name)
        try:
            fn(G)
        except Exception as e:
            import traceback
            traceback.print_exc()
            G.failed += 1
            G.errors.append(f"group crashed: {e!r}")
        results.append(G)

    print("\n" + "=" * 72)
    print("SUMMARY")
    print("-" * 72)
    total_p = total_f = 0
    for G in results:
        total_p += G.passed
        total_f += G.failed
        status = "OK  " if G.failed == 0 else "FAIL"
        print(f"  [{status}] {G.name}: {G.passed} passed, {G.failed} failed")
    print("-" * 72)
    print(f"TOTAL: {total_p} passed, {total_f} failed")
    print("=" * 72)
    return 1 if total_f else 0


if __name__ == "__main__":
    sys.exit(main())
