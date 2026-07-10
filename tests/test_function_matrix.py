#!/usr/bin/env python3
"""
jetxl full-function matrix suite
================================

Addresses two gaps in earlier coverage:

  1. Earlier suites leaned on `write_sheet_arrow_to_bytes` (the least-used
     entry point). This suite drives EVERY public function directly — with the
     file-based ones (`write_sheet_arrow`, `write_sheets_arrow`) and the legacy
     dict API (`write_sheet`, `write_sheets`) treated as first-class, not
     afterthoughts.

  2. Every arrow function is fed from real **Polars** and **Pandas**
     DataFrames (via `.to_arrow()` / `pa.Table.from_pandas`) as well as native
     PyArrow, because that's how users actually call jetxl.

Edge cases are drawn from what openpyxl / XlsxWriter / rust_xlsxwriter test:
NaN & Infinity, numbers-as-text (leading zeros), the Excel-1900 phantom leap
day, unicode & XML escapes, type coercion (bool vs int), empty / single-cell /
very-wide / very-tall data, and quote/space handling.

Every write is round-tripped through openpyxl and CRC-checked. Run with -v to
see each passing assertion.
"""
from __future__ import annotations

import datetime as dt
import io
import math
import os
import sys
import tempfile
import zipfile

import pyarrow as pa
import polars as pl
import pandas as pd
import openpyxl

import jetxl

VERBOSE = "-v" in sys.argv or "--verbose" in sys.argv

_p = _f = 0
_fails = []


def ok(d):
    global _p
    _p += 1
    if VERBOSE:
        print(f"    pass  {d}")


def bad(d):
    global _f
    _f += 1
    _fails.append(d)
    print(f"    FAIL  {d}")


def check(d, c):
    ok(d) if c else bad(d)


def sec(n):
    print(f"\n=== {n} ===")


# --------------------------------------------------------------------------
# read-back helpers: open the result of every write path and CRC-check it
# --------------------------------------------------------------------------
def rb_file_single(arrow, **kw):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        jetxl.write_sheet_arrow(arrow, path, **kw)
        with open(path, "rb") as fh:
            data = fh.read()
    finally:
        os.unlink(path)
    _crc(data)
    return openpyxl.load_workbook(io.BytesIO(data))


def rb_bytes_single(arrow, **kw):
    data = jetxl.write_sheet_arrow_to_bytes(arrow, **kw)
    _crc(data)
    return openpyxl.load_workbook(io.BytesIO(data))


def rb_file_multi(sheets, threads=2):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        jetxl.write_sheets_arrow(sheets, path, threads)
        with open(path, "rb") as fh:
            data = fh.read()
    finally:
        os.unlink(path)
    _crc(data)
    return openpyxl.load_workbook(io.BytesIO(data))


def rb_bytes_multi(sheets, threads=2):
    data = jetxl.write_sheets_arrow_to_bytes(sheets, threads)
    _crc(data)
    return openpyxl.load_workbook(io.BytesIO(data))


def rb_dict_single(cols, **kw):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        jetxl.write_sheet(cols, path, **kw)
        with open(path, "rb") as fh:
            data = fh.read()
    finally:
        os.unlink(path)
    _crc(data)
    return openpyxl.load_workbook(io.BytesIO(data))


def rb_dict_multi(sheets, threads=2):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        jetxl.write_sheets(sheets, path, threads)
        with open(path, "rb") as fh:
            data = fh.read()
    finally:
        os.unlink(path)
    _crc(data)
    return openpyxl.load_workbook(io.BytesIO(data))


def _crc(data):
    if zipfile.ZipFile(io.BytesIO(data)).testzip() is not None:
        raise AssertionError("bad CRC in archive")


# Every arrow single-sheet write path, applied to one arrow table + kwargs.
def each_single_arrow_path(arrow, **kw):
    yield "file", rb_file_single(arrow, **kw)
    yield "bytes", rb_bytes_single(arrow, **kw)
    sheet = {"data": arrow, "name": "S1"}
    sheet.update(kw)
    yield "multi_file", rb_file_multi([dict(sheet)])
    yield "multi_bytes", rb_bytes_multi([dict(sheet)])


# --------------------------------------------------------------------------
def base_frames():
    """Same logical data as polars, pandas, pyarrow."""
    data = {
        "id": [1, 2, 3, 4],
        "name": ["Alice", "Bob", "Carol", "Dan"],
        "score": [9.5, 8.25, 7.0, 10.0],
        "active": [True, False, True, False],
    }
    pldf = pl.DataFrame(data)
    pddf = pd.DataFrame(data)
    patbl = pa.table(data)
    return pldf, pddf, patbl


# ==========================================================================
# 1. Every function, every DataFrame source — the core matrix
# ==========================================================================
def t_function_x_dataframe_matrix():
    sec("Function × DataFrame matrix — Polars / Pandas / PyArrow through every path")
    pldf, pddf, patbl = base_frames()
    sources = {
        "polars": pldf.to_arrow(),
        "pandas": pa.Table.from_pandas(pddf),
        "pyarrow": patbl,
    }
    for src_name, arrow in sources.items():
        for path_name, wb in each_single_arrow_path(arrow):
            ws = wb.active
            good = (ws["A1"].value == "id" and ws["B2"].value == "Alice"
                    and abs(ws["C3"].value - 8.25) < 1e-9
                    and ws["D2"].value is True and ws["D3"].value is False)
            check(f"[{src_name} → {path_name}] values + types round-trip", good)

    # multi-sheet with mixed sources in one workbook, both multi paths
    sheets = [
        {"data": sources["polars"], "name": "FromPolars"},
        {"data": sources["pandas"], "name": "FromPandas"},
        {"data": sources["pyarrow"], "name": "FromArrow"},
    ]
    for label, wb in [("multi_file", rb_file_multi(sheets)), ("multi_bytes", rb_bytes_multi(sheets))]:
        check(f"[{label}] three sources, sheet order preserved",
              wb.sheetnames == ["FromPolars", "FromPandas", "FromArrow"])
        check(f"[{label}] FromPandas sheet intact", wb["FromPandas"]["B3"].value == "Bob")


# ==========================================================================
# 2. Legacy dict API — write_sheet / write_sheets (first-class coverage)
# ==========================================================================
def t_dict_api_full():
    sec("Legacy dict API — write_sheet / write_sheets, type fidelity")
    cols = {
        "flag": [True, False, True],
        "count": [10, 20, 30],
        "price": [1.5, 2.25, 3.75],
        "label": ["a", "b", "c"],
        "when": [dt.datetime(2024, 1, 1, 12, 0, 0)] * 3,
    }
    wb = rb_dict_single(cols, sheet_name="Types")
    ws = wb["Types"]
    check("dict: bool True stays bool", ws["A2"].value is True)
    check("dict: bool False stays bool", ws["A3"].value is False)
    check("dict: int fidelity", ws["B2"].value == 10)
    check("dict: float fidelity", abs(ws["C3"].value - 2.25) < 1e-9)
    check("dict: string fidelity", ws["D2"].value == "a")
    check("dict: datetime fidelity", isinstance(ws["E2"].value, dt.datetime))

    # multi-sheet dict with thread counts
    for threads in (1, 2, 3):
        sheets = [
            {"name": "Q1", "columns": {"x": [1, 2, 3]}},
            {"name": "Q2", "columns": {"y": [4, 5, 6]}},
            {"name": "Q3", "columns": {"z": [7, 8, 9]}},
        ]
        wb = rb_dict_multi(sheets, threads)
        check(f"dict multi [{threads}t] order", wb.sheetnames == ["Q1", "Q2", "Q3"])
        check(f"dict multi [{threads}t] Q3 value", wb["Q3"]["A2"].value == 7)


# ==========================================================================
# 3. NaN / Infinity — must not corrupt (references test this explicitly)
# ==========================================================================
def t_nan_inf():
    sec("NaN / Infinity floats — no corruption, empty cells")
    t = pa.table({"v": pa.array([1.0, float("nan"), float("inf"), float("-inf"), 2.0], pa.float64())})
    for path_name, wb in each_single_arrow_path(t):
        ws = wb.active
        # 1.0 and 2.0 preserved; NaN/±Inf must NOT become literal text
        good = (ws["A2"].value == 1.0 and ws["A6"].value == 2.0
                and ws["A3"].value in (None, "") and ws["A4"].value in (None, ""))
        check(f"[{path_name}] NaN/Inf → empty, finite values intact", good)
    # raw-xml guard: no 'NaN'/'inf' literal leaked into a numeric cell
    data = jetxl.write_sheet_arrow_to_bytes(t)
    sx = zipfile.ZipFile(io.BytesIO(data)).read("xl/worksheets/sheet1.xml").decode()
    check("no NaN/inf literal in sheet xml", "NaN" not in sx and "inf" not in sx.lower())


# ==========================================================================
# 4. Numbers-as-text — leading zeros / plus signs stay strings
# ==========================================================================
def t_numbers_as_text():
    sec("Numbers-as-text — leading zeros preserved (openpyxl/xlsxwriter case)")
    # via all three frame sources to be sure the string type survives conversion
    data = {"code": ["007", "0042", "+5", "1e3", "00"]}
    for src, arrow in [("polars", pl.DataFrame(data).to_arrow()),
                       ("pandas", pa.Table.from_pandas(pd.DataFrame(data))),
                       ("pyarrow", pa.table(data))]:
        ws = rb_bytes_single(arrow).active
        vals = [ws.cell(row=r, column=1).value for r in range(2, 7)]
        check(f"[{src}] leading-zero codes stay strings",
              vals == ["007", "0042", "+5", "1e3", "00"])


# ==========================================================================
# 5. Excel-1900 date system incl. phantom leap day (all paths)
# ==========================================================================
def t_date_serials():
    sec("Excel 1900 date serials — phantom leap-day boundary, all paths")
    dates = [dt.date(1900, 1, 1), dt.date(1900, 2, 28), dt.date(1900, 3, 1),
             dt.date(2000, 1, 1), dt.date(2024, 6, 15), dt.date(9999, 12, 31)]
    expected = [1, 59, 61, 36526, 45458, 2958465]
    t = pa.table({"d": pa.array(dates)})
    for path_name in ["bytes", "file", "multi_bytes", "multi_file"]:
        if path_name == "bytes":
            data = jetxl.write_sheet_arrow_to_bytes(t, column_formats={"d": "date"})
        elif path_name == "file":
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
                p = tf.name
            jetxl.write_sheet_arrow(t, p, column_formats={"d": "date"})
            data = open(p, "rb").read(); os.unlink(p)
        elif path_name == "multi_bytes":
            data = jetxl.write_sheets_arrow_to_bytes([{"data": t, "name": "S1", "column_formats": {"d": "date"}}], 1)
        else:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
                p = tf.name
            jetxl.write_sheets_arrow([{"data": t, "name": "S1", "column_formats": {"d": "date"}}], p, 1)
            data = open(p, "rb").read(); os.unlink(p)
        _crc(data)
        import re
        sx = zipfile.ZipFile(io.BytesIO(data)).read("xl/worksheets/sheet1.xml").decode()
        serials = [round(float(s)) for s in re.findall(r'<c r="A[2-9]"[^>]*><v>([^<]+)</v>', sx)]
        check(f"[{path_name}] all six date serials match Excel", serials == expected)

    # datetime with time component keeps the fractional part
    tt = pa.table({"ts": pa.array([dt.datetime(2024, 6, 15, 14, 30, 0)])})
    ws = rb_bytes_single(tt).active
    check("datetime with time round-trips", ws["A2"].value == dt.datetime(2024, 6, 15, 14, 30, 0))


# ==========================================================================
# 6. Unicode & XML escapes — through all frame sources and paths
# ==========================================================================
def t_unicode_escapes():
    sec("Unicode & XML escapes — hostile strings via Polars/Pandas/PyArrow")
    data = {"s": ["<tag>", "a & b", '"q"', "café ☕ 日本 🚀", "line1\nline2"]}
    for src, arrow in [("polars", pl.DataFrame(data).to_arrow()),
                       ("pandas", pa.Table.from_pandas(pd.DataFrame(data))),
                       ("pyarrow", pa.table(data))]:
        for path_name, wb in each_single_arrow_path(arrow):
            ws = wb.active
            good = (ws["A2"].value == "<tag>" and ws["A3"].value == "a & b"
                    and ws["A4"].value == '"q"'
                    and ws["A5"].value == "café ☕ 日本 🚀"
                    and ws["A6"].value == "line1\nline2")
            check(f"[{src} → {path_name}] escapes + unicode intact", good)


# ==========================================================================
# 7. Shape edge cases — empty-ish, single cell, very wide, very tall
# ==========================================================================
def t_shapes():
    sec("Shape edge cases — single cell, very wide, very tall, header-only")
    # single cell
    ws = rb_bytes_single(pa.table({"a": pa.array([42], pa.int64())})).active
    check("single-cell value", ws["A2"].value == 42)

    # very wide (200 columns)
    wide = pa.table({f"c{i}": pa.array([i], pa.int64()) for i in range(200)})
    for path_name, wb in each_single_arrow_path(wide):
        check(f"[{path_name}] 200-column wide sheet last col",
              wb.active.cell(row=2, column=200).value == 199)

    # very tall handled elsewhere at 50k; here a 1-row-per-col mix
    # header-only via write_header_row on zero-ish data: a 1-row table
    onerow = pa.table({"x": pa.array([7], pa.int64()), "y": pa.array(["z"])})
    ws = rb_file_single(onerow, styled_headers=True).active
    check("1-row table header present", ws["A1"].value == "x")
    check("1-row table data present", ws["A2"].value == 7)


# ==========================================================================
# 8. Column ref by index vs name — parity across file & bytes
# ==========================================================================
def t_column_refs():
    sec("Column references — int index vs str name parity (file & bytes)")
    t = pa.table({"id": pa.array([1, 2]), "region": pa.array(["N", "S"]),
                  "amt": pa.array([1.0, 2.0], pa.float64())})
    # by name
    for path_name, wb in each_single_arrow_path(t, hidden_columns=["region"],
                                                column_widths={"amt": 20.0}):
        check(f"[{path_name}] hidden by NAME", wb.active.column_dimensions["B"].hidden)
    # by index (region == 1)
    for path_name, wb in each_single_arrow_path(t, hidden_columns=[1],
                                                column_widths={2: 20.0}):
        check(f"[{path_name}] hidden by INDEX", wb.active.column_dimensions["B"].hidden)


# ==========================================================================
# 9. Error handling parity — bad refs raise on every arrow path
# ==========================================================================
def t_column_name_casing():
    sec("Column references are exact-match; wrong name hard-errors with column list")
    t = pa.table({"Apple": pa.array([1.5, 2.5], pa.float64()),
                  "Banana": pa.array([10, 20]),
                  "myColumn": pa.array(["a", "b"])})
    # exact match works and preserves data casing in the header
    b = rb_bytes_single(t, column_formats={"Apple": "currency"})
    ws = b.active
    check("[column_formats 'Apple'] exact match applies",
          ws.cell(row=2, column=1).number_format == "$#,##0.00")
    check("output header preserves data casing",
          [ws.cell(row=1, column=c).value for c in range(1, 4)] == ["Apple", "Banana", "myColumn"])
    # any mismatch (wrong case OR wrong name) must HARD ERROR, not resolve
    for bad in ["APPLE", "apple", "Cherry", "apple "]:
        raised = False
        msg = ""
        try:
            rb_bytes_single(t, column_formats={bad: "currency"})
        except Exception as e:
            raised = True
            msg = str(e)
        check(f"[column_formats {bad!r}] hard-errors (no silent resolve)", raised)
        # error prints the available columns as a Python list literal
        check(f"[{bad!r}] error lists columns as Python list",
              '["Apple", "Banana", "myColumn"]' in msg)
    # same behavior for column_widths and hidden_columns
    for param, val in [("column_widths", {"cherry": 20.0}), ("hidden_columns", ["CHERRY"])]:
        raised = False
        msg = ""
        try:
            rb_bytes_single(t, **{param: val})
        except Exception as e:
            raised = True
            msg = str(e)
        check(f"[{param}] wrong name hard-errors", raised)
        check(f"[{param}] error lists columns as Python list",
              '["Apple", "Banana", "myColumn"]' in msg)


def t_column_name_exact_case_sensitive():
    sec("Columns differing only by case are matched exactly (no collision)")
    t = pa.table({"col": pa.array([1.0, 2.0], pa.float64()),
                  "Col": pa.array([3.0, 4.0], pa.float64()),
                  "COL": pa.array([5.0, 6.0], pa.float64())})
    for name, idx in [("col", 1), ("Col", 2), ("COL", 3)]:
        b = rb_bytes_single(t, column_formats={name: "currency"})
        fmts = [b.active.cell(row=2, column=c).number_format for c in range(1, 4)]
        expected = ["$#,##0.00" if c == idx else "General" for c in range(1, 4)]
        check(f"[exact {name!r}] targets only column {idx}", fmts == expected)
    # a name matching none exactly errors even though it case-folds to several
    raised = False
    try:
        rb_bytes_single(t, column_formats={"CoL": "currency"})
    except Exception:
        raised = True
    check("['CoL'] with no exact match hard-errors", raised)


def t_direct_dataframe_pycapsule():
    sec("Direct DataFrame via Arrow PyCapsule — no .to_arrow() needed")
    data = {
        "id": [1, 2, 3, 4],
        "name": ["Alice", "Bob", "Carol", "Dan"],
        "score": [9.5, 8.25, 7.0, 10.0],
        "active": [True, False, True, False],
    }
    # Polars exports strings as Utf8View via the capsule; pandas exports Utf8.
    # Both must work when passed DIRECTLY (the whole point of this feature).
    pldf = pl.DataFrame(data)
    pddf = pd.DataFrame(data)
    for label, frame in [("polars-direct", pldf), ("pandas-direct", pddf)]:
        for path_name, wb in each_single_arrow_path(frame):
            ws = wb.active
            names = [ws.cell(row=r, column=2).value for r in range(2, 6)]
            check(f"[{label}/{path_name}] names round-trip",
                  names == ["Alice", "Bob", "Carol", "Dan"])
            scores = [ws.cell(row=r, column=3).value for r in range(2, 6)]
            check(f"[{label}/{path_name}] scores round-trip",
                  scores == [9.5, 8.25, 7.0, 10.0])
    # Utf8View / BinaryView explicit coverage (Polars' native string repr)
    import pyarrow as _pa
    sv = _pa.table({"s": _pa.array(["a", "bb", "ccc"], _pa.string_view())})
    for path_name, wb in each_single_arrow_path(sv):
        ws = wb.active
        check(f"[utf8view/{path_name}] string_view resolves",
              [ws.cell(row=r, column=1).value for r in range(2, 5)] == ["a", "bb", "ccc"])
    # backward-compat: .to_arrow() must still work identically
    for path_name, wb in each_single_arrow_path(pldf.to_arrow()):
        ws = wb.active
        check(f"[polars.to_arrow/{path_name}] still works",
              [ws.cell(row=r, column=2).value for r in range(2, 6)] == ["Alice", "Bob", "Carol", "Dan"])


def t_categorical_columns():
    sec("Categorical / dictionary columns — pandas & polars, all paths")
    import pandas as _pd
    import polars as _pl
    # pandas Categorical
    pdf = _pd.DataFrame({"cat": _pd.Categorical(["apple", "banana", "apple", "cherry"]),
                         "val": [1, 2, 3, 4]})
    arrow = pa.Table.from_pandas(pdf)
    for path_name, wb in each_single_arrow_path(arrow):
        ws = wb.active
        check(f"[pandas-cat/{path_name}] categorical resolves to values",
              [ws.cell(row=r, column=1).value for r in range(2, 6)] == ["apple", "banana", "apple", "cherry"])
    # polars Categorical
    pldf = _pl.DataFrame({"cat": _pl.Series(["x", "y", "x"], dtype=_pl.Categorical), "val": [1, 2, 3]})
    for path_name, wb in each_single_arrow_path(pldf.to_arrow()):
        ws = wb.active
        check(f"[polars-cat/{path_name}] categorical resolves to values",
              [ws.cell(row=r, column=1).value for r in range(2, 5)] == ["x", "y", "x"])
    # integer-keyed dictionary
    dint = pa.array([10, 20, 10]).dictionary_encode()
    for path_name, wb in each_single_arrow_path(pa.table({"d": dint})):
        ws = wb.active
        check(f"[int-dict/{path_name}] int dictionary resolves",
              [ws.cell(row=r, column=1).value for r in range(2, 5)] == [10, 20, 10])


def t_error_parity():
    sec("Error handling — invalid column ref raises on file AND bytes")
    t = pa.table({"id": pa.array([1, 2])})

    def expect_raise(fn, exc, label):
        try:
            fn()
            bad(f"{label}: expected {exc.__name__}, no error")
        except exc:
            ok(f"{label}: raised {exc.__name__}")
        except Exception as e:
            bad(f"{label}: expected {exc.__name__}, got {type(e).__name__}")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = tf.name
    try:
        expect_raise(lambda: jetxl.write_sheet_arrow(t, p, column_widths={"nope": 10.0}),
                     ValueError, "file unknown-name")
        expect_raise(lambda: jetxl.write_sheet_arrow_to_bytes(t, column_widths={"nope": 10.0}),
                     ValueError, "bytes unknown-name")
        expect_raise(lambda: jetxl.write_sheet_arrow(t, p, hidden_columns=[99]),
                     IndexError, "file oob-index")
        expect_raise(lambda: jetxl.write_sheet_arrow_to_bytes(t, hidden_columns=[99]),
                     IndexError, "bytes oob-index")
    finally:
        os.unlink(p)


def main():
    print("=" * 74)
    print("jetxl full-function matrix suite")
    print(f"polars {pl.__version__} | pandas {pd.__version__} | pyarrow {pa.__version__}")
    print("=" * 74)
    for fn in [
        t_function_x_dataframe_matrix, t_dict_api_full, t_nan_inf,
        t_numbers_as_text, t_date_serials, t_unicode_escapes,
        t_shapes, t_column_refs, t_column_name_casing, t_column_name_exact_case_sensitive,
        t_direct_dataframe_pycapsule, t_categorical_columns, t_error_parity,
    ]:
        try:
            fn()
        except Exception as e:
            import traceback
            traceback.print_exc()
            bad(f"{fn.__name__} crashed: {e!r}")
    print("\n" + "=" * 74)
    print(f"TOTAL: {_p} passed, {_f} failed")
    if _fails:
        print("\nFailures:")
        for x in _fails:
            print("  -", x)
    print("=" * 74)
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())
